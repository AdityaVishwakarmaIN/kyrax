"""W2 styles round-trip: turbo write -> openpyxl load (silo B oracle checks)."""

from __future__ import annotations

import re
import warnings
import zipfile
from datetime import date, datetime
from pathlib import Path

import pytest

from kyrax import write_excel_turbo, write_excel_turbo_bytes

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock


def _load(path: Path, **kwargs):
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        return load_workbook(path, **kwargs)


def _rgb_of(color) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    if color.type == "theme":
        return f"theme:{color.theme}"
    return str(color.value) if color.value is not None else None


def _cell_xfs_count(path: Path) -> int:
    with zipfile.ZipFile(path) as z:
        styles = z.read("xl/styles.xml").decode()
    m = re.search(r'<cellXfs count="(\d+)"', styles)
    assert m, "cellXfs count missing"
    return int(m.group(1))


def test_smoke_styles(tmp_path: Path):
    path = tmp_path / "smoke.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Sheet",
                "rows": [[42, 100, "hi", 0.5], ["plain", None, None, None]],
                "cell_styles": {
                    (0, 0): {
                        "font": {
                            "name": "Arial",
                            "sz": 14,
                            "bold": True,
                            "color": "FF0000",
                        },
                        "fill": {"patternType": "solid", "fg": "FFFF00"},
                        "border": {"style": "thin", "color": "FF0000"},
                        "alignment": {
                            "horizontal": "center",
                            "vertical": "center",
                            "wrap_text": True,
                        },
                        "num_fmt": "0.00",
                        "protection": {"locked": False, "hidden": True},
                    },
                    (0, 1): {"num_fmt": '"USD"#,##0.00'},
                    (0, 2): {"named_style": "Highlight"},
                    (0, 3): {"num_fmt": "0%"},
                },
                "named_styles": [
                    {
                        "name": "Highlight",
                        "font": {"bold": True, "color": "0000FF"},
                        "fill": {"patternType": "solid", "fg": "CCFFCC"},
                    }
                ],
            }
        ],
    )
    wb = _load(path)
    ws = wb.active
    c = ws["A1"]
    assert c.value == 42
    assert c.font.name == "Arial"
    assert c.font.bold is True
    assert c.font.sz == 14.0
    assert _rgb_of(c.font.color) == "00FF0000"
    assert c.fill.patternType == "solid"
    assert _rgb_of(c.fill.fgColor) == "00FFFF00"
    assert c.border.left.style == "thin"
    assert _rgb_of(c.border.left.color) == "00FF0000"
    assert c.alignment.horizontal == "center"
    assert c.alignment.wrap_text is True
    assert c.number_format == "0.00"
    assert c.protection.locked is False
    assert c.protection.hidden is True

    assert ws["B1"].number_format == '"USD"#,##0.00'
    assert ws["D1"].number_format == "0%"
    assert ws["C1"].style == "Highlight"
    assert "Highlight" in wb.style_names
    assert "Normal" in wb.style_names
    assert ws["A2"].has_style is False

    with zipfile.ZipFile(path) as z:
        styles = z.read("xl/styles.xml").decode()
    assert "numFmtId=\"164\"" in styles or 'numFmtId="164"' in styles
    assert "USD" in styles or "&quot;USD&quot;" in styles
    assert "gray125" in styles


def test_cf_families(tmp_path: Path):
    path = tmp_path / "cf.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "CF",
                "rows": [[5, 10, 15, 20, 25]],
                "conditional_formatting": [
                    {
                        "sqref": "A1:E1",
                        "rules": [
                            {
                                "type": "colorScale",
                                "priority": 1,
                                "cfvos": [
                                    {"type": "min"},
                                    {"type": "percentile", "val": "50"},
                                    {"type": "max"},
                                ],
                                "colors": ["FF0000", "FFFF00", "00FF00"],
                            },
                            {
                                "type": "dataBar",
                                "priority": 2,
                                "cfvos": [{"type": "min"}, {"type": "max"}],
                                "color": "638EC6",
                            },
                            {
                                "type": "iconSet",
                                "priority": 3,
                                "iconSet": "3TrafficLights1",
                                "cfvos": [
                                    {"type": "percent", "val": "0"},
                                    {"type": "percent", "val": "33"},
                                    {"type": "percent", "val": "67"},
                                ],
                            },
                            {
                                "type": "cellIs",
                                "priority": 4,
                                "operator": "greaterThan",
                                "formulas": ["10"],
                                "dxf": {
                                    "font": {"color": "9C0006"},
                                    "fill": {
                                        "patternType": "solid",
                                        "fg": "FFC7CE",
                                    },
                                },
                            },
                        ],
                    }
                ],
            }
        ],
    )
    wb = _load(path)
    ws = wb.active
    rules = list(ws.conditional_formatting)
    assert rules
    types = []
    for cf in rules:
        for r in cf.rules:
            types.append(r.type)
    for need in ("colorScale", "dataBar", "iconSet", "cellIs"):
        assert need in types, f"missing {need} in {types}"
    cell_is = None
    for cf in rules:
        for r in cf.rules:
            if r.type == "cellIs":
                cell_is = r
    assert cell_is is not None
    assert cell_is.dxf is not None
    assert cell_is.dxf.fill is not None
    assert cell_is.operator == "greaterThan"


def test_data_validations(tmp_path: Path):
    path = tmp_path / "dv.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "DV",
                "rows": [["A", None, None, None]],
                "data_validations": [
                    {
                        "type": "list",
                        "formula1": '"A,B,C"',
                        "sqref": "A1",
                        "allow_blank": True,
                    },
                    {
                        "type": "decimal",
                        "operator": "between",
                        "formula1": "0",
                        "formula2": "100",
                        "sqref": "B1",
                    },
                    {
                        "type": "date",
                        "operator": "greaterThan",
                        "formula1": "44927",
                        "sqref": "C1",
                    },
                    {
                        "type": "custom",
                        "formula1": "ISNUMBER(D1)",
                        "sqref": "D1",
                    },
                ],
            }
        ],
    )
    wb = _load(path)
    ws = wb.active
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) == 4
    by_type = {d.type: d for d in dvs}
    assert by_type["list"].formula1 == '"A,B,C"'
    assert by_type["decimal"].operator == "between"
    assert by_type["decimal"].formula1 == "0"
    assert by_type["decimal"].formula2 == "100"
    assert by_type["date"].operator == "greaterThan"
    assert by_type["custom"].formula1 == "ISNUMBER(D1)"


def test_rich_text(tmp_path: Path):
    path = tmp_path / "rich.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Rich",
                "rows": [
                    [
                        {
                            "rich": [
                                {
                                    "text": "Bold",
                                    "font": {
                                        "bold": True,
                                        "sz": 12,
                                        "color": "FF0000",
                                    },
                                },
                                " plain ",
                                {
                                    "text": "italic",
                                    "font": {
                                        "italic": True,
                                        "color": "0000FF",
                                    },
                                },
                            ]
                        }
                    ],
                    [
                        {
                            "rich": [
                                {
                                    "text": "code",
                                    "font": {
                                        "name": "Consolas",
                                        "sz": 14,
                                        "bold": True,
                                        "color": "008000",
                                    },
                                }
                            ]
                        }
                    ],
                ],
            }
        ],
    )
    wb = _load(path, rich_text=True)
    ws = wb.active
    v = ws["A1"].value
    assert isinstance(v, CellRichText)
    texts = [str(x) for x in v]
    assert "Bold" in "".join(texts)
    assert "italic" in "".join(texts)
    bold_runs = [x for x in v if isinstance(x, TextBlock) and x.font.b]
    assert bold_runs
    assert _rgb_of(bold_runs[0].font.color) == "00FF0000"
    italic_runs = [x for x in v if isinstance(x, TextBlock) and x.font.i]
    assert italic_runs
    assert _rgb_of(italic_runs[0].font.color) == "000000FF"

    v2 = ws["A2"].value
    assert isinstance(v2, CellRichText)
    block = next(x for x in v2 if isinstance(x, TextBlock))
    assert block.font.rFont == "Consolas"


def test_row_col_styles(tmp_path: Path):
    path = tmp_path / "rowcol.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "RowCol",
                "rows": [["row+col style", 1]],
                "row_dims": [
                    {
                        "row": 1,
                        "height": 20.0,
                        "style": {
                            "font": {
                                "name": "Arial",
                                "sz": 12,
                                "bold": True,
                                "color": "0000FF",
                            }
                        },
                    }
                ],
                "col_dims": [
                    {
                        "min": 1,
                        "width": 18.0,
                        "style": {
                            "fill": {"patternType": "solid", "fg": "FFFF99"}
                        },
                    }
                ],
            }
        ],
    )
    wb = _load(path)
    ws = wb.active
    rd = ws.row_dimensions[1]
    assert rd.customFormat or (rd.style is not None)
    cd = ws.column_dimensions["A"]
    assert cd.width is not None and cd.width > 10


def test_dense_styles_collapse(tmp_path: Path):
    n_cells = 100_000
    n_styles = 200
    palette = []
    for i in range(n_styles):
        fg = f"{(i * 37) % 256:02X}{(i * 59) % 256:02X}{(i * 97) % 256:02X}"
        palette.append(
            {
                "font": {
                    "name": ["Calibri", "Arial", "Consolas", "Georgia", "Tahoma"][
                        i % 5
                    ],
                    "sz": [9, 10, 11, 12, 14, 16][i % 6],
                    "bold": i % 3 == 0,
                    "color": fg,
                },
                "fill": {"patternType": "solid", "fg": fg},
                "num_fmt": (
                    f'0.000"u{i % 17}"'
                    if i % 7 == 0
                    else ["General", "0.00", "0%", "#,##0", "0.00E+00", "mm-dd-yy"][
                        i % 6
                    ]
                ),
            }
        )
    path = tmp_path / "dense.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Dense",
                "columns": [list(range(n_cells))],
                "style_palette": palette,
            }
        ],
    )
    n = _cell_xfs_count(path)
    assert n <= 250, f"dense cellXfs={n} expected <=250"
    # silo B: 201
    assert n == 201 or n <= 220


def test_unique_styles_linear(tmp_path: Path):
    n = 10_000
    path = tmp_path / "unique.xlsx"
    # Build via style_palette of unique styles + one cell each
    palette = []
    for i in range(n):
        r = i & 0xFF
        g = (i >> 8) & 0xFF
        b = (i * 3) & 0xFF
        fg = f"{r:02X}{g:02X}{b:02X}"
        palette.append(
            {
                "fill": {"patternType": "solid", "fg": fg},
                "font": {
                    "name": "Arial",
                    "sz": 10.0 + (i % 5),
                    "bold": i % 2 == 0,
                    "color": fg,
                },
                "num_fmt": f'0.0"x{i}"',
            }
        )
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Unique",
                "columns": [list(range(n))],
                "style_palette": palette,
            }
        ],
    )
    n_xfs = _cell_xfs_count(path)
    assert n_xfs >= 9000, f"unique cellXfs={n_xfs} expected >=9000"
    assert n_xfs == n + 1  # xf0 + n unique


def test_dates_display_as_dates(tmp_path: Path):
    path = tmp_path / "dates.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "D",
                "columns": [
                    [date(2020, 1, 15), datetime(2020, 1, 15, 12, 30, 0)],
                ],
            }
        ],
    )
    wb = _load(path)
    ws = wb.active
    # openpyxl should treat cells as dates (is_date) after numFmt xf
    c1 = ws["A1"]
    c2 = ws["A2"]
    assert c1.is_date is True or isinstance(c1.value, (date, datetime))
    assert c2.is_date is True or isinstance(c2.value, (date, datetime))
    # values should round-trip as date-like
    v1 = c1.value
    if isinstance(v1, datetime):
        assert v1.date() == date(2020, 1, 15)
    else:
        assert v1 == date(2020, 1, 15) or float(v1) == pytest.approx(43845.0)


def test_named_styles_by_name(tmp_path: Path):
    path = tmp_path / "named.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "S",
                "rows": [["x"]],
                "cell_styles": {(0, 0): {"named_style": "Highlight"}},
            }
        ],
        named_styles=[
            {
                "name": "Highlight",
                "font": {"bold": True, "color": "0000FF"},
                "fill": {"patternType": "solid", "fg": "CCFFCC"},
            }
        ],
    )
    wb = _load(path)
    assert "Highlight" in wb.style_names
    assert wb.active["A1"].style == "Highlight"


def test_unstyled_path_still_works(tmp_path: Path):
    """Pay-for-what-you-use: unstyled W1 path still round-trips."""
    path = tmp_path / "plain.xlsx"
    write_excel_turbo(
        str(path),
        [{"name": "P", "columns": [[1, 2, 3], ["a", "b", "c"]]}],
    )
    wb = _load(path)
    assert wb.active["A1"].value == 1
    assert wb.active["B2"].value == "b"
    assert wb.active["A1"].has_style is False
    with zipfile.ZipFile(path) as z:
        styles = z.read("xl/styles.xml").decode()
    assert "gray125" in styles
    assert "Normal" in styles


def test_bytes_api_styles():
    data = write_excel_turbo_bytes(
        [
            {
                "name": "S",
                "rows": [[1]],
                "cell_styles": {
                    (0, 0): {"font": {"bold": True, "name": "Arial"}}
                },
            }
        ]
    )
    assert data[:2] == b"PK"
    import io

    wb = load_workbook(io.BytesIO(data))
    assert wb.active["A1"].font.bold is True
