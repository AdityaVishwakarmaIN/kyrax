"""Stream C gap-parts oracle tests for turbo reader.

Charts / pivots / VBA / threaded comments vs gaplab/parts oracle_expected.json.
Fixtures copied into testdata/ (charts.xlsx, pivot.xlsx, vba.xlsm, threaded.xlsx).
"""

from __future__ import annotations

import json
import sys
import time
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from kyrax import read_excel_turbo  # noqa: E402

TESTDATA = ROOT / "testdata"
ORACLE_PATH = TESTDATA / "gap_parts_oracle.json"
CHARTS = TESTDATA / "charts.xlsx"
PIVOT = TESTDATA / "pivot.xlsx"
VBA = TESTDATA / "vba.xlsm"
THREADED = TESTDATA / "threaded.xlsx"

FEATURES_ALL = "all"


@pytest.fixture(scope="module")
def oracle():
    return json.loads(ORACLE_PATH.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# C1 / C5 — Charts + chartsheet
# ---------------------------------------------------------------------------


def test_charts_sheet_kinds(oracle):
    exp = oracle["charts"]
    reader = read_excel_turbo(str(CHARTS))
    assert reader.sheet_names == [s["name"] for s in exp["sheets"]]
    for s in exp["sheets"]:
        sh = reader.load_sheet(s["name"], features=FEATURES_ALL)
        assert sh.sheet_kind == s["kind"].lower()
        assert sh.sheet_state == s["state"]


def test_chartsheet_empty_grid():
    reader = read_excel_turbo(str(CHARTS))
    sh = reader.load_sheet("ChartOnly", features=FEATURES_ALL)
    assert sh.sheet_kind == "chartsheet"
    assert sh.nrows == 0
    assert sh.ncols == 0
    batch = sh.to_arrow()
    assert batch.num_rows == 0
    charts = sh.charts() or []
    assert len(charts) == 1
    assert charts[0]["type"] == "bar"
    assert charts[0]["title"] == "Chartsheet Bar"
    assert charts[0]["anchor"]["kind"] == "absolute"


def test_charts_metadata_vs_oracle(oracle):
    exp_charts = oracle["charts"]["charts"]
    reader = read_excel_turbo(str(CHARTS))
    all_charts = []
    for name in reader.sheet_names:
        sh = reader.load_sheet(name, features=["charts"])
        for c in sh.charts() or []:
            c = dict(c)
            c["sheet_name"] = name
            c["sheet_kind"] = sh.sheet_kind
            all_charts.append(c)

    assert len(all_charts) == oracle["charts"]["chart_count"] == 5

    # Match by (sheet_name, type, title)
    def key(c):
        return (c.get("sheet_name") or c.get("sheet"), c["type"], c.get("title"))

    got_by = {key(c): c for c in all_charts}
    for exp in exp_charts:
        k = (exp["sheet_name"], exp["type"], exp["title"])
        assert k in got_by, f"missing chart {k}; have {list(got_by)}"
        g = got_by[k]
        assert g["x_axis_title"] == exp["x_axis_title"]
        assert g["y_axis_title"] == exp["y_axis_title"]
        assert g["anchor"]["kind"] == exp["anchor"]["kind"]
        if "from" in exp["anchor"]:
            assert g["anchor"]["from"]["col"] == exp["anchor"]["from"]["col"]
            assert g["anchor"]["from"]["row"] == exp["anchor"]["from"]["row"]
        assert len(g["series"]) == len(exp["series"])
        for gs, es in zip(g["series"], exp["series"]):
            assert gs["title_ref"] == es["title_ref"]
            assert gs["title_cache"] == es["title_cache"]
            assert gs["categories_ref"] == es["categories_ref"]
            assert gs["categories_cache"] == es["categories_cache"]
            assert gs["values_ref"] == es["values_ref"]
            assert gs["values_cache"] == es["values_cache"]


def test_charts_injected_numcache(oracle):
    """Bar series on Sales has hand-injected numCache/strCache."""
    reader = read_excel_turbo(str(CHARTS))
    sh = reader.load_sheet("Sales", features=["charts"])
    bar = next(c for c in sh.charts() if c["type"] == "bar")
    ser = bar["series"][0]
    assert ser["title_cache"] == ["Rev"]
    assert ser["values_cache"] == [10.0, 20.0, 30.0]
    assert oracle["charts"]["injected_cache_visible_to_openpyxl"] is True


def test_charts_selective_skip():
    reader = read_excel_turbo(str(CHARTS))
    sh = reader.load_sheet(0, features=["values"])
    assert sh.charts() is None
    sh2 = reader.load_sheet(0, features=["charts"])
    assert sh2.charts() is not None
    assert len(sh2.charts()) == 2


def test_openpyxl_readonly_drops_worksheet_charts(oracle):
    """Enhancement claim (b): openpyxl read_only drops worksheet charts; turbo does not."""
    contrast = oracle["charts_readonly_contrast"]
    wb_full = openpyxl.load_workbook(CHARTS, read_only=False)
    full_ws = sum(len(getattr(wb_full[n], "_charts", []) or []) for n in ("Sales", "Mix"))
    full_cs = len(getattr(wb_full["ChartOnly"], "_charts", []) or [])
    wb_full.close()

    wb_ro = openpyxl.load_workbook(CHARTS, read_only=True)
    ro_ws = 0
    for n in ("Sales", "Mix"):
        ws = wb_ro[n]
        ro_ws += len(getattr(ws, "_charts", []) or [])
        assert type(ws).__name__ == "ReadOnlyWorksheet"
    ro_cs = len(getattr(wb_ro["ChartOnly"], "_charts", []) or [])
    wb_ro.close()

    assert full_ws == contrast["full_mode_worksheet_charts"]
    assert full_cs == contrast["full_mode_chartsheet_charts"]
    assert ro_ws == contrast["read_only_worksheet_charts"] == 0
    assert ro_cs == contrast["read_only_chartsheet_charts"]
    assert contrast["worksheet_charts_dropped"] is True

    # turbo single path surfaces all under Features::CHARTS
    reader = read_excel_turbo(str(CHARTS))
    turbo_total = 0
    for name in reader.sheet_names:
        sh = reader.load_sheet(name, features=["charts"])
        turbo_total += len(sh.charts() or [])
    assert turbo_total == contrast["full_mode_chart_count"] == 5


# ---------------------------------------------------------------------------
# C2 — Pivots
# ---------------------------------------------------------------------------


def test_pivot_metadata(oracle):
    exp = oracle["pivots"]["pivots"][0]
    reader = read_excel_turbo(str(PIVOT))
    assert reader.sheet_names == [exp["sheet_name"]]
    sh = reader.load_sheet(0, features=["pivots"])
    pivots = sh.pivots()
    assert pivots is not None
    assert len(pivots) == oracle["pivots"]["pivot_count"] == 1
    p = pivots[0]
    assert p["name"] == exp["name"]
    assert p["cache_id"] == exp["cache_id"]
    assert p["location_ref"] == exp["location_ref"]
    assert p["row_fields"] == exp["row_fields"]
    assert p["col_fields"] == exp["col_fields"]
    assert p["data_fields"] == exp["data_fields"]
    assert p["cache_field_names"] == exp["cache_field_names"]
    assert p["cache_source"]["type"] == exp["cache_source"]["type"]
    assert p["cache_source"]["sheet"] == exp["cache_source"]["sheet"]
    assert p["cache_source"]["ref"] == exp["cache_source"]["ref"]


def test_pivot_selective_skip():
    reader = read_excel_turbo(str(PIVOT))
    sh = reader.load_sheet(0, features=["values"])
    assert sh.pivots() is None


# ---------------------------------------------------------------------------
# C3 — VBA
# ---------------------------------------------------------------------------


def test_vba_bytes_roundtrip(oracle):
    exp = oracle["vba"]["keep_vba_True"]
    reader = read_excel_turbo(str(VBA))
    _ = reader.load_sheet(0, features=["vba"])
    assert reader.has_vba is True
    blob = reader.vba_project()
    assert blob is not None
    assert len(blob) == exp["blob_len"]
    assert blob.hex()[: len(exp["blob_hex_prefix"])] == exp["blob_hex_prefix"]
    # fixture content
    assert blob.startswith(b"FAKE_VBA_PROJECT_BLOB_FOR_GAPLAB_STREAM_C")


def test_vba_absent_on_xlsx():
    reader = read_excel_turbo(str(CHARTS))
    _ = reader.load_sheet(0, features=["vba"])
    assert reader.has_vba is False
    assert reader.vba_project() is None


def test_vba_selective_skip():
    reader = read_excel_turbo(str(VBA))
    _ = reader.load_sheet(0, features=["values"])
    # not requested → has_vba false / no bytes
    assert reader.has_vba is False
    assert reader.vba_project() is None


# ---------------------------------------------------------------------------
# C4 — Threaded comments (openpyxl cannot)
# ---------------------------------------------------------------------------


def test_threaded_vs_hand_encoded(oracle):
    exp = oracle["threaded"]
    reader = read_excel_turbo(str(THREADED))
    sh = reader.load_sheet(0, features=FEATURES_ALL)

    persons = reader.persons()
    assert persons is not None
    assert len(persons) == len(exp["hand_encoded_persons"])
    by_id = {p["id"]: p for p in persons}
    for ep in exp["hand_encoded_persons"]:
        g = by_id[ep["id"]]
        assert g["display_name"] == ep["displayName"]
        assert g["user_id"] == ep["userId"]
        assert g["provider_id"] == ep["providerId"]

    threaded = sh.threaded_comments()
    assert threaded is not None
    assert len(threaded) == len(exp["hand_encoded_threaded"])
    for gt, et in zip(threaded, exp["hand_encoded_threaded"]):
        assert gt["ref"] == et["ref"]
        assert gt["id"] == et["id"]
        assert gt["person_id"] == et["personId"]
        assert gt["parent_id"] == et["parentId"]
        assert gt["done"] is et["done"]
        assert gt["text"] == et["text"]
        assert gt["datetime"] == et["dT"]
        # display name resolved
        assert gt["person_display_name"] in ("Alice", "Bob")


def test_threaded_dedupe_legacy_mirror(oracle):
    """Prefer threaded; legacy still exposed with legacy_is_mirror=True."""
    exp = oracle["threaded"]
    reader = read_excel_turbo(str(THREADED))
    sh = reader.load_sheet(0, features=["comments"])
    assert sh.legacy_is_mirror is True

    legacy = sh.comments()
    assert legacy is not None
    assert legacy.num_rows == exp["openpyxl_view"]["legacy_comments_openpyxl_count"] == 4

    refs = [legacy.column("ref")[i].as_py() for i in range(legacy.num_rows)]
    assert exp["legacy_only_control_cell"] in refs  # E5 pure legacy remains

    threaded = sh.threaded_comments()
    assert len(threaded) == 4
    # openpyxl claim: no persons / no threaded
    assert exp["openpyxl_view"]["has_persons_attr"] is False
    assert exp["openpyxl_view"]["threaded_via_openpyxl"] is None


def test_openpyxl_cannot_read_threaded(oracle):
    """Enhancement claim (a): openpyxl returns only legacy mirrors."""
    wb = openpyxl.load_workbook(THREADED)
    has_persons = hasattr(wb, "persons")
    ws = wb.active
    legacy = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                legacy.append((cell.coordinate, cell.comment.author, cell.comment.text))
    wb.close()
    assert len(legacy) == 4
    assert has_persons is False or getattr(openpyxl.Workbook(), "persons", None) is None
    # turbo surfaces 4 threaded + 4 legacy
    reader = read_excel_turbo(str(THREADED))
    sh = reader.load_sheet(0, features="all")
    assert len(sh.threaded_comments()) == 4
    assert sh.comments().num_rows == 4


# ---------------------------------------------------------------------------
# Timing (features=all on charts.xlsx — small parts, expect trivially fast)
# ---------------------------------------------------------------------------


def test_charts_all_timing_smoke():
    path = str(CHARTS)
    reader = read_excel_turbo(path)
    # warmup
    _ = reader.load_sheet(0, features=FEATURES_ALL).charts()
    times = []
    for _ in range(20):
        t0 = time.perf_counter()
        r = read_excel_turbo(path)
        for name in r.sheet_names:
            sh = r.load_sheet(name, features=FEATURES_ALL)
            _ = sh.charts()
            _ = sh.to_arrow()
        times.append(time.perf_counter() - t0)
    mean_ms = (sum(times) / len(times)) * 1000.0
    # not a hard perf gate — just ensure it runs and is in a sane ballpark
    assert mean_ms < 500.0, f"unexpectedly slow mean {mean_ms:.2f} ms"
    # stash for human report via pytest -s
    print(f"\n[timing] charts.xlsx features=all mean={mean_ms:.2f} ms (n=20)")
