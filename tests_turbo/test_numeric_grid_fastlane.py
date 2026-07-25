import pytest
import tempfile
from pathlib import Path
import kyrax

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

@pytest.mark.skipif(not HAS_NUMPY, reason="NumPy not installed")
def test_numeric_grid_fastlane_float64():
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "dense_float64.xlsx"
        arr = np.arange(100, dtype=np.float64).reshape((10, 10))

        # Test single sheet numpy array pass-through
        kyrax.write_excel_turbo(
            str(output_path),
            [{"name": "Sheet1", "data": arr}],
        )
        assert output_path.exists()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path, data_only=True)
            ws = wb["Sheet1"]
            assert ws.cell(row=1, column=1).value == 0.0
            assert ws.cell(row=10, column=10).value == 99.0
        except ImportError:
            pass

@pytest.mark.skipif(not HAS_NUMPY, reason="NumPy not installed")
def test_numeric_grid_fastlane_float32():
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "dense_float32.xlsx"
        arr = np.array([[1.5, 2.5], [3.5, 4.5]], dtype=np.float32)

        kyrax.write_excel_turbo(
            str(output_path),
            [{"name": "Sheet1", "data": arr}],
        )
        assert output_path.exists()

@pytest.mark.skipif(not HAS_NUMPY, reason="NumPy not installed")
def test_numeric_grid_transposed_fallback():
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "transposed_fallback.xlsx"
        arr = np.arange(20, dtype=np.float64).reshape((4, 5)).T # Non C-contiguous view

        # Should fall back safely without memory corruption
        kyrax.write_excel_turbo(
            str(output_path),
            [{"name": "Sheet1", "data": arr}],
        )
        assert output_path.exists()
