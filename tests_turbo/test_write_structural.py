"""W3 structural write path — silo C oracle semantics via turbo write."""
from __future__ import annotations

import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart.bar_chart import BarChart
from openpyxl.chart.line_chart import LineChart
from openpyxl.chart.pie_chart import PieChart
from openpyxl.chart.scatter_chart import ScatterChart

from kyrax import write_excel_turbo, write_excel_turbo_bytes

DATA_SHEET = {
    "name": "Data",
    "columns": [
        ["Category", "A", "B", "C", "D"],
        ["Sales", 10, 20, 30, 40],
        ["Cost", 5, 10, 15, 20],
        [None, None, None, None, None],
        ["Merged", None, None, None, None],
        [None, None, None, None, None],
        ["link", "internal", None, None, None],
        ["X", 1, 2, 3, 4],
        ["Y", 1, 4, 9, 16],
    ],
}


def _base_dense_sheet() -> dict:
    sh = dict(DATA_SHEET)
    sh.update(
        {
            "tab_color": "FF0000",
            "freeze_panes": "B2",
            "protection": {"sheet": True, "password": "secret"},
            "scenarios": [{"name": "Base", "cells": {"B2": "10"}}],
            "auto_filter": "A1:C5",
            "merges": ["E1:F1"],
            "hyperlinks": [
                {
                    "ref": "G1",
                    "target": "https://example.com",
                    "display": "link",
                },
                {
                    "ref": "G2",
                    "location": "Data!A1",
                    "display": "internal",
                },
            ],
            "print_options": {"horizontal_centered": True},
            "page_margins": {"left": 0.5},
            "page_setup": {"orientation": "landscape", "fit_to_page": True},
            "header_footer": {
                "odd_header_center": "Header",
                "odd_footer_left": "Footer",
            },
            "row_breaks": [3],
            "col_breaks": [2],
            "print_area": "A1:C5",
            "print_titles": "'Data'!$1:$1",
            "tables": [
                {
                    "display_name": "SalesTable",
                    "ref": "A1:C5",
                    "columns": ["Category", "Sales", "Cost"],
                    "style_name": "TableStyleMedium9",
                    "show_row_stripes": True,
                }
            ],
            "comments": [
                {"ref": "A1", "author": "Alice", "text": "Hello comment", "height": 50, "width": 100},
                {"ref": "B2", "author": "Bob", "text": "Second"},
            ],
            "charts": [
                {
                    "type": "col",
                    "title": "Bar Title",
                    "series": [
                        {
                            "title_ref": "'Data'!B1",
                            "cat_ref": "'Data'!$A$2:$A$5",
                            "val_ref": "'Data'!$B$2:$B$5",
                        }
                    ],
                    "anchor": {"cell": "E3", "width_cm": 15, "height_cm": 7.5},
                },
                {
                    "type": "line",
                    "title": "Line Title",
                    "series": [
                        {
                            "cat_ref": "'Data'!$A$2:$A$5",
                            "val_ref": "'Data'!$B$2:$B$5",
                        }
                    ],
                    "anchor": "E15",
                },
                {
                    "type": "pie",
                    "title": "Pie Title",
                    "series": [
                        {
                            "cat_ref": "'Data'!$A$2:$A$5",
                            "val_ref": "'Data'!$B$2:$B$5",
                        }
                    ],
                    "anchor": "E25",
                },
                {
                    "type": "scatter",
                    "title": "Scatter Title",
                    "series": [
                        {
                            "x_ref": "'Data'!$H$2:$H$5",
                            "y_ref": "'Data'!$I$2:$I$5",
                        }
                    ],
                    "anchor": "E35",
                },
            ],
        }
    )
    return sh


def title_text(chart) -> str | None:
    if chart.title is None:
        return None
    try:
        tx = chart.title.tx
        if tx is None:
            return None
        if tx.rich is not None:
            parts = []
            for p in tx.rich.paragraphs:
                for r in p.r or []:
                    if getattr(r, "t", None):
                        parts.append(r.t)
            return "".join(parts) or None
    except Exception:
        return None
    return None


def series_val_ref(chart) -> str | None:
    if not chart.series:
        return None
    s0 = chart.series[0]
    if getattr(s0, "val", None) is not None and s0.val is not None:
        if s0.val.numRef is not None:
            return s0.val.numRef.f
    if getattr(s0, "yVal", None) is not None and s0.yVal is not None:
        if s0.yVal.numRef is not None:
            return s0.yVal.numRef.f
    return None


def test_feature_dense_roundtrip(tmp_path: Path):
    path = tmp_path / "feature_dense.xlsx"
    write_excel_turbo(
        str(path),
        [_base_dense_sheet()],
        props={
            "title": "Test Title",
            "creator": "SiloC",
            "custom": {"Dept": "Engineering"},
        },
        defined_names=[{"name": "MyRange", "value": "'Data'!$A$1:$B$2"}],
        lock_structure=True,
        features="all",
    )

    wb = load_workbook(path)
    ws = wb.active

    assert wb.properties.title == "Test Title"
    assert wb.properties.creator == "SiloC"
    assert any(p.name == "Dept" and p.value == "Engineering" for p in wb.custom_doc_props)
    assert wb.security is not None and wb.security.lockStructure

    assert ws.freeze_panes == "B2"
    tab = ws.sheet_properties.tabColor
    assert tab is not None and (tab.rgb or "").endswith("FF0000")
    assert ws.auto_filter.ref == "A1:C5"
    merges = {str(r) for r in ws.merged_cells.ranges}
    assert "E1:F1" in merges
    assert ws.protection.sheet is True
    assert (ws.protection.password or "").upper() == "DAA7"

    hl = ws["G1"].hyperlink
    assert hl is not None and hl.target == "https://example.com"
    hl2 = ws["G2"].hyperlink
    assert hl2 is not None and (
        hl2.location == "Data!A1" or (hl2.target and "Data" in str(hl2.target))
    )

    assert ws.page_setup.orientation == "landscape"
    assert abs(float(ws.page_margins.left) - 0.5) < 1e-9
    assert ws.print_options.horizontalCentered is True
    assert "Header" in (ws.oddHeader.center.text or "")
    assert "Footer" in (ws.oddFooter.left.text or "")
    assert any(b.id == 3 for b in ws.row_breaks.brk)
    assert any(b.id == 2 for b in ws.col_breaks.brk)
    assert "A1:C5" in (ws.print_area or "").replace("$", "")

    names = list(wb.defined_names.keys())
    assert "MyRange" in names

    assert "SalesTable" in ws.tables
    t = ws.tables["SalesTable"]
    assert t.ref == "A1:C5"
    assert [c.name for c in t.tableColumns] == ["Category", "Sales", "Cost"]
    assert t.tableStyleInfo is not None and t.tableStyleInfo.name == "TableStyleMedium9"

    c1 = ws["A1"].comment
    assert c1 is not None and c1.text == "Hello comment"
    assert c1.author == "Alice"
    c2 = ws["B2"].comment
    assert c2 is not None and c2.author == "Bob"

    assert ws.scenarios is not None and len(ws.scenarios.scenario) >= 1

    charts = ws._charts
    assert len(charts) == 4
    types = [type(c).__name__ for c in charts]
    assert any(t == "BarChart" for t in types)
    assert any(t == "LineChart" for t in types)
    assert any(t == "PieChart" for t in types)
    assert any(t == "ScatterChart" for t in types)

    for c in charts:
        tt = title_text(c)
        if isinstance(c, BarChart):
            assert tt == "Bar Title"
            assert series_val_ref(c) is not None and "B" in series_val_ref(c)
        if isinstance(c, LineChart):
            assert tt == "Line Title"
        if isinstance(c, PieChart):
            assert tt == "Pie Title"
        if isinstance(c, ScatterChart):
            assert tt == "Scatter Title"

    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        for p in [
            "xl/drawings/drawing1.xml",
            "xl/charts/chart1.xml",
            "xl/comments/comment1.xml",
            "xl/drawings/commentsDrawing1.vml",
            "xl/tables/table1.xml",
            "docProps/custom.xml",
        ]:
            assert p in names, f"missing {p}"

    with zipfile.ZipFile(path) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8", "replace")
        order = [
            "sheetProtection",
            "scenarios",
            "autoFilter",
            "mergeCells",
            "hyperlinks",
            "printOptions",
            "pageMargins",
            "pageSetup",
            "headerFooter",
            "rowBreaks",
            "colBreaks",
            "drawing",
            "legacyDrawing",
            "tableParts",
        ]
        last = -1
        positions = []
        for tag in order:
            i = sheet_xml.find(f"<{tag}")
            assert i >= 0, f"missing {tag}"
            assert i > last, f"order broken at {tag}: {positions}"
            positions.append((tag, i))
            last = i


def test_chart_types_and_chartsheet(tmp_path: Path):
    path = tmp_path / "chart_types.xlsx"
    cat = "'Data'!$A$2:$A$5"
    val = "'Data'!$B$2:$B$5"
    series = [{"cat_ref": cat, "val_ref": val}]
    scatter_ser = [{"x_ref": "'Data'!$H$2:$H$5", "y_ref": "'Data'!$I$2:$I$5"}]
    bubble_ser = [
        {
            "x_ref": "'Data'!$H$2:$H$5",
            "y_ref": "'Data'!$I$2:$I$5",
            "bubble_size_ref": "'Data'!$B$2:$B$5",
        }
    ]
    types = [
        "bar",
        "bar3d",
        "col",
        "col3d",
        "line",
        "line3d",
        "area",
        "area3d",
        "pie",
        "pie3d",
        "doughnut",
        "scatter",
        "bubble",
        "radar",
        "stock",
        "surface",
        "surface3d",
    ]
    charts = []
    for i, t in enumerate(types):
        ser = bubble_ser if t == "bubble" else (scatter_ser if t == "scatter" else series)
        charts.append(
            {
                "type": t,
                "title": t,
                "series": ser,
                "anchor": {"cell": f"E{3 + i * 10}", "width_cm": 12, "height_cm": 6},
            }
        )
    sheet = dict(DATA_SHEET)
    sheet["charts"] = charts
    write_excel_turbo(
        str(path),
        [sheet],
        chartsheets=[
            {
                "name": "ChartSheet1",
                "charts": [
                    {
                        "type": "col",
                        "title": "CS Chart",
                        "series": series,
                    }
                ],
            }
        ],
        features="all",
    )
    wb = load_workbook(path)
    ws = wb["Data"]
    found = {type(c).__name__ for c in ws._charts}
    expected = {
        "BarChart",
        "BarChart3D",
        "LineChart",
        "LineChart3D",
        "AreaChart",
        "AreaChart3D",
        "PieChart",
        "PieChart3D",
        "DoughnutChart",
        "ScatterChart",
        "BubbleChart",
        "RadarChart",
        "StockChart",
        "SurfaceChart",
        "SurfaceChart3D",
    }
    for t in expected:
        assert t in found, f"missing {t}; have {found}"
    assert len(wb.chartsheets) >= 1
    assert len(wb.chartsheets[0]._charts) >= 1


def test_external_link_stub(tmp_path: Path):
    path = tmp_path / "extlink.xlsx"
    write_excel_turbo(
        str(path),
        [{"name": "S", "columns": [["a", 1]]}],
        external_links=["file:///C:/other.xlsx"],
        features="all",
    )
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        assert any(
            n.startswith("xl/externalLinks/externalLink") and n.endswith(".xml") for n in names
        )
    load_workbook(path)  # opens cleanly


def test_plain_values_no_structural(tmp_path: Path):
    path = tmp_path / "plain.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Data",
                "columns": [
                    ["Category", "A", "B"],
                    ["Sales", 10, 20],
                ],
            }
        ],
    )
    wb = load_workbook(path)
    assert wb.active["A1"].value == "Category"
    assert len(wb.active._charts) == 0
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        assert not any("drawing" in n for n in names)
        assert not any("comment" in n for n in names)


def test_bytes_api_structural():
    data = write_excel_turbo_bytes(
        [_base_dense_sheet()],
        props={"title": "Bytes", "creator": "T"},
        features="all",
    )
    assert data[:2] == b"PK"
    # openpyxl from bytes
    from io import BytesIO

    wb = load_workbook(BytesIO(data))
    assert wb.properties.title == "Bytes"
    assert len(wb.active._charts) == 4


def test_merges_only_feature_flag(tmp_path: Path):
    path = tmp_path / "merges.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "S",
                "columns": [["A", 1], ["B", 2]],
                "merges": ["A1:B1"],
            }
        ],
    )
    wb = load_workbook(path)
    assert "A1:B1" in {str(r) for r in wb.active.merged_cells.ranges}


def test_auto_filter_value_filters_surface(tmp_path: Path):
    path = tmp_path / "autofilter_filters.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Data",
                "columns": [["Category", "Alice", "Bob", "Carol"]],
                "auto_filter": {
                    "ref": "A1:A4",
                    "columns": [
                        {
                            "col_id": 0,
                            "hidden_button": False,
                            "show_button": True,
                            "values": ["Alice", "Carol"],
                            "blank": False,
                        }
                    ],
                },
            }
        ],
        features="all",
    )

    wb = load_workbook(path)
    ws = wb.active
    assert ws.auto_filter.ref == "A1:A4"
    assert len(ws.auto_filter.filterColumn) == 1
    fc = ws.auto_filter.filterColumn[0]
    assert fc.colId == 0
    assert [str(f) for f in fc.filters.filter] == ["Alice", "Carol"]
    assert fc.filters.blank is False

    # read-back via kyrax sees the same value filters
    from kyrax import read_excel_turbo

    reader = read_excel_turbo(str(path))
    sh = reader.load_sheet("Data", features=["sheet_meta"])
    af = sh.auto_filter()
    assert af is not None
    assert af["ref"] == "A1:A4"
    assert af["columns"][0]["values"] == ["Alice", "Carol"]
    assert af["columns"][0]["blank"] is False


def test_auto_filter_values_only_workload_emits_nothing(tmp_path: Path):
    # No auto_filter in the model: a values-only workload must not emit any
    # autoFilter XML (the MERGES feature flag stays off).
    path = tmp_path / "autofilter_off.xlsx"
    write_excel_turbo(
        str(path),
        [{"name": "Data", "columns": [["Category", "Alice", "Bob"]]}],
    )
    with zipfile.ZipFile(path) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "<autoFilter" not in xml
    assert "<filterColumn" not in xml


def test_defined_names_print_forms(tmp_path: Path):
    path = tmp_path / "names.xlsx"
    write_excel_turbo(
        str(path),
        [
            {
                "name": "Data",
                "columns": [["H", "a"], ["V", 1]],
                "print_area": "A1:B2",
                "auto_filter": "A1:B2",
            }
        ],
        defined_names=[{"name": "GlobalX", "value": "'Data'!$A$1"}],
    )
    wb = load_workbook(path)
    keys = list(wb.defined_names.keys())
    assert "GlobalX" in keys
    # print area visible on sheet
    assert "A1:B2" in (wb.active.print_area or "").replace("$", "")
