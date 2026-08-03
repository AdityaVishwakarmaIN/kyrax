"""Python-level tests for row/column insert and delete through the edit overlay.

Everything here goes through the public API: ``kyrax.load_workbook(path,
edit_mode=True)`` (or ``kyrax.edit_excel``), sheet operations like
``insert_rows`` / ``delete_rows`` / ``insert_cols`` / ``delete_cols`` (all
1-based, openpyxl semantics), then ``save()``.

Shifts are applied to the sheet XML before cell edits, so an edit coordinate is
final while a shift moves the grid under it. A refusal is all-or-nothing: the
destination file is never written and the source file is never modified.
"""

import zipfile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

import kyrax


def _write_sheet_grid(path, rows, cols, start_val=1):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=start_val + (r - 1) * cols + (c - 1))
    wb.save(path)
    wb.close()


def _write_simple(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "a"
    ws["A2"] = "b"
    ws["A3"] = "c"
    ws["B1"] = 10
    ws["C1"] = "x"
    ws["C2"] = "y"
    wb.save(path)
    wb.close()


def _sheet_xml_bytes(path):
    """Extract the first worksheet part's raw XML bytes from ``path``."""
    with zipfile.ZipFile(path) as z:
        names = sorted(
            n for n in z.namelist() if n.startswith("xl/worksheets/") and n.endswith(".xml")
        )
        assert names, "no worksheet part found"
        return z.read(names[0])


# ---------------------------------------------------------------------------
# insert then read back
# ---------------------------------------------------------------------------

def test_insert_rows_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _write_simple(str(src))

    wb = kyrax.load_workbook(str(src), edit_mode=True)
    wb["Data"].insert_rows(2, 2)
    wb.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    # rows 2-3 are the new blank rows; the originals shift down by 2
    assert ws["A1"].value == "a"
    assert ws["A2"].value is None
    assert ws["A3"].value is None
    assert ws["A4"].value == "b"
    assert ws["A5"].value == "c"
    ld.close()


def test_insert_rows_at_one_shifts_everything(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _write_simple(str(src))

    wb = kyrax.edit_excel(str(src))
    wb["Data"].insert_rows(1, 1)
    wb.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value is None
    assert ws["A2"].value == "a"
    assert ws["B2"].value == 10
    assert ws["C2"].value == "x"
    assert ws["A3"].value == "b"
    assert ws["A4"].value == "c"
    ld.close()


def test_insert_cols_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _write_simple(str(src))

    wb = kyrax.edit_excel(str(src))
    wb["Data"].insert_cols(2, 1)  # A stays, B is the new blank col
    wb.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value == "a"
    assert ws["B1"].value is None
    assert ws["C1"].value == 10   # old B1 shifted right
    assert ws["D1"].value == "x"  # old C1 shifted right
    assert ws["D2"].value == "y"  # old C2 shifted right
    ld.close()


# ---------------------------------------------------------------------------
# delete then read back
# ---------------------------------------------------------------------------

def test_delete_rows_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _write_simple(str(src))

    wb = kyrax.edit_excel(str(src))
    wb["Data"].delete_rows(2, 1)  # remove "b"
    wb.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value == "a"
    assert ws["A2"].value == "c"
    assert ws["A3"].value is None
    ld.close()


def test_delete_cols_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _write_simple(str(src))

    wb = kyrax.edit_excel(str(src))
    wb["Data"].delete_cols(2, 1)  # remove column B
    wb.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value == "a"
    assert ws["B1"].value == "x"  # old C1 shifted left into deleted B's place
    assert ws["B2"].value == "y"  # old C2 shifted left
    assert ws["C1"].value is None
    ld.close()


# ---------------------------------------------------------------------------
# a formula cell referencing a shifted cell has moved
# ---------------------------------------------------------------------------
# Assert on the saved sheet XML directly, NOT through a reader: the turbo arrow
# reader treats row 1 as a header and its view would mislead here. The mutate
# splice shifts the CELL (its r= coordinate) while keeping formula bodies
# shifted, so assert both the coordinate and the references move.

def test_formula_cell_moves_on_insert(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=A1+A2"
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_rows(1, 1)
    ed.save(str(out))

    xml = _sheet_xml_bytes(str(out)).decode("utf-8")
    # The formula cell moved from row 3 to row 4.
    assert 'r="4"' in xml, xml
    assert '<c r="A4"' in xml, xml
    # ...and its REFERENCES moved with the grid. Inserting a row at 1 pushes the
    # operands from A1/A2 down to A2/A3, so the formula must read A2+A3. An
    # earlier version of this test asserted the body passed through untouched,
    # which encoded the T1-1f defect: the cell moved but the formula did not,
    # so the workbook silently computed from the wrong cells.
    assert "<f>A2+A3</f>" in xml, xml
    assert "<f>A1+A2</f>" not in xml, xml


# ---------------------------------------------------------------------------
# merged ranges: survive an insert above, trim on a delete inside
# ---------------------------------------------------------------------------

def test_merge_survives_insert_above(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "h"
    ws["A2"] = "m"
    ws.merge_cells("A2:A3")
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_rows(1, 1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert [str(m) for m in ws.merged_cells.ranges] == ["A3:A4"]
    ld.close()


def test_merge_trimmed_by_delete_inside(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "m"
    ws.merge_cells("A1:A5")
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].delete_rows(3, 1)  # delete a row inside the merge
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert [str(m) for m in ws.merged_cells.ranges] == ["A1:A4"]
    ld.close()


# ---------------------------------------------------------------------------
# refusal: raises, and the ORIGINAL file is left byte-identical on disk
# ---------------------------------------------------------------------------

def _refusal_asserts(src, src_before, out, exc, ed):
    with pytest.raises(exc):
        ed.save(str(out))
    assert src.read_bytes() == src_before, "source file must be untouched"
    assert not out.exists(), "no output may be written on a refusal"


def test_refusal_row_overflow_leaves_file_unmodified(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1048576, column=1, value=1)  # last row of the grid
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_rows(1048576, 1)  # would push past row 1048576
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


def test_refusal_table_header_delete_leaves_file_unmodified(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "H1"
    ws["B1"] = "H2"
    ws["A2"] = 1
    ws["B2"] = 2
    ws["A3"] = 3
    ws["B3"] = 4
    tab = Table(displayName="Tab1", ref="A1:B3")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].delete_rows(1, 1)  # removes the table's header row
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


# ---------------------------------------------------------------------------
# insert(n) then delete(n) == no net change, byte-identical to a plain save
# ---------------------------------------------------------------------------

def test_insert_then_delete_is_byte_identical_to_no_ops(tmp_path):
    src = tmp_path / "src.xlsx"
    with_ops = tmp_path / "with_ops.xlsx"
    without = tmp_path / "without.xlsx"
    _write_sheet_grid(str(src), 6, 4)

    ed = kyrax.edit_excel(str(src))
    sh = ed["Data"]
    sh.set_cell(1, 1, 999)
    sh.insert_rows(2, 1)
    sh.delete_rows(2, 1)
    ed.save(str(with_ops))

    ed = kyrax.edit_excel(str(src))
    ed["Data"].set_cell(1, 1, 999)
    ed.save(str(without))

    assert with_ops.read_bytes() == without.read_bytes()


# ---------------------------------------------------------------------------
# move_range: relocate a block without shifting the rest of the grid
# ---------------------------------------------------------------------------

def test_move_range_down_right_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _write_simple(str(src))  # A1=a, A2=b, A3=c, B1=10, C1=x, C2=y

    ed = kyrax.edit_excel(str(src))
    ed["Data"].move_range("A1:B2", rows=2, cols=1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    # A1:a -> B3, B1:10 -> C3, A2:b -> B4, C1:C2 are untouched (outside the
    # source A1:B2 and outside the destination B3:C4).
    assert ws["B3"].value == "a"
    assert ws["C3"].value == 10
    assert ws["B4"].value == "b"
    assert ws["A1"].value is None
    assert ws["A2"].value is None
    assert ws["C1"].value == "x"
    assert ws["C2"].value == "y"
    ld.close()


def test_move_range_up_left_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B2"] = 2
    ws["B3"] = 4
    ws["C2"] = 3
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    # Move B2:C3 up 1 / left 1 -> A1:B2. B2:2 -> A1, C2:3 -> B1, B3:4 -> A2.
    ed["Data"].move_range("B2:C3", rows=-1, cols=-1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value == 2
    assert ws["B1"].value == 3
    assert ws["A2"].value == 4
    assert ws["B2"].value is None
    assert ws["B3"].value is None
    assert ws["C2"].value is None
    ld.close()


def test_move_range_overlap_down_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, 5):
        ws.cell(row=r, column=1, value=r)
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    # Move A1:A3 down by 1 (destination A2:A4 overlaps the source). Each cell
    # must keep its ORIGINAL value: A1=1 -> A2, A2=2 -> A3, A3=3 -> A4.
    ed["Data"].move_range("A1:A3", rows=1, cols=0)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value is None
    assert ws["A2"].value == 1
    assert ws["A3"].value == 2
    assert ws["A4"].value == 3
    ld.close()


def test_move_range_overlap_up_read_back(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, 5):
        ws.cell(row=r, column=1, value=r)
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    # Move A2:A4 up by 1 (destination A1:A3 overlaps the source).
    ed["Data"].move_range("A2:A4", rows=-1, cols=0)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value == 2
    assert ws["A2"].value == 3
    assert ws["A3"].value == 4
    assert ws["A4"].value is None
    ld.close()


def test_move_range_translate_true_shifts_formulas_inside_range(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["B1"] = "=A1+A2"
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    # Move B1 down 1 / right 1; translate=True shifts the formula's references
    # by the same offset (A1+A2 -> B2+B3).
    ed["Data"].move_range("B1:B1", rows=1, cols=1, translate=True)
    ed.save(str(out))

    xml = _sheet_xml_bytes(str(out)).decode("utf-8")
    assert '<c r="C2"' in xml, xml
    assert "<f>B2+B3</f>" in xml, xml
    assert "<f>A1+A2</f>" not in xml, xml


def test_move_range_translate_false_leaves_formulas_alone(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["B1"] = "=A1+A2"
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    # translate=False (default): the cell moves but its formula text is kept.
    ed["Data"].move_range("B1:B1", rows=1, cols=1)
    ed.save(str(out))

    xml = _sheet_xml_bytes(str(out)).decode("utf-8")
    assert '<c r="C2"' in xml, xml
    assert "<f>A1+A2</f>" in xml, xml


def test_move_range_refusal_out_of_grid_leaves_file_unmodified(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1048576, column=1, value=1)  # last row of the grid
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    # Moving A1048576 down by 1 would push it past the grid: refused.
    ed["Data"].move_range("A1048576:A1048576", rows=1, cols=0)
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


def test_move_range_merged_range_inside_block_follows(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "m"
    ws.merge_cells("A2:A3")
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    # The merge A2:A3 is fully contained in the moved block A1:B4, so it moves
    # with the range (down 1 / right 1 -> B3:B4).
    ed["Data"].move_range("A1:B4", rows=1, cols=1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert [str(m) for m in ws.merged_cells.ranges] == ["B3:B4"]
    ld.close()


def test_move_range_zero_offset_is_byte_identical_to_plain_save(tmp_path):
    src = tmp_path / "src.xlsx"
    with_move = tmp_path / "with_move.xlsx"
    without = tmp_path / "without.xlsx"
    _write_sheet_grid(str(src), 6, 4)

    ed = kyrax.edit_excel(str(src))
    ed["Data"].move_range("B2:D4", rows=0, cols=0)
    ed.save(str(with_move))

    ed = kyrax.edit_excel(str(src))
    ed.save(str(without))

    assert with_move.read_bytes() == without.read_bytes()
