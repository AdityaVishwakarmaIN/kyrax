"""Generate adversarial stress fixtures for turbo vs openpyxl.

Run from repo root:
  .venv/Scripts/python tests_turbo/gen_stress.py

Writes testdata/stress_*.xlsx (small; fixed seed). Safe to re-run.
"""

from __future__ import annotations

import random
import zipfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parents[1]
TESTDATA = ROOT / "testdata"
SEED = 42


def _ensure_dir() -> None:
    TESTDATA.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# 1. stress_sparse.xlsx
# ---------------------------------------------------------------------------


def gen_sparse(path: Path) -> None:
    """5000 data rows x 10 cols, sparse population, styles, formulas."""
    rng = random.Random(SEED)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"

    ncols = 10
    nrows = 5000
    # Header row 1
    for c in range(1, ncols + 1):
        ws.cell(1, c, f"C{c}")

    # Force a fully empty first data-row region: rows 2..4 empty (no cells).
    # Rows that start at col C+ and fully empty rows sprinkled throughout.
    for r in range(5, nrows + 2):  # sheet rows 5 .. 5001 → 4997 data rows after 3 empty
        # Map so total data rows (sheet 2..5001) = 5000
        pass

    # Build exactly 5000 data rows at sheet rows 2..5001
    for r in range(2, nrows + 2):
        mode = rng.randrange(100)
        if r <= 4:
            # fully empty first row region (rows 2-4)
            continue
        if mode < 8:
            # fully empty row
            continue
        if mode < 20:
            # row starts at col C+ (cols 3..10 only)
            cols = [c for c in range(3, ncols + 1) if rng.random() < 0.45]
        else:
            # random subset of columns (may include A/B)
            cols = [c for c in range(1, ncols + 1) if rng.random() < 0.4]
            if not cols and rng.random() < 0.5:
                # ensure some mid-row sparse: only high columns
                cols = [rng.randint(4, ncols)]

        for c in cols:
            cell = ws.cell(r, c)
            # Numeric + formulas only (avoid mixed-type sticky columns; that is a
            # separate LIMITATION covered in test_stress_types / onecell docs).
            kind = rng.randrange(10)
            if kind < 8:
                cell.value = float(r * 10 + c) + rng.random()
            else:
                # formula referencing col A of same row (may be empty → still stored)
                cell.value = f"=A{r}+1"
            if rng.random() < 0.15:
                cell.number_format = "0.00"
                cell.font = Font(bold=True, color="FF0000")
            if rng.random() < 0.05:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Guarantee a few known sparse patterns for hard assertions
    ws["A10"] = 100.0
    ws["C10"] = 300.0  # no B10
    ws["D15"] = 400.0  # first cell at D
    ws["A20"] = 1.0
    ws["C20"] = "=A20+1"
    ws["C20"].number_format = "0.00"

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# 2. stress_multisheet.xlsx
# ---------------------------------------------------------------------------


def gen_multisheet(path: Path) -> None:
    wb = Workbook()

    # Sheet 1: 100x5 values + table
    ws1 = wb.active
    ws1.title = "Values"
    for c in range(1, 6):
        ws1.cell(1, c, f"V{c}")
    for r in range(2, 102):
        for c in range(1, 6):
            ws1.cell(r, c, r * 10 + c)
    tab1 = Table(displayName="TblValues", ref="A1:E101")
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws1.add_table(tab1)

    # Sheet 2: 50x3 with merges + hyperlinks + sheet-scoped defined name
    ws2 = wb.create_sheet("MergedLinks")
    for c in range(1, 4):
        ws2.cell(1, c, f"M{c}")
    for r in range(2, 52):
        for c in range(1, 4):
            ws2.cell(r, c, f"s2r{r}c{c}")
    ws2.merge_cells("A2:B3")
    ws2.merge_cells("C5:C7")
    ws2["A2"] = "merged_top"
    ws2["A10"] = "link_here"
    ws2["A10"].hyperlink = "https://example.com/sheet2"
    ws2["B10"] = "mailto"
    ws2["B10"].hyperlink = "mailto:test@example.com"
    # sheet-scoped defined name on sheet 2 (localSheetId)
    dn2 = DefinedName(name="Sheet2Anchor", attr_text="'MergedLinks'!$A$2")
    # openpyxl: sheet-scoped via wb.defined_names with localSheetId
    try:
        dn2.localSheetId = 1  # 0-based sheet index
    except Exception:
        pass
    wb.defined_names.add(dn2)
    # also via worksheet API when available
    try:
        ws2.defined_names.add(DefinedName(name="LocalRange2", attr_text="'MergedLinks'!$A$1:$C$5"))
    except Exception:
        dn2b = DefinedName(name="LocalRange2", attr_text="'MergedLinks'!$A$1:$C$5")
        dn2b.localSheetId = 1
        wb.defined_names.add(dn2b)

    # Sheet 3: 30x2 with comments + table
    ws3 = wb.create_sheet("Comments")
    for c in range(1, 3):
        ws3.cell(1, c, f"K{c}")
    for r in range(2, 32):
        ws3.cell(r, 1, r)
        ws3.cell(r, 2, f"txt{r}")
    ws3["A2"].comment = Comment("hello from A2", "author1")
    ws3["B5"].comment = Comment("note on B5 with <entities> & ampersand", "author2")
    ws3["A10"].comment = Comment("emoji 🎉 and CJK 中文", "作者")
    tab3 = Table(displayName="TblComments", ref="A1:B31")
    tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws3.add_table(tab3)

    # Sheet 4: empty sheet + sheet-scoped defined name
    ws4 = wb.create_sheet("Empty")
    # leave completely empty (no cells)
    try:
        ws4.defined_names.add(DefinedName(name="EmptySheetFlag", attr_text="1"))
    except Exception:
        dn4 = DefinedName(name="EmptySheetFlag", attr_text="1")
        dn4.localSheetId = 3
        wb.defined_names.add(dn4)

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# 3. stress_onecell_everything.xlsx
# ---------------------------------------------------------------------------


def gen_onecell_everything(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Everything"

    # Small grid with headers
    headers = ["A", "B", "C", "D"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)

    # Merged range B2:C3 — top-left B2 has formula with cached value path
    ws.merge_cells("B2:C3")
    ws["B2"] = "=1+2"  # formula; openpyxl stores formula; cache may be empty until Excel
    ws["B2"].number_format = "0.00"
    ws["B2"].font = Font(bold=True, italic=True, name="Calibri", sz=14, color="0000FF")
    ws["B2"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ws["B2"].comment = Comment("merged formula cell", "tester")
    ws["B2"].hyperlink = "https://example.com/merged"

    # Several cells with value + custom format + font/fill + comment + hyperlink.
    # Keep each column mono-typed so sticky columnar typing does not drop values
    # (mixed-type LIMITATION is asserted separately in test_stress).
    for addr, val, fmt in [
        ("A2", 1234.5, "#,##0.00"),
        ("D2", -7.5, "0.000"),
        ("A4", 99, "0%"),
        ("D4", 42.125, "0.000"),
    ]:
        cell = ws[addr]
        cell.value = val
        cell.number_format = fmt
        cell.font = Font(bold=True, color="FF0000", name="Arial", sz=11)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.comment = Comment(f"comment on {addr}", "authorX")
        cell.hyperlink = f"https://example.com/{addr}"

    # String column C (mono-typed string) with full feature set
    ws["C4"] = "hello"
    ws["C4"].number_format = "@"
    ws["C4"].font = Font(bold=True, color="FF0000", name="Arial", sz=11)
    ws["C4"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["C4"].comment = Comment("comment on C4", "authorX")
    ws["C4"].hyperlink = "https://example.com/C4"

    # Plain numeric neighbors
    ws["A5"] = 0
    ws["D5"] = 1

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# 4. stress_types.xlsx
# ---------------------------------------------------------------------------


def gen_types(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Types"

    headers = [
        "bools",
        "unicode",
        "entities",
        "empty_str",
        "long_str",
        "floats",
        "date_edge",
        "formula_ent",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)

    # Row 2: primary interesting values
    ws["A2"] = True
    ws["B2"] = "emoji 🎉 CJK 中文日本語 한국어"
    ws["C2"] = "ents & < > \" ' amp"
    ws["D2"] = ""  # empty string
    ws["E2"] = "L" * 32000
    ws["F2"] = -1.0e308
    ws["G2"] = date(1900, 2, 28)  # serial edge before 1900-03-01
    ws["H2"] = '="a&b","c,d"'  # formula text with entities-ish and quoted comma

    # Row 3: more edge values
    ws["A3"] = False
    ws["B3"] = "ñáéíóú"
    ws["C3"] = "<>&\"'"
    ws["D3"] = ""
    ws["E3"] = "short"
    ws["F3"] = 1.0e-300
    ws["G3"] = date(1899, 12, 31) if False else date(1900, 1, 1)
    # openpyxl may not accept pre-1900 dates well; use serial via number + date format
    ws["G3"] = 0.0  # Excel serial 0 edge
    ws["G3"].number_format = "yyyy-mm-dd"
    ws["H3"] = '=CONCATENATE("x&y","a,b")'

    # Row 4: very large / small / negative
    ws["A4"] = True
    ws["F4"] = -0.0
    ws["F4"] = -123456789.123456789
    ws["G4"] = 1.0  # serial 1 = 1900-01-01 under Excel 1900 system
    ws["G4"].number_format = "yyyy-mm-dd"

    # Row 5: another bool + unicode
    ws["A5"] = False
    ws["B5"] = "零一二三"
    ws["F5"] = 2.2250738585072014e-308  # near tiny

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# 5. stress_noshared.xlsx + stress_nostyles.xlsx
# ---------------------------------------------------------------------------


def gen_noshared(path: Path) -> None:
    """Workbook whose values are inline strings / numbers (sharedStrings optional).

    openpyxl may still emit sharedStrings for some strings; we strip the part
    and rewrite sheet cells to inlineStr / numbers via zip surgery so the
    package truly has no xl/sharedStrings.xml.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "NoShared"
    ws["A1"] = "h1"
    ws["B1"] = "h2"
    ws["A2"] = 11
    ws["B2"] = 22
    ws["A3"] = 33
    ws["B3"] = 44
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    raw = buf.getvalue()

    # Rebuild zip without sharedStrings; force sheet to use inline / numeric only.
    out = BytesIO()
    with zipfile.ZipFile(BytesIO(raw), "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            name = item.filename
            if name == "xl/sharedStrings.xml":
                continue
            if name == "xl/workbook.xml.rels" or name == "xl/_rels/workbook.xml.rels":
                # drop sharedStrings relationship
                text = data.decode("utf-8")
                # remove Relationship that targets sharedStrings
                parts = text.split("<Relationship ")
                kept = [parts[0]]
                for p in parts[1:]:
                    if "sharedStrings" in p:
                        continue
                    kept.append("<Relationship " + p)
                data = "".join(kept).encode("utf-8")
            if name == "[Content_Types].xml":
                text = data.decode("utf-8")
                # remove Override for sharedStrings
                lines = text.split(">")
                rebuilt = []
                for ln in lines:
                    if "sharedStrings" in ln:
                        continue
                    rebuilt.append(ln)
                data = ">".join(rebuilt).encode("utf-8")
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                # Replace any t="s" cells with inline numbers we know; sheet is simple.
                data = _sheet_without_shared(data)
            zout.writestr(item, data)
    path.write_bytes(out.getvalue())


def _sheet_without_shared(sheet_xml: bytes) -> bytes:
    """Ensure no t=\"s\" cells remain; convert to numbers/inlineStr if needed."""
    text = sheet_xml.decode("utf-8")
    # Our generator used only ints + short headers; openpyxl may use shared for headers.
    # Rewrite a known-good sheetData.
    sheet_data = """<sheetData>
<row r="1">
<c r="A1" t="inlineStr"><is><t>h1</t></is></c>
<c r="B1" t="inlineStr"><is><t>h2</t></is></c>
</row>
<row r="2">
<c r="A2" t="n"><v>11</v></c>
<c r="B2" t="n"><v>22</v></c>
</row>
<row r="3">
<c r="A3" t="n"><v>33</v></c>
<c r="B3" t="n"><v>44</v></c>
</row>
</sheetData>"""
    import re

    text2, n = re.subn(
        r"<sheetData[\s\S]*?</sheetData>",
        sheet_data,
        text,
        count=1,
    )
    if n == 0:
        text2, n = re.subn(
            r"<sheetData[^/]*/>",
            sheet_data,
            text,
            count=1,
        )
    return text2.encode("utf-8")


def gen_nostyles(path: Path) -> None:
    """Minimal package with no styles.xml; values still readable; xf defaults to 0."""
    # Build minimal xlsx by hand.
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""
    wb_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="NoStyles" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>"""
    wb_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>"""
    sheet = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" t="inlineStr"><is><t>h1</t></is></c>
      <c r="B1" t="inlineStr"><is><t>h2</t></is></c>
    </row>
    <row r="2">
      <c r="A2" t="n"><v>1</v></c>
      <c r="B2" t="n"><v>2</v></c>
    </row>
    <row r="3">
      <c r="A3" t="n"><v>3</v></c>
      <c r="C3" t="n"><v>9</v></c>
    </row>
  </sheetData>
</worksheet>"""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("xl/workbook.xml", wb_xml)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/worksheets/sheet1.xml", sheet)
        # deliberately no xl/styles.xml, no sharedStrings


def main() -> None:
    _ensure_dir()
    targets = {
        "stress_sparse.xlsx": gen_sparse,
        "stress_multisheet.xlsx": gen_multisheet,
        "stress_onecell_everything.xlsx": gen_onecell_everything,
        "stress_types.xlsx": gen_types,
        "stress_noshared.xlsx": gen_noshared,
        "stress_nostyles.xlsx": gen_nostyles,
    }
    for name, fn in targets.items():
        p = TESTDATA / name
        print(f"writing {p} ...")
        fn(p)
        print(f"  -> {p.stat().st_size} bytes")
    print("done.")


if __name__ == "__main__":
    main()
