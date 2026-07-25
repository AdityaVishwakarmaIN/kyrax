import tempfile
from pathlib import Path
import kyrax

try:
    import pyarrow as pa
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False

import pytest

@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_arrow_table_export():
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "arrow_table.xlsx"

        table = pa.Table.from_arrays(
            [
                pa.array([1.0, 2.0, 3.0]),
                pa.array(["A", "B", "C"]),
                pa.array([True, False, True]),
            ],
            names=["col1", "col2", "col3"],
        )

        kyrax.write_excel_turbo(
            str(output_path),
            [{"name": "Sheet1", "columns": table}],
        )
        assert output_path.exists()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Sheet1"]
            assert ws.cell(row=1, column=1).value == 1.0
            assert ws.cell(row=1, column=2).value == "A"
        except ImportError:
            pass
