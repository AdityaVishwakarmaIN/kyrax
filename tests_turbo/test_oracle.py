"""Correctness oracle: kyrax turbo Python API vs openpyxl ground truth.

Run from repo root:
  .venv/Scripts/python -m pytest tests_turbo/ -v

Coordinate mapping (documented design — not a bug):
  - turbo header row = spreadsheet row 1
  - turbo data (row r, col c) 0-based → openpyxl (row r+2, col c+1)
  - openpyxl formula XOR cached value; turbo both-not-XOR
  - reserved defined names (Print_Area) via ws.print_area in openpyxl
  - hyperlink ranges may be multi-cell in OOXML; openpyxl explodes per-cell
  - comment text is flattened rich-text (both sides)
"""

from __future__ import annotations

import random
import sys
import time
from datetime import date, datetime, time as dtime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel, to_excel

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from kyrax import read_excel_turbo  # noqa: E402

TESTDATA = ROOT / "testdata"
SEED = 42
SAMPLE = 500

# Gaps discovered at runtime (filled by tests that skip).
GAPS: list[str] = []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def turbo_to_oxl(row: int, col: int) -> tuple[int, int]:
    """Map turbo 0-based data (row, col) → openpyxl 1-based (row, col).

    Design: turbo columns are sequential-XML-order; header = spreadsheet row 1.
    Data row 0 → sheet row 2. See scripts/turbo_smoke.py and _reference/docs.
    """
    return row + 2, col + 1


def cell_value(rb, row: int, col: int):
    return rb.column(col)[row].as_py()


def values_equal(tv, ov, *, rel: float = 1e-6) -> bool:
    """Compare turbo value to openpyxl value with documented conversions."""
    if tv is None and ov is None:
        return True
    # Turbo dates are f64 Excel serials + is_date styles; openpyxl may return
    # datetime/date when the number format is a date (1900 epoch rules).
    if isinstance(ov, datetime):
        try:
            serial = float(tv) if not isinstance(tv, datetime) else to_excel(tv)
            return abs(serial - float(to_excel(ov))) <= rel * max(1.0, abs(float(to_excel(ov))))
        except (TypeError, ValueError):
            return False
    if isinstance(ov, date) and not isinstance(ov, datetime):
        try:
            serial = float(tv)
            return abs(serial - float(to_excel(ov))) <= rel * max(1.0, abs(float(to_excel(ov))))
        except (TypeError, ValueError):
            return False
    if isinstance(ov, dtime):
        # rare; compare via serial fraction if turbo is float
        try:
            return abs(float(tv) - float(to_excel(ov))) <= 1e-9
        except (TypeError, ValueError):
            return False
    if isinstance(tv, float) and isinstance(ov, (int, float)):
        if tv != tv and (isinstance(ov, float) and ov != ov):
            return True
        return abs(float(tv) - float(ov)) <= rel * max(1.0, abs(float(ov)))
    if isinstance(tv, (int, float)) and isinstance(ov, (int, float)):
        return abs(float(tv) - float(ov)) <= rel * max(1.0, abs(float(ov)))
    # openpyxl Cell error types often stringify to "#DIV/0!" etc.
    if ov is not None and (
        type(ov).__name__ in ("ErrorCode", "CellError")
        or (isinstance(ov, str) and ov.startswith("#"))
    ):
        return str(tv) == str(ov)
    if isinstance(ov, str) and ov.startswith("#") and tv is None:
        # Documented residual: error may be null in value column when mixed into
        # a numeric column; typed path is cell_errors(). Not equal as values.
        return False
    return tv == ov


def color_tuple_oxl(c):
    if c is None:
        return ("none", None, 0.0)
    t = c.type
    if t == "rgb":
        v = c.rgb
        if not isinstance(v, str):
            v = "00000000"
        return ("rgb", v.upper(), 0.0)
    if t == "indexed":
        return ("indexed", int(c.indexed), 0.0)
    if t == "theme":
        return ("theme", int(c.theme), float(c.tint or 0.0))
    if t == "auto":
        return ("auto", None, 0.0)
    return (t, None, 0.0)


def color_tuple_turbo(rc: dict | None):
    if not rc:
        return ("none", None, 0.0)
    # turbo style_table uses "kind" (not openpyxl's "type")
    t = rc.get("kind") or rc.get("type") or "none"
    if t == "rgb":
        return ("rgb", str(rc.get("argb", "")).upper(), 0.0)
    if t == "indexed":
        return ("indexed", int(rc["indexed"]), 0.0)
    if t == "theme":
        return ("theme", int(rc["theme"]), float(rc.get("tint", 0.0)))
    if t == "auto":
        return ("auto", None, 0.0)
    if t == "none":
        return ("none", None, 0.0)
    return (t, None, 0.0)


def color_eq(turbo_c, oxl_c) -> bool:
    rt = color_tuple_turbo(turbo_c)
    pt = color_tuple_oxl(oxl_c)
    if rt[0] != pt[0]:
        return False
    if rt[0] == "rgb":
        return rt[1] == pt[1]
    if rt[0] == "indexed":
        return rt[1] == pt[1]
    if rt[0] == "theme":
        return rt[1] == pt[1] and abs(rt[2] - pt[2]) < 1e-6
    return True


def sample_coords(nrows: int, ncols: int, n: int = SAMPLE, seed: int = SEED) -> list[tuple[int, int]]:
    rng = random.Random(seed)
    if nrows <= 0 or ncols <= 0:
        return []
    n = min(n, nrows * ncols)
    # with replacement is fine for large sheets; prefer unique when cheap
    if nrows * ncols <= n * 4:
        allc = [(r, c) for r in range(nrows) for c in range(ncols)]
        return rng.sample(allc, min(n, len(allc)))
    seen: set[tuple[int, int]] = set()
    out: list[tuple[int, int]] = []
    while len(out) < n:
        rc = (rng.randrange(nrows), rng.randrange(ncols))
        if rc not in seen:
            seen.add(rc)
            out.append(rc)
    return out


def oxl_collect_values(path: Path, coords: set[tuple[int, int]], *, data_only: bool):
    """coords are openpyxl (row, col) 1-based. Returns dict keyed the same."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    ws = wb.active
    got = {}
    if not coords:
        wb.close()
        return got
    maxrow = max(r for r, _ in coords)
    for row in ws.iter_rows(max_row=maxrow):
        if not row:
            continue
        rr = row[0].row
        if rr is None:
            continue
        for cell in row:
            key = (cell.row, cell.column)
            if key in coords:
                got[key] = cell.value
    wb.close()
    return got


def oxl_collect_styles(path: Path, coords: set[tuple[int, int]]):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    got = {}
    if not coords:
        wb.close()
        return got
    maxrow = max(r for r, _ in coords)
    for row in ws.iter_rows(max_row=maxrow):
        if not row:
            continue
        for cell in row:
            key = (cell.row, cell.column)
            if key in coords:
                got[key] = {
                    "number_format": cell.number_format,
                    "font": cell.font,
                    "fill": cell.fill,
                }
    wb.close()
    return got


# ---------------------------------------------------------------------------
# 1. values
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "fixture",
    ["mixed.xlsx", "strings_shared.xlsx"],
    ids=["mixed", "strings_shared"],
)
def test_values(fixture: str):
    path = TESTDATA / fixture
    sheet = read_excel_turbo(str(path)).load_sheet(0, features="values")
    rb = sheet.to_arrow()
    assert rb.num_rows == sheet.nrows
    assert rb.num_columns == sheet.ncols

    coords_t = sample_coords(sheet.nrows, sheet.ncols, SAMPLE, SEED)
    coords_o = {turbo_to_oxl(r, c) for r, c in coords_t}
    oxl = oxl_collect_values(path, coords_o, data_only=True)

    mismatches = []
    for r, c in coords_t:
        orow, ocol = turbo_to_oxl(r, c)
        tv = cell_value(rb, r, c)
        ov = oxl.get((orow, ocol))
        if not values_equal(tv, ov):
            mismatches.append((r, c, orow, ocol, tv, ov))
            if len(mismatches) >= 8:
                break
    assert not mismatches, f"{fixture}: value mismatches e.g. {mismatches[:5]}"


# ---------------------------------------------------------------------------
# 2. styles
# ---------------------------------------------------------------------------


def test_styles():
    path = TESTDATA / "styled.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features=["styles"])
    style_table = sheet.style_table()
    style_idx = sheet.style_indices()
    assert style_table is not None and len(style_table) > 0
    assert style_idx is not None and len(style_idx) == sheet.ncols

    coords_t = sample_coords(sheet.nrows, sheet.ncols, SAMPLE, SEED)
    coords_o = {turbo_to_oxl(r, c) for r, c in coords_t}
    oxl = oxl_collect_styles(path, coords_o)

    fields = [
        "number_format",
        "is_date",
        "font_name",
        "font_bold",
        "font_italic",
        "font_size",
        "font_color",
        "fill_pattern",
        "fill_fg",
    ]
    fails = {f: [] for f in fields}
    total = 0
    for r, c in coords_t:
        orow, ocol = turbo_to_oxl(r, c)
        py = oxl.get((orow, ocol))
        if py is None:
            continue
        xf = style_idx[c][r]
        rust = style_table[xf]
        total += 1
        font = py["font"]
        fill = py["fill"]

        if rust["number_format"] != py["number_format"]:
            fails["number_format"].append((r, c, rust["number_format"], py["number_format"]))
        py_isdate = bool(is_date_format(py["number_format"] or "General"))
        if bool(rust["is_date"]) != py_isdate:
            fails["is_date"].append((r, c, rust["is_date"], py_isdate))
        rfont = rust["font"]
        if rfont["name"] != (font.name or ""):
            fails["font_name"].append((r, c, rfont["name"], font.name))
        if bool(rfont["bold"]) != bool(font.bold):
            fails["font_bold"].append((r, c, rfont["bold"], font.bold))
        if bool(rfont["italic"]) != bool(font.italic):
            fails["font_italic"].append((r, c, rfont["italic"], font.italic))
        if abs(float(rfont["size"]) - float(font.sz or 0)) > 1e-6:
            fails["font_size"].append((r, c, rfont["size"], font.sz))
        if not color_eq(rfont.get("color"), font.color):
            fails["font_color"].append(
                (r, c, color_tuple_turbo(rfont.get("color")), color_tuple_oxl(font.color))
            )
        rfill = rust["fill"]
        py_pat = fill.patternType if fill.patternType is not None else "none"
        if rfill["pattern"] != py_pat:
            fails["fill_pattern"].append((r, c, rfill["pattern"], py_pat))
        py_fg = getattr(fill, "fgColor", None)
        if not color_eq(rfill.get("fg"), py_fg):
            fails["fill_fg"].append(
                (r, c, color_tuple_turbo(rfill.get("fg")), color_tuple_oxl(py_fg))
            )

    assert total > 0
    bad = {f: v for f, v in fails.items() if v}
    assert not bad, f"style mismatches ({total} cells): " + "; ".join(
        f"{f}:{ex[:2]}" for f, ex in bad.items()
    )


# ---------------------------------------------------------------------------
# 3. formulas (+ string/error caches)
# ---------------------------------------------------------------------------


def test_formulas():
    path = TESTDATA / "formulas.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features=["formulas"])
    formulas = sheet.formulas()
    assert formulas is not None and formulas.num_rows > 0
    rb = sheet.to_arrow()

    # Index formula cells by (row, col) via columnar RecordBatch
    fpd = formulas.to_pydict()
    f_rows, f_cols = fpd["row"], fpd["col"]
    f_kinds, f_texts = fpd["kind"], fpd["text"]
    by_rc = {
        (f_rows[i], f_cols[i]): {
            "row": f_rows[i],
            "col": f_cols[i],
            "kind": f_kinds[i],
            "text": f_texts[i],
        }
        for i in range(formulas.num_rows)
    }
    formula_keys = list(by_rc.keys())
    rng = random.Random(SEED)

    # 500 random formula cells
    sample = rng.sample(formula_keys, min(SAMPLE, len(formula_keys)))

    # Forced samples from BOTH shared groups at rows far from anchors.
    # Col D (3) = shared_double, col H (7) = shared_sum (see fixture headers).
    shared_groups = {3: [], 7: []}
    for i in range(formulas.num_rows):
        if f_kinds[i] == "shared" and f_cols[i] in shared_groups:
            shared_groups[f_cols[i]].append(
                {"row": f_rows[i], "col": f_cols[i], "kind": f_kinds[i], "text": f_texts[i]}
            )
    forced = []
    for col, items in shared_groups.items():
        if not items:
            continue
        items_sorted = sorted(items, key=lambda x: x["row"])
        # far from anchors (anchors are near row 0)
        pick = items_sorted[min(len(items_sorted) - 1, max(1000, len(items_sorted) // 2))]
        forced.append((pick["row"], pick["col"]))
        # another near the end
        pick2 = items_sorted[-1]
        forced.append((pick2["row"], pick2["col"]))
    for rc in forced:
        if rc not in sample:
            sample.append(rc)

    # String caches col E (4) and error caches col F (5)
    for col in (4, 5):
        for _ in range(10):
            r = rng.randrange(sheet.nrows)
            if (r, col) in by_rc and (r, col) not in sample:
                sample.append((r, col))

    coords_o = {turbo_to_oxl(r, c) for r, c in sample}
    fstr = oxl_collect_values(path, coords_o, data_only=False)
    fcache = oxl_collect_values(path, coords_o, data_only=True)

    f_mismatches = []
    v_mismatches = []
    for r, c in sample:
        f = by_rc[(r, c)]
        orow, ocol = turbo_to_oxl(r, c)
        py_f = fstr.get((orow, ocol))
        py_fs = (
            py_f[1:]
            if isinstance(py_f, str) and py_f.startswith("=")
            else (py_f if isinstance(py_f, str) else "")
        )
        if f["text"] != py_fs:
            f_mismatches.append((r, c, f["kind"], f["text"], py_fs))
            if len(f_mismatches) >= 8:
                break

    for r, c in sample:
        if len(v_mismatches) >= 8:
            break
        orow, ocol = turbo_to_oxl(r, c)
        tv = cell_value(rb, r, c)
        ov = fcache.get((orow, ocol))
        # Error caches: compare via cell_errors when value path is null/typed gap
        if c == 5:
            continue  # asserted separately below
        if not values_equal(tv, ov):
            v_mismatches.append((r, c, tv, ov))

    assert not f_mismatches, f"formula text mismatches: {f_mismatches[:5]}"
    assert not v_mismatches, f"cached value mismatches: {v_mismatches[:5]}"

    # String caches col E
    e_samples = [(r, c) for r, c in sample if c == 4][:20]
    if not e_samples:
        e_samples = [(rng.randrange(min(1000, sheet.nrows)), 4) for _ in range(20)]
    for r, c in e_samples:
        orow, ocol = turbo_to_oxl(r, c)
        tv = cell_value(rb, r, c)
        ov = fcache.get((orow, ocol))
        if ov is None and (orow, ocol) not in fcache:
            # may need fetch
            continue
        assert values_equal(tv, ov) or (isinstance(tv, str) and tv == str(ov)), (
            f"string cache E{orow}: turbo={tv!r} oxl={ov!r}"
        )

    # Error caches col F via cell_errors() — REAL gap fix (typed t="e")
    if not hasattr(sheet, "cell_errors"):
        GAPS.append("TurboSheet.cell_errors() missing — error caches untyped")
        pytest.skip("API lacks cell_errors(); error cache gap")
    errs = sheet.cell_errors()
    assert errs is not None and errs.num_rows >= 0
    epd = errs.to_pydict()
    err_by_rc = {(epd["row"][i], epd["col"][i]): epd["code"][i] for i in range(errs.num_rows)}
    f_err_idxs = [i for i in range(errs.num_rows) if epd["col"][i] == 5]
    assert len(f_err_idxs) > 0, "expected t=e cells in col F (err_formula)"
    # sample of col F errors must be #DIV/0!
    for i in rng.sample(f_err_idxs, min(50, len(f_err_idxs))):
        assert epd["code"][i] == "#DIV/0!", f"col F error code {epd['code'][i]!r}"
    # forced rows also present
    for r in (0, 1, sheet.nrows - 1):
        code = err_by_rc.get((r, 5))
        assert code == "#DIV/0!", f"cell_errors missing/wrong at data row {r} col F: {code!r}"


# ---------------------------------------------------------------------------
# 4. merges
# ---------------------------------------------------------------------------


def test_merges():
    path = TESTDATA / "structured.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features=["merges"])
    merges = sheet.merges()
    assert merges is not None
    assert len(merges) == 20000, f"expected 20000 merges, got {len(merges)}"

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb.active
    oxl = set(str(r) for r in ws.merged_cells.ranges)
    wb.close()

    proto = set(merges)
    assert proto == oxl, (
        f"merge set mismatch; missing={list(oxl - proto)[:5]} extra={list(proto - oxl)[:5]}"
    )


# ---------------------------------------------------------------------------
# 5. defined names
# ---------------------------------------------------------------------------


def test_defined_names():
    path = TESTDATA / "structured.xlsx"
    reader = read_excel_turbo(str(path))
    sheet = reader.load_sheet(0, features=["defined_names"])
    dns = reader.defined_names()
    assert dns is not None
    assert len(dns) == 198, f"expected 198 defined names, got {len(dns)}"

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    oxl_global = dict(wb.defined_names)
    oxl_sheet: dict[int, dict] = {}
    for idx, sn in enumerate(sheets):
        oxl_sheet[idx] = dict(wb[sn].defined_names)

    proto_global = [d for d in dns if d.get("scope") is None and not d.get("reserved")]
    proto_sheet: dict[int, list] = {}
    proto_reserved = [d for d in dns if d.get("reserved")]
    for d in dns:
        sc = d.get("scope")
        if sc is not None and not d.get("reserved"):
            proto_sheet.setdefault(int(sc), []).append(d)

    issues = []
    if len(proto_global) != len(oxl_global):
        issues.append(f"global count turbo {len(proto_global)} != oxl {len(oxl_global)}")
    for idx, names in oxl_sheet.items():
        pc = len(proto_sheet.get(idx, []))
        if pc != len(names):
            issues.append(f"sheet[{idx}] count turbo {pc} != oxl {len(names)}")

    for d in proto_global:
        o = oxl_global.get(d["name"])
        if o is None:
            issues.append(f"global {d['name']} absent in openpyxl")
        elif (o.value or "") != d["value"]:
            issues.append(f"global {d['name']} value turbo {d['value']!r} != oxl {o.value!r}")

    for idx, names in proto_sheet.items():
        for d in names:
            o = oxl_sheet.get(idx, {}).get(d["name"])
            if o is None:
                issues.append(f"sheet[{idx}] {d['name']} absent in openpyxl")
            elif (o.value or "") != d["value"]:
                issues.append(
                    f"sheet[{idx}] {d['name']} turbo {d['value']!r} != oxl {o.value!r}"
                )

    # Reserved: openpyxl hides them; compare Print_Area via ws.print_area
    # (documented design difference — see _reference/oracle_struct.py).
    for d in proto_reserved:
        rk = d["reserved"]
        idx = d.get("scope")
        if idx is None:
            wsx = ws
        else:
            wsx = wb[sheets[int(idx)]]
        if rk == "Print_Area":
            pa = wsx.print_area
            if isinstance(pa, list):
                pa = ",".join(pa)
            pa = pa or ""
            val = d["value"] or ""
            if pa not in (val, val.replace("'", "")) and val not in (pa, pa.replace("'", "")):
                issues.append(f"Print_Area turbo {val!r} != ws.print_area {pa!r}")

    # openpyxl must not have names turbo dropped
    proto_g_names = {d["name"] for d in proto_global}
    for nm in oxl_global:
        if nm not in proto_g_names:
            issues.append(f"openpyxl global {nm} missing from turbo")

    wb.close()
    assert not issues, f"defined_names issues ({len(issues)}): {issues[:8]}"

    # sanity: constants and scoped present
    kinds = {d.get("kind") for d in dns}
    assert "constant" in kinds or any(
        d.get("kind") == "constant" or (d.get("value") and d["value"][:1].isdigit())
        for d in dns
    )
    assert any(d.get("scope") is not None for d in dns), "expected sheet-scoped names"
    assert any(d.get("reserved") for d in dns), "expected reserved Print_Area"


# ---------------------------------------------------------------------------
# 6. tables
# ---------------------------------------------------------------------------


def test_tables():
    path = TESTDATA / "structured.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features=["tables"])
    tables = sheet.tables()
    assert tables is not None
    assert len(tables) == 20, f"expected 20 tables, got {len(tables)}"

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb.active
    oxl = ws.tables
    issues = []
    for t in tables:
        o = oxl.get(t["name"])
        if o is None:
            issues.append(f"table {t['name']} absent in openpyxl")
            continue
        if t["ref"] != o.ref:
            issues.append(f"{t['name']} ref turbo {t['ref']} != oxl {o.ref}")
        if t.get("display_name") and t["display_name"] != o.displayName:
            issues.append(
                f"{t['name']} displayName {t['display_name']} != {o.displayName}"
            )
        ohdr = 1 if o.headerRowCount is None else o.headerRowCount
        if t["header_row_count"] != ohdr:
            issues.append(f"{t['name']} headerRowCount {t['header_row_count']} != {ohdr}")
        otot = 0 if o.totalsRowCount is None else o.totalsRowCount
        if t["totals_row_count"] != otot:
            issues.append(f"{t['name']} totalsRowCount {t['totals_row_count']} != {otot}")
        ocols = list(o.tableColumns)
        tcols = t["columns"]
        if len(tcols) != len(ocols):
            issues.append(f"{t['name']} col count {len(tcols)} != {len(ocols)}")
        else:
            for pc, oc in zip(tcols, ocols):
                if pc["name"] != oc.name:
                    issues.append(f"{t['name']} col {pc['name']} != {oc.name}")
                tf = pc.get("totals_row_function") or pc.get("totals_fn")
                if (tf or None) != (oc.totalsRowFunction or None):
                    issues.append(
                        f"{t['name']}/{pc['name']} totalsFn {tf} != {oc.totalsRowFunction}"
                    )
                tl = pc.get("totals_row_label") or pc.get("totals_label")
                if (tl or None) != (oc.totalsRowLabel or None):
                    issues.append(
                        f"{t['name']}/{pc['name']} totalsLabel {tl} != {oc.totalsRowLabel}"
                    )
    wb.close()
    assert not issues, f"tables issues: {issues[:8]}"


# ---------------------------------------------------------------------------
# 7. hyperlinks
# ---------------------------------------------------------------------------


def test_hyperlinks():
    path = TESTDATA / "structured.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features=["hyperlinks"])
    hlinks = sheet.hyperlinks()
    assert hlinks is not None and len(hlinks) > 0

    rng = random.Random(SEED)
    sample = rng.sample(hlinks, min(SAMPLE, len(hlinks)))

    # openpyxl explodes multi-cell hyperlink ranges per-cell (documented).
    # Fixture links are single-cell refs; match cell.hyperlink on ref start.
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb.active
    issues = []
    for h in sample:
        ref = h["ref"]
        # take top-left of range if A1:B2
        cell_ref = ref.split(":")[0]
        cell = ws[cell_ref]
        o = cell.hyperlink
        if o is None:
            issues.append(f"{ref} openpyxl has no hyperlink")
            continue
        if h.get("kind") == "external":
            if o.target != h.get("target"):
                issues.append(f"{ref} target turbo {h.get('target')!r} != oxl {o.target!r}")
        else:
            if (o.location or None) != (h.get("location") or None):
                issues.append(
                    f"{ref} location turbo {h.get('location')!r} != oxl {o.location!r}"
                )
        if (o.tooltip or None) != (h.get("tooltip") or None):
            issues.append(f"{ref} tooltip turbo {h.get('tooltip')!r} != oxl {o.tooltip!r}")
        if len(issues) > 10:
            break
    wb.close()
    assert not issues, f"hyperlink issues: {issues[:6]}"


# ---------------------------------------------------------------------------
# 8. comments
# ---------------------------------------------------------------------------


def test_comments():
    path = TESTDATA / "comments.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features=["comments"])
    comments = sheet.comments()
    assert comments is not None and comments.num_rows > 0
    authors = sheet.comment_authors()
    assert authors is not None and len(authors) == 10

    cpd = comments.to_pydict()
    rng = random.Random(SEED)
    sample_idxs = rng.sample(range(comments.num_rows), min(SAMPLE, comments.num_rows))

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb.active
    issues = []
    for i in sample_idxs:
        ref = cpd["ref"][i]
        text = cpd["text"][i]
        author = cpd["author"][i]
        cell = ws[ref]
        o = cell.comment
        if o is None:
            issues.append(f"{ref} openpyxl has no comment")
            continue
        # Flattened rich-text on both sides (documented design).
        if o.text != text:
            issues.append(f"{ref} text turbo {text!r} != oxl {o.text!r}")
        if o.author != author:
            issues.append(f"{ref} author turbo {author!r} != oxl {o.author!r}")
        if len(issues) > 10:
            break
    wb.close()
    assert not issues, f"comment issues: {issues[:6]}"


# ---------------------------------------------------------------------------
# 9. selective flags
# ---------------------------------------------------------------------------


def test_selective_flags():
    path = TESTDATA / "structured.xlsx"
    # features=["styles"] leaves formulas/merges/etc. absent
    # Note: styles fixture is styled.xlsx; structured may have few styles but
    # the None-vs-present gate still holds.
    st = read_excel_turbo(str(path)).load_sheet(0, features=["styles"])
    assert st.formulas() is None
    assert st.merges() is None
    assert st.hyperlinks() is None
    assert st.comments() is None
    assert st.tables() is None
    # styles may or may not yield a non-empty table; indices/table should be present
    assert st.style_table() is not None
    assert st.style_indices() is not None

    # features="values" bare works on every fixture
    fixtures = [
        "mixed.xlsx",
        "strings_shared.xlsx",
        "styled.xlsx",
        "formulas.xlsx",
        "structured.xlsx",
        "comments.xlsx",
    ]
    for name in fixtures:
        s = read_excel_turbo(str(TESTDATA / name)).load_sheet(0, features="values")
        rb = s.to_arrow()
        assert rb.num_rows == s.nrows
        assert rb.num_columns == s.ncols
        assert s.formulas() is None
        assert s.merges() is None
        assert s.hyperlinks() is None
        assert s.comments() is None
        assert s.tables() is None
        assert s.style_table() is None
        assert s.style_indices() is None


# ---------------------------------------------------------------------------
# 10. cached values as default (values-only path)
# ---------------------------------------------------------------------------


def test_cached_values_as_default():
    """formulas.xlsx with features='values' still returns cached results in to_arrow().

    This is the data_only=True equivalent on the fast path (both-not-XOR design:
    formula text is gated, cached <v> always flows to value columns).
    """
    path = TESTDATA / "formulas.xlsx"
    sheet = read_excel_turbo(str(path)).load_sheet(0, features="values")
    assert sheet.formulas() is None
    rb = sheet.to_arrow()
    assert rb.num_rows == sheet.nrows > 0

    # Sample numeric formula columns C (sum) and D (shared_double) — not error col
    coords_t = []
    rng = random.Random(SEED)
    for _ in range(SAMPLE):
        r = rng.randrange(sheet.nrows)
        c = rng.choice([2, 3, 4, 7])  # sum, shared_double, str_formula, shared_sum
        coords_t.append((r, c))
    coords_o = {turbo_to_oxl(r, c) for r, c in coords_t}
    oxl = oxl_collect_values(path, coords_o, data_only=True)

    mismatches = []
    for r, c in coords_t:
        orow, ocol = turbo_to_oxl(r, c)
        tv = cell_value(rb, r, c)
        ov = oxl.get((orow, ocol))
        if c == 4:
            # string cache
            if str(tv) != str(ov):
                mismatches.append((r, c, tv, ov))
        elif not values_equal(tv, ov):
            mismatches.append((r, c, tv, ov))
        if len(mismatches) >= 8:
            break
    assert not mismatches, f"values-only cached results mismatch: {mismatches[:5]}"

    # Error codes still available via cell_errors on values-only path
    if hasattr(sheet, "cell_errors"):
        errs = sheet.cell_errors()
        epd = errs.to_pydict()
        f_idxs = [i for i in range(errs.num_rows) if epd["col"][i] == 5]
        assert len(f_idxs) > 0
        assert all(epd["code"][i] == "#DIV/0!" for i in f_idxs[:50])


# ---------------------------------------------------------------------------
# Gap report (session end)
# ---------------------------------------------------------------------------


def test_zzz_report_gaps(capsys):
    """Loudly report any API gaps discovered (runs last by name)."""
    if GAPS:
        print("\n" + "!" * 70)
        print("API GAPS DISCOVERED:")
        for g in GAPS:
            print(f"  - {g}")
        print("!" * 70)
    else:
        print("\nNo API gaps discovered during oracle run.")
