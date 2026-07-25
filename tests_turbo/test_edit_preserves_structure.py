"""Regression: kyrax.edit_excel round-trip must preserve sheet-level structure.

Editing a single cell through the overlay must not destroy merges, column
widths, row heights, freeze panes, autofilter or hyperlinks, and must not
change the part list of the archive.
"""

import zipfile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import kyrax


def _build_source(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for r in range(1, 21):
        for c in range(1, 5):
            ws.cell(row=r, column=c, value=float(r * c))

    ws.merge_cells("B2:C3")
    ws.column_dimensions["A"].width = 42.0
    ws.row_dimensions[5].height = 33.0
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:D20"
    ws["D1"].hyperlink = "https://example.com/target"
    ws["A1"].font = Font(bold=True, color="FFFF0000")

    wb.save(path)
    wb.close()


@pytest.fixture()
def roundtrip(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _build_source(str(src))

    wb = kyrax.edit_excel(str(src))
    wb["Data"].set_cell(1, 1, 12345.0)
    wb.save(str(out))

    return src, out


def test_edit_preserves_sheet_structure(roundtrip):
    src, out = roundtrip
    wb = load_workbook(str(out))
    ws = wb["Data"]

    # 1. edited cell + untouched cells
    assert ws["A1"].value == 12345.0
    assert ws["B10"].value == 20.0
    assert ws["D20"].value == 80.0

    # 2. structures that used to be destroyed
    assert [str(m) for m in ws.merged_cells.ranges] == ["B2:C3"]
    assert ws.column_dimensions["A"].width == 42.0
    assert ws.row_dimensions[5].height == 33.0
    assert ws.freeze_panes == "B2"
    assert ws.auto_filter.ref == "A1:D20"
    assert ws["D1"].hyperlink is not None
    assert ws["D1"].hyperlink.target == "https://example.com/target"

    # 3. per-cell style indices still resolve against pass-through styles.xml
    assert ws["A1"].font.bold is True

    wb.close()


def test_edit_preserves_part_list(roundtrip):
    src, out = roundtrip
    with zipfile.ZipFile(str(src)) as zsrc, zipfile.ZipFile(str(out)) as zout:
        assert zout.testzip() is None
        assert sorted(zsrc.namelist()) == sorted(zout.namelist())


def test_edit_string_and_new_cells(tmp_path):
    """String edits, edits outside the used range, and new rows."""
    src = tmp_path / "src2.xlsx"
    out = tmp_path / "out2.xlsx"
    _build_source(str(src))

    wb = kyrax.edit_excel(str(src))
    sh = wb["Data"]
    sh.set_cell(1, 1, "hello world")   # replace existing, keeps s= index
    sh.set_cell(21, 1, 99.0)           # brand-new row past the last one
    sh.set_cell(3, 8, "col H")         # new column past the used range
    wb.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["A1"].value == "hello world"
    assert ws["A1"].font.bold is True
    assert ws["A21"].value == 99.0
    assert ws["H3"].value == "col H"
    assert ws["B10"].value == 20.0
    assert [str(m) for m in ws.merged_cells.ranges] == ["B2:C3"]
    assert ws.freeze_panes == "B2"
    ld.close()
