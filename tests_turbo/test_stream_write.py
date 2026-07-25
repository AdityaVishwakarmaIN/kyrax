"""Streaming export (openpyxl write_only analogue) must match the default writer."""

import os
import zipfile

import openpyxl
import pytest

import kyrax


def _sheets():
    return [
        {
            "name": "Alpha",
            "columns": [
                [float(i) for i in range(500)],
                [f"row_{i}" for i in range(500)],
                [i % 2 == 0 for i in range(500)],
            ],
        },
        {"name": "Beta", "columns": [[1.5, 2.5], ["x", "y"]]},
    ]


def test_stream_writer_is_exported():
    assert hasattr(kyrax, "write_excel_turbo_stream"), "streaming API not exported"


def test_stream_output_opens_and_matches_values(tmp_path):
    out = str(tmp_path / "stream.xlsx")
    kyrax.write_excel_turbo_stream(out, _sheets())

    assert zipfile.is_zipfile(out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Alpha", "Beta"]

    ws = wb["Alpha"]
    assert ws.cell(row=1, column=1).value == 0.0
    assert ws.cell(row=1, column=2).value == "row_0"
    assert ws.cell(row=1, column=3).value is True
    assert ws.cell(row=500, column=1).value == 499.0
    assert ws.cell(row=500, column=2).value == "row_499"

    wsb = wb["Beta"]
    assert wsb.cell(row=1, column=1).value == 1.5
    assert wsb.cell(row=2, column=2).value == "y"


def test_stream_matches_default_writer_cell_for_cell(tmp_path):
    a = str(tmp_path / "default.xlsx")
    b = str(tmp_path / "streamed.xlsx")
    kyrax.write_excel_turbo(a, _sheets())
    kyrax.write_excel_turbo_stream(b, _sheets())

    wa = openpyxl.load_workbook(a)
    wb = openpyxl.load_workbook(b)
    assert wa.sheetnames == wb.sheetnames

    for name in wa.sheetnames:
        sa, sb = wa[name], wb[name]
        assert (sa.max_row, sa.max_column) == (sb.max_row, sb.max_column), name
        for r in range(1, sa.max_row + 1):
            for c in range(1, sa.max_column + 1):
                assert sa.cell(row=r, column=c).value == sb.cell(row=r, column=c).value, (
                    f"{name}!{r},{c}"
                )


def test_stream_part_set_matches_default_writer(tmp_path):
    a = str(tmp_path / "d.xlsx")
    b = str(tmp_path / "s.xlsx")
    kyrax.write_excel_turbo(a, _sheets())
    kyrax.write_excel_turbo_stream(b, _sheets())
    with zipfile.ZipFile(a) as za, zipfile.ZipFile(b) as zb:
        assert sorted(za.namelist()) == sorted(zb.namelist())


def test_stream_large_sheet_round_trips(tmp_path):
    out = str(tmp_path / "big.xlsx")
    n = 50_000
    kyrax.write_excel_turbo_stream(
        out, [{"name": "Big", "columns": [[float(i) for i in range(n)]]}]
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb["Big"]
    assert ws.max_row == n
    rows = list(ws.iter_rows(min_row=n, max_row=n, values_only=True))
    assert rows[0][0] == float(n - 1)
    wb.close()
