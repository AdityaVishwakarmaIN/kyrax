"""Regression tests for silent data loss when a column's dtype is inferred from a sample.

`kyrax.read_excel(...).load_sheet(...)` infers each column's dtype from only the first
`schema_sample_rows` rows (1000 by default). Values further down that do not fit the inferred
dtype used to be silently converted to null at materialization time -- a float column with a
string in its tail came back as float64 with NaNs where the strings were.

The reader now detects, while building the arrow array, that a cell holding a real value had to
be nulled, and re-derives the column dtype from *every* selected row before rebuilding it. The
column comes back widened (a schema change the user is warned about) instead of silently gutted.

The boundary that matters on this code path is `schema_sample_rows`, not the turbo scanner's
byte-based rayon chunking: `read_excel` goes through the calamine-backed `src/types/` reader,
which is single-threaded and never touches `src/turbo/scan.rs`. Row counts below are therefore
sized to cross the 1000-row sampling boundary.
"""

import warnings
from pathlib import Path

import openpyxl
import pytest

import kyrax

# Default `schema_sample_rows` for `load_sheet` (see `py_load_sheet` in
# src/types/excelreader/python.rs). Every case below must exceed this for the sampled dtype to
# disagree with the column's real contents.
SCHEMA_SAMPLE_ROWS = 1000


def _write(path: Path, values, header="colA"):
    """Writes a single column (header + `values`) to a one-sheet workbook named "S"."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = header
    for offset, value in enumerate(values):
        ws.cell(row=offset + 2, column=1, value=value)
    wb.save(path)
    wb.close()


def _read(path: Path, **kwargs):
    return kyrax.read_excel(str(path)).load_sheet(0, **kwargs).to_pandas()


def test_trailing_strings_in_long_float_column_survive(tmp_path):
    """The original repro: 60000 floats followed by two strings, far past the sample window."""
    path = tmp_path / "late_string.xlsx"
    n_rows = 60000
    values = [float(r) for r in range(1, n_rows + 1)]
    values[-2:] = ["LATE_STRING", "LATE_STRING"]
    _write(path, values)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = _read(path)

    col = df["colA"]

    assert len(col) == n_rows
    # The two trailing strings must survive rather than becoming NaN.
    assert col.isna().sum() == 0
    assert list(col.tail(2)) == ["LATE_STRING", "LATE_STRING"]
    # ...and the leading numeric values must still be correct.
    assert col.iloc[0] == "1"
    assert col.iloc[100] == "101"
    assert list(col.head(3)) == ["1", "2", "3"]


def test_type_switch_just_past_the_sample_boundary(tmp_path):
    """Numeric for the sampled prefix, string afterwards -- the smallest case that reproduces.

    The switch happens a few rows after `schema_sample_rows`, so the sampled dtype is float while
    the bulk of the column is string.
    """
    path = tmp_path / "boundary.xlsx"
    numeric_rows = SCHEMA_SAMPLE_ROWS + 5
    string_rows = 200
    values = [float(r) for r in range(1, numeric_rows + 1)]
    values += [f"S{i}" for i in range(string_rows)]
    _write(path, values)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = _read(path)

    col = df["colA"]

    assert len(col) == numeric_rows + string_rows
    assert col.isna().sum() == 0
    assert col.iloc[0] == "1"
    assert col.iloc[numeric_rows - 1] == str(numeric_rows)
    assert col.iloc[numeric_rows] == "S0"
    assert col.iloc[-1] == f"S{string_rows - 1}"


def test_bools_followed_by_strings_survive(tmp_path):
    """A bool column with strings past the sample window widens to string, losing nothing."""
    path = tmp_path / "bool_string.xlsx"
    bool_rows = SCHEMA_SAMPLE_ROWS + 5
    values = [bool(r % 2) for r in range(bool_rows)]
    values += ["MAYBE", "MAYBE"]
    _write(path, values)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = _read(path)

    col = df["colA"]

    assert len(col) == bool_rows + 2
    assert col.isna().sum() == 0
    assert list(col.tail(2)) == ["MAYBE", "MAYBE"]
    assert col.iloc[0] == "false"
    assert col.iloc[1] == "true"


def test_ints_followed_by_floats_widen_to_float(tmp_path):
    """Promotion follows the coercion lattice: int + float widens to float, not to string."""
    path = tmp_path / "int_float.xlsx"
    int_rows = SCHEMA_SAMPLE_ROWS + 5
    values = [r for r in range(1, int_rows + 1)]
    values += [1.5, 2.5]
    _write(path, values)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = _read(path)

    col = df["colA"]

    assert col.dtype == "float64"
    assert col.isna().sum() == 0
    assert col.iloc[0] == 1.0
    assert list(col.tail(2)) == [1.5, 2.5]


def test_homogeneous_column_keeps_its_dtype_and_does_not_warn(tmp_path):
    """No promotion, no warning, no dtype change when the column really is homogeneous."""
    path = tmp_path / "clean.xlsx"
    n_rows = SCHEMA_SAMPLE_ROWS + 500
    _write(path, [float(r) for r in range(1, n_rows + 1)])

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        df = _read(path)

    col = df["colA"]

    assert col.dtype == "float64"
    assert col.isna().sum() == 0
    assert col.iloc[0] == 1.0
    assert col.iloc[-1] == float(n_rows)
    assert [w for w in caught if issubclass(w.category, UserWarning)] == []


def test_genuinely_empty_column_stays_null(tmp_path):
    """A column that is empty everywhere must not be widened to string by the promotion path."""
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "colA"
    ws["B1"] = "colB"
    for r in range(2, SCHEMA_SAMPLE_ROWS + 102):
        ws.cell(row=r, column=2, value=float(r))
    wb.save(path)
    wb.close()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = _read(path)

    assert df["colA"].isna().all()
    assert df["colB"].isna().sum() == 0


def test_promotion_emits_a_user_warning_naming_the_column(tmp_path):
    """The schema change is announced rather than silent."""
    path = tmp_path / "warns.xlsx"
    numeric_rows = SCHEMA_SAMPLE_ROWS + 5
    values = [float(r) for r in range(1, numeric_rows + 1)]
    values += ["LATE_STRING"]
    _write(path, values)

    with pytest.warns(UserWarning) as caught:
        df = _read(path)

    messages = [str(w.message) for w in caught]
    assert any('"colA"' in m and "float" in m and "string" in m for m in messages), messages
    assert df["colA"].iloc[-1] == "LATE_STRING"


def test_strict_coercion_reports_the_mismatch_instead_of_promoting(tmp_path):
    """Under strict coercion a value that does not fit the inferred dtype is an error."""
    path = tmp_path / "strict.xlsx"
    numeric_rows = SCHEMA_SAMPLE_ROWS + 5
    values = [float(r) for r in range(1, numeric_rows + 1)]
    values += ["LATE_STRING"]
    _write(path, values)

    sheet = kyrax.read_excel(str(path)).load_sheet(0, dtype_coercion="strict")
    with pytest.raises(Exception) as excinfo:
        sheet.to_pandas()

    assert "colA" in str(excinfo.value)


def test_explicit_dtype_is_never_second_guessed(tmp_path):
    """Promotion only rescues *inferred* dtypes.

    Asking for `float` on a column that also holds strings is a deliberate coercion request, and
    must keep nulling the values that do not fit -- silently widening it would override the user.
    """
    path = tmp_path / "explicit.xlsx"
    numeric_rows = SCHEMA_SAMPLE_ROWS + 5
    values = [float(r) for r in range(1, numeric_rows + 1)]
    values += ["LATE_STRING"]
    _write(path, values)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        df = _read(path, dtypes={"colA": "float"})

    col = df["colA"]
    assert col.dtype == "float64"
    assert col.iloc[0] == 1.0
    # The explicitly requested coercion still applies to the value that does not fit...
    assert col.isna().sum() == 1
    # ...and no promotion warning is raised, because nothing was promoted.
    assert [w for w in caught if issubclass(w.category, UserWarning)] == []
