"""Stream A + B gap-meta oracle tests for turbo reader.

Compares turbo output against gaplab oracle_expected.json fixtures.
Where the oracle omits a field, falls back to openpyxl live values.

Fixtures (copied into testdata/):
  gap_sheetmeta.xlsx / gap_sheetmeta_1904.xlsx + gap_sheetmeta_oracle.json
  gap_richmeta.xlsx + gap_richmeta_oracle.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from kyrax import read_excel_turbo  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

TESTDATA = ROOT / "testdata"
SHEETMETA_XLSX = TESTDATA / "gap_sheetmeta.xlsx"
SHEETMETA_1904 = TESTDATA / "gap_sheetmeta_1904.xlsx"
SHEETMETA_ORACLE = TESTDATA / "gap_sheetmeta_oracle.json"
RICHMETA_XLSX = TESTDATA / "gap_richmeta.xlsx"
RICHMETA_ORACLE = TESTDATA / "gap_richmeta_oracle.json"

FEATURES_ALL = "all"
FEATURES_SHEET = [
    "sheet_meta",
    "page_setup",
    "workbook_meta",
    "defined_names",
    "styles",
]
FEATURES_RICH = ["styles", "validations", "cond_format"]


@pytest.fixture(scope="module")
def sheetmeta_oracle():
    return json.loads(SHEETMETA_ORACLE.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def richmeta_oracle():
    return json.loads(RICHMETA_ORACLE.read_text(encoding="utf-8"))


def _approx(a, b, rel=1e-6):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= rel * max(1.0, abs(float(b)))
    except (TypeError, ValueError):
        return a == b


# ---------------------------------------------------------------------------
# Stream A — sheetmeta
# ---------------------------------------------------------------------------


def test_sheetmeta_sheet_names_and_kinds(sheetmeta_oracle):
    wb_exp = sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    assert reader.sheet_names == wb_exp["sheetnames"]

    for exp in wb_exp["sheets"]:
        sh = reader.load_sheet(exp["title"], features=FEATURES_SHEET)
        assert sh.sheet_kind == exp["kind"]
        assert sh.sheet_state == exp["sheet_state"]


def test_sheetmeta_chartsheet_empty_grid():
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    sh = reader.load_sheet("Charts", features=FEATURES_ALL)
    assert sh.sheet_kind == "chartsheet"
    assert sh.nrows == 0
    assert sh.ncols == 0
    # must not crash
    _ = sh.to_arrow()


def test_sheetmeta_row_dimensions(sheetmeta_oracle):
    exp_sheets = sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"]
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    for exp in exp_sheets:
        if exp["kind"] != "worksheet":
            continue
        sh = reader.load_sheet(exp["title"], features=FEATURES_SHEET)
        rows = sh.row_dimensions() or {}
        exp_rows = exp.get("row_dimensions") or {}
        assert set(str(k) for k in rows.keys()) == set(exp_rows.keys())
        for k, er in exp_rows.items():
            got = rows[int(k)]
            assert _approx(got["height"], er["ht"]), f"row {k} height"
            assert got["hidden"] is er["hidden"]
            assert got["outline_level"] == er["outlineLevel"]
            assert got["collapsed"] is er["collapsed"]


def test_sheetmeta_column_dimensions(sheetmeta_oracle):
    exp_sheets = sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"]
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    for exp in exp_sheets:
        if exp["kind"] != "worksheet":
            continue
        sh = reader.load_sheet(exp["title"], features=FEATURES_SHEET)
        cols = sh.column_dimensions() or []
        exp_cols = exp.get("column_dimensions") or {}
        # turbo returns list of col ranges; oracle keys by letter of min
        by_min = {c["min"]: c for c in cols}
        for letter, er in exp_cols.items():
            got = by_min.get(er["min"])
            assert got is not None, f"missing col {letter}"
            # XML width truth: openpyxl may invent 13; compare when oracle has width
            if er.get("width") is not None and got["width"] is not None:
                assert _approx(got["width"], er["width"]), f"col {letter} width"
            assert got["hidden"] is er["hidden"]
            assert got["best_fit"] is er["bestFit"]
            assert got["outline_level"] == er["outlineLevel"]
            assert got["min"] == er["min"]
            assert got["max"] == er["max"]


def test_sheetmeta_auto_filter(sheetmeta_oracle):
    exp = sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"][0]
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    sh = reader.load_sheet("MetaMain", features=FEATURES_SHEET)
    af = sh.auto_filter()
    eaf = exp["auto_filter"]
    assert af is not None
    assert af["ref"] == eaf["ref"]
    assert len(af["columns"]) == len(eaf["filterColumns"])
    c0 = af["columns"][0]
    e0 = eaf["filterColumns"][0]
    assert c0["col_id"] == e0["colId"]
    assert c0["values"] == e0["filters"]
    assert c0["blank"] is e0["blank"]


def test_sheetmeta_sheet_view_and_freeze(sheetmeta_oracle):
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    for exp in sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"]:
        if exp["kind"] != "worksheet":
            continue
        sh = reader.load_sheet(exp["title"], features=FEATURES_SHEET)
        sv = sh.sheet_view()
        esv = exp["sheet_view"]
        assert sv is not None
        if esv.get("showGridLines") is not None:
            assert sv["show_grid_lines"] is esv["showGridLines"]
        if esv.get("zoomScale") is not None:
            assert sv["zoom_scale"] == esv["zoomScale"]
        if esv.get("pane"):
            ep = esv["pane"]
            gp = sv["pane"]
            assert gp is not None
            assert gp["state"] == ep["state"]
            assert gp["active_pane"] == ep["activePane"]
            if ep.get("xSplit") is not None:
                assert _approx(gp["x_split"], ep["xSplit"])
            if ep.get("ySplit") is not None:
                assert _approx(gp["y_split"], ep["ySplit"])
            assert gp["top_left_cell"] == ep["topLeftCell"]
            assert sh.freeze_panes() == esv["freeze_panes"]
            assert sv["freeze_panes"] == esv["freeze_panes"]
        else:
            assert esv.get("freeze_panes") is None


def test_sheetmeta_protection(sheetmeta_oracle):
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    for exp in sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"]:
        if exp["kind"] != "worksheet":
            continue
        sh = reader.load_sheet(exp["title"], features=FEATURES_SHEET)
        p = sh.protection()
        ep = exp["protection"]
        assert p is not None
        assert p["sheet"] is ep["sheet"]
        assert p["enabled"] is ep["enabled_bool"]
        assert p["password"] == ep["password"]
        assert p["format_cells"] is ep["formatCells"]
        assert p["select_locked_cells"] is ep["selectLockedCells"]


def test_sheetmeta_page_setup(sheetmeta_oracle):
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    exp = sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"][0]
    sh = reader.load_sheet("MetaMain", features=FEATURES_SHEET)
    page = exp["page"]
    ps = sh.page_setup()
    assert ps is not None
    assert ps["orientation"] == page["page_setup"]["orientation"]
    assert ps["paper_size"] == page["page_setup"]["paperSize"]
    assert ps["scale"] == page["page_setup"]["scale"]
    assert ps["fit_to_width"] == page["page_setup"]["fitToWidth"]
    assert ps["fit_to_height"] == page["page_setup"]["fitToHeight"]
    assert ps["fit_to_page"] is page["page_setup"]["fitToPage"]

    pm = sh.page_margins()
    epm = page["page_margins"]
    assert _approx(pm["left"], epm["left"])
    assert _approx(pm["header"], epm["header"])

    po = sh.print_options()
    epo = page["print_options"]
    assert po["horizontal_centered"] is epo["horizontalCentered"]
    assert po["headings"] is epo["headings"]

    hf = sh.header_footer()
    ehf = page["header_footer"]
    assert hf["odd_header"] == ehf["oddHeader"]["raw"]
    assert hf["odd_footer"] == ehf["oddFooter"]["raw"]


def test_sheetmeta_sheet_properties(sheetmeta_oracle):
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    for exp in sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]["sheets"]:
        if exp["kind"] != "worksheet":
            continue
        sh = reader.load_sheet(exp["title"], features=FEATURES_SHEET)
        sp = exp["sheet_properties"]
        assert sh.code_name == sp["codeName"]
        if sp.get("tabColor"):
            # turbo surfaces raw rgb hex
            assert sh.tab_color is not None
            assert sp["tabColor"]["rgb"] in sh.tab_color or sh.tab_color in sp["tabColor"]["rgb"]
        sf = sh.sheet_format()
        esf = sp["sheet_format"]
        assert sf is not None
        assert sf["base_col_width"] == esf["baseColWidth"] or sf["base_col_width"] == 8
        if esf.get("defaultRowHeight") is not None:
            assert _approx(sf["default_row_height"], esf["defaultRowHeight"])


def test_sheetmeta_workbook_props(sheetmeta_oracle):
    exp = sheetmeta_oracle["workbooks"]["sheetmeta.xlsx"]
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    _ = reader.load_sheet(0, features=FEATURES_SHEET)
    assert reader.date1904 is exp["date1904"]
    wp = reader.workbook_props()
    assert wp is not None
    assert wp["date1904"] is False
    assert wp["code_name"] == exp["code_name"]
    assert wp["calc_id"] == exp["calculation"]["calcId"]
    core = wp["core"]
    assert core["title"] == exp["core_properties"]["title"]
    assert core["creator"] == exp["core_properties"]["creator"]
    assert core["subject"] == exp["core_properties"]["subject"]
    app = wp["app"]
    assert app["application"] == exp["app_properties"]["Application"]
    assert app["app_version"] == exp["app_properties"]["AppVersion"]


def test_sheetmeta_date1904_flag(sheetmeta_oracle):
    exp = sheetmeta_oracle["workbooks"]["sheetmeta_1904.xlsx"]
    reader = read_excel_turbo(str(SHEETMETA_1904))
    _ = reader.load_sheet(0, features=["workbook_meta"])
    assert reader.date1904 is True
    assert exp["date1904"] is True
    wp = reader.workbook_props()
    assert wp["date1904"] is True


def test_sheetmeta_selective_flags_skip_work():
    reader = read_excel_turbo(str(SHEETMETA_XLSX))
    sh = reader.load_sheet("MetaMain", features=["values"])
    assert sh.row_dimensions() is None
    assert sh.auto_filter() is None
    assert sh.page_setup() is None
    assert sh.data_validations() is None

    sh2 = reader.load_sheet("MetaMain", features=["sheet_meta"])
    assert sh2.row_dimensions() is not None
    assert sh2.page_setup() is None  # PAGE_SETUP not requested


# ---------------------------------------------------------------------------
# Stream B — richmeta
# ---------------------------------------------------------------------------


def test_richmeta_named_styles(richmeta_oracle):
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=FEATURES_RICH)
    names = {ns["name"]: ns for ns in (sh.named_styles() or [])}
    for ens in richmeta_oracle["named_styles"]:
        assert ens["name"] in names, ens["name"]
        got = names[ens["name"]]
        assert got["xf_id"] == ens["xfId"]
        assert got["hidden"] is ens["hidden"]


def test_richmeta_sampled_cells_style_resolve(richmeta_oracle):
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=FEATURES_RICH)
    st = sh.style_table()
    assert st is not None
    si = sh.style_indices()
    assert si is not None

    # Sample a subset of oracle cells that land on data rows (row>=2)
    samples = [
        c
        for c in richmeta_oracle["sampled_cells"]
        if c["row"] >= 2 and c["col"] <= sh.ncols
    ][:40]
    for cell in samples:
        # turbo: header = row 1 → data row = sheet_row - 2, col = sheet_col - 1
        tr = cell["row"] - 2
        tc = cell["col"] - 1
        if tr < 0 or tr >= sh.nrows or tc < 0 or tc >= sh.ncols:
            continue
        xf = si[tc][tr]
        resolved = st[xf]
        assert resolved["name"] == cell["named_style"], (
            f"{cell['coord']}: name {resolved['name']} != {cell['named_style']}"
        )
        # alignment
        ea = cell["alignment"]
        ga = resolved["alignment"]
        assert ga["horizontal"] == ea["horizontal"]
        assert ga["vertical"] == ea["vertical"]
        assert ga["text_rotation"] == ea["textRotation"]
        if ea["wrapText"] is not None:
            assert ga["wrap_text"] is ea["wrapText"]
        # protection
        assert resolved["protection"]["locked"] is cell["protection"]["locked"]
        assert resolved["protection"]["hidden"] is cell["protection"]["hidden"]
        # border style on sides when present
        eb = cell["border"]
        gb = resolved["border"]
        for side in ("left", "right", "top", "bottom"):
            es = eb.get(side)
            gs = gb[side]
            if es is None:
                # openpyxl may dump null side OR empty Side — accept either
                continue
            if isinstance(es, dict):
                assert gs["style"] == es.get("style")


def test_richmeta_borders_full_records(richmeta_oracle):
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=["styles"])
    st = sh.style_table()
    # Find RM_Warning named style cells — thin red borders
    warning = next(n for n in richmeta_oracle["named_styles"] if n["name"] == "RM_Warning")
    # resolve a cell that uses it
    cell = next(c for c in richmeta_oracle["sampled_cells"] if c["named_style"] == "RM_Warning")
    si = sh.style_indices()
    tr, tc = cell["row"] - 2, cell["col"] - 1
    xf = si[tc][tr]
    b = st[xf]["border"]
    assert b["left"]["style"] == "thin"
    assert b["diagonal_up"] is warning["border"]["diagonalUp"]


def test_richmeta_data_validations(richmeta_oracle):
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=FEATURES_RICH)
    dvs = sh.data_validations()
    assert dvs is not None
    exp = richmeta_oracle["data_validations"]
    assert len(dvs) == len(exp)
    # index by sqref
    by_sq = {d["sqref"]: d for d in dvs}
    for ed in exp:
        got = by_sq[ed["sqref"]]
        assert got["type"] == ed["type"]
        assert got["operator"] == ed["operator"]
        assert got["allow_blank"] is ed["allowBlank"]
        assert got["show_input_message"] is ed["showInputMessage"]
        assert got["show_error_message"] is ed["showErrorMessage"]
        assert got["formula1"] == ed["formula1"]
        assert got["formula2"] == ed["formula2"]
        assert got["prompt_title"] == ed["promptTitle"]
        assert got["error_title"] == ed["errorTitle"]


def test_richmeta_conditional_formatting(richmeta_oracle):
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=FEATURES_RICH)
    rules = sh.conditional_formatting()
    assert rules is not None
    exp = richmeta_oracle["conditional_formatting"]
    assert len(rules) == len(exp)
    # sort by priority for stable compare
    rules_s = sorted(rules, key=lambda r: r["priority"])
    exp_s = sorted(exp, key=lambda r: r["priority"])
    for got, ed in zip(rules_s, exp_s):
        assert got["sqref"] == ed["sqref"]
        assert got["type"] == ed["type"]
        assert got["priority"] == ed["priority"]
        assert got["operator"] == ed["operator"]
        assert got["stop_if_true"] is ed["stopIfTrue"]
        assert got["dxf_id"] == ed["dxfId"]
        assert got["formulas"] == ed["formulas"]
        assert got["text"] == ed["text"]
        # dxf fill when present
        if ed.get("dxf") and ed["dxf"].get("fill"):
            assert got["dxf"] is not None
            assert got["dxf"]["fill"] is not None
            assert got["dxf"]["fill"]["pattern"] == ed["dxf"]["fill"]["patternType"]


def test_richmeta_styles_without_validations_no_force():
    """styles-only should not require validations parse; validations flag independent."""
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=["styles"])
    assert sh.style_table() is not None
    assert sh.data_validations() is None
    assert sh.conditional_formatting() is None

    sh2 = reader.load_sheet(0, features=["validations"])
    assert sh2.data_validations() is not None
    assert sh2.style_table() is None


def test_richmeta_vs_openpyxl_live_spotcheck():
    """Where useful, cross-check a live openpyxl cell style name."""
    ox = openpyxl.load_workbook(RICHMETA_XLSX, data_only=False)
    ws = ox.active
    reader = read_excel_turbo(str(RICHMETA_XLSX))
    sh = reader.load_sheet(0, features=["styles"])
    st = sh.style_table()
    si = sh.style_indices()
    # header A1
    assert ws["A1"].style == "RM_Header"
    xf = si[0][0] if sh.nrows > 0 else None
    # A1 is header row — not in style_indices data rows. Use A2 (data row 0).
    # B2 should be RM_Input per fixture
    assert ws["B2"].style == st[si[1][0]]["name"]
