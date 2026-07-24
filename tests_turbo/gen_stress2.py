"""Generate adversarial stress2 fixtures for gap features (S1).

Run from repo root:
  .venv/Scripts/python tests_turbo/gen_stress2.py

Writes testdata/stress2_*.xlsx (and a few hand-built variants).
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chartsheet import Chartsheet
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
TESTDATA = ROOT / "testdata"

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_ODR = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_THREAD = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"

PERSON_ALICE = "{11111111-1111-1111-1111-111111111111}"
PERSON_BOB = "{22222222-2222-2222-2222-222222222222}"
TC_ID = "{aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa}"
TC_UNKNOWN_PERSON = "{99999999-9999-9999-9999-999999999999}"


def _ensure_dir() -> None:
    TESTDATA.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# 1. stress2_sparse_meta.xlsx
# ---------------------------------------------------------------------------


def gen_sparse_meta(path: Path) -> None:
    """Sparse values + row heights/hidden on empty rows + cols past used range + filter/freeze."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SparseMeta"

    # Header
    for c in range(1, 6):
        ws.cell(1, c, f"C{c}")

    # Sparse data with gaps: rows 2-4 empty, values at A5, C5, D10, A20
    ws["A5"] = 100.0
    ws["C5"] = 300.0  # no B5
    ws["D10"] = 400.0
    ws["A20"] = 1.0
    ws["C20"] = 2.0
    # denser islands
    for r in (6, 7, 15, 30, 50):
        ws.cell(r, 1, float(r))
        if r % 2 == 0:
            ws.cell(r, 3, float(r * 10))

    # Row dims on EMPTY rows (no cells) — regression: @r still recorded
    ws.row_dimensions[2].height = 30.0  # empty
    ws.row_dimensions[3].hidden = True  # empty
    ws.row_dimensions[4].height = 18.0
    ws.row_dimensions[4].outlineLevel = 1
    ws.row_dimensions[8].height = 25.0  # empty row between data
    ws.row_dimensions[8].hidden = True
    ws.row_dimensions[5].height = 22.0  # has cells
    ws.row_dimensions[20].height = 16.0

    # Col properties beyond used range (used cols ~A-D; set F/G/Z)
    ws.column_dimensions["A"].width = 14.0
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["F"].width = 20.0  # beyond sparse used
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["Z"].width = 5.0  # far beyond

    # Autofilter + freeze on gappy sheet
    ws.auto_filter.ref = "A1:E50"
    ws.freeze_panes = "B5"

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# 2. stress2_multisheet_meta.xlsx
# ---------------------------------------------------------------------------


def gen_multisheet_meta(path: Path) -> None:
    """4 worksheets with DIFFERENT meta/features + chartsheet; isolation check."""
    wb = Workbook()

    # --- Sheet 0: ValidCF — validations + CF + dims ---
    ws0 = wb.active
    ws0.title = "ValidCF"
    ws0["A1"] = "val"
    ws0["B1"] = "cat"
    for r in range(2, 12):
        ws0.cell(r, 1, r * 10)
        ws0.cell(r, 2, "A" if r % 2 == 0 else "B")
    ws0.row_dimensions[2].height = 28.0
    ws0.column_dimensions["A"].width = 12.0
    ws0.freeze_panes = "A2"
    dv0 = DataValidation(type="whole", operator="between", formula1="1", formula2="200", allow_blank=True)
    dv0.add("A2:A11")
    ws0.add_data_validation(dv0)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws0.conditional_formatting.add(
        "A2:A11",
        CellIsRule(operator="greaterThan", formula=["50"], fill=red_fill),
    )
    ws0.protection.sheet = True
    ws0.protection.enable()

    # --- Sheet 1: ChartsOnly — bar chart, no validations/CF/threaded ---
    ws1 = wb.create_sheet("ChartsOnly")
    ws1["A1"] = "Month"
    ws1["B1"] = "Rev"
    for i, m in enumerate(["Jan", "Feb", "Mar"], start=2):
        ws1.cell(i, 1, m)
        ws1.cell(i, 2, i * 10)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Sheet2 Bar"
    bar.add_data(Reference(ws1, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    bar.set_categories(Reference(ws1, min_col=1, min_row=2, max_row=4))
    ws1.add_chart(bar, "D2")
    ws1.row_dimensions[1].height = 20.0
    ws1.column_dimensions["B"].width = 15.0
    ws1.freeze_panes = "B2"
    # different protection
    ws1.protection.sheet = False

    # --- Sheet 2: Threaded — threaded comments only (injected post-save) ---
    ws2 = wb.create_sheet("Threaded")
    ws2["A1"] = "note"
    ws2["B1"] = "v"
    ws2["A2"] = "root-cell"
    ws2["B2"] = 42
    ws2["C3"] = "other"
    ws2.row_dimensions[3].height = 40.0
    ws2.column_dimensions["C"].width = 18.0
    ws2.freeze_panes = "A3"
    # different validation type
    dv2 = DataValidation(type="list", formula1='"X,Y,Z"', allow_blank=True)
    dv2.add("B2")
    ws2.add_data_validation(dv2)
    # CF expression (different from sheet 0)
    ws2.conditional_formatting.add(
        "B2",
        FormulaRule(formula=["$B$2>10"], fill=PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")),
    )
    ws2.protection.sheet = True
    ws2.protection.password = "iso"

    # --- Sheet 3: DimsPanes — dims/panes/filter only, no charts/CF/threaded ---
    ws3 = wb.create_sheet("DimsPanes")
    ws3["A1"] = "h1"
    ws3["B1"] = "h2"
    for r in range(2, 8):
        ws3.cell(r, 1, r)
        ws3.cell(r, 2, r * 2)
    ws3.row_dimensions[1].height = 24.0
    ws3.row_dimensions[4].hidden = True
    ws3.column_dimensions["A"].width = 10.0
    ws3.column_dimensions["D"].width = 9.0  # beyond used
    ws3.auto_filter.ref = "A1:B7"
    ws3.freeze_panes = "C2"
    ws3.protection.sheet = False

    # --- Chartsheet ---
    try:
        cs = wb.create_chartsheet("ChartSheetOnly")
        # openpyxl chartsheet: attach a chart
        bar2 = BarChart()
        bar2.title = "Chartsheet Title"
        bar2.add_data(Reference(ws1, min_col=2, min_row=1, max_row=4), titles_from_data=True)
        cs.add_chart(bar2)
    except Exception:
        # Fallback: leave workbook without chartsheet; tests will skip kind check
        pass

    wb.save(path)
    wb.close()

    # Inject threaded comments on sheet "Threaded" only (sheet index 2 → sheet3.xml typically)
    _inject_threaded_on_sheet(path, sheet_title="Threaded", ref="A2", text="iso-threaded on sheet3")


def _sheet_path_for_title(parts: dict[str, bytes], title: str) -> str | None:
    """Resolve worksheets/sheetN.xml for a workbook sheet name."""
    wb_xml = parts.get("xl/workbook.xml", b"").decode("utf-8", "replace")
    # sheet name → r:id
    m = re.search(
        rf'<sheet[^>]*name="{re.escape(title)}"[^>]*r:id="(rId\d+)"',
        wb_xml,
    )
    if not m:
        m = re.search(
            rf'<sheet[^>]*r:id="(rId\d+)"[^>]*name="{re.escape(title)}"',
            wb_xml,
        )
    if not m:
        return None
    rid = m.group(1)
    rels = parts.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8", "replace")
    rm = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', rels)
    if not rm:
        rm = re.search(rf'Target="([^"]+)"[^>]*Id="{rid}"', rels)
    if not rm:
        return None
    target = rm.group(1).lstrip("/")
    if not target.startswith("xl/"):
        target = "xl/" + target if not target.startswith("worksheets") else "xl/" + target
    if target.startswith("worksheets/"):
        target = "xl/" + target
    return target.replace("\\", "/")


def _inject_threaded_on_sheet(path: Path, sheet_title: str, ref: str, text: str) -> None:
    """Add persons + threadedComments part linked only to one worksheet."""
    buf = path.read_bytes()
    with zipfile.ZipFile(io.BytesIO(buf), "r") as zin:
        parts = {n: zin.read(n) for n in zin.namelist()}

    sheet_path = _sheet_path_for_title(parts, sheet_title)
    if not sheet_path:
        # try common names
        for cand in ("xl/worksheets/sheet3.xml", "xl/worksheets/sheet2.xml"):
            if cand in parts:
                sheet_path = cand
                break
    if not sheet_path or sheet_path not in parts:
        return

    sheet_file = sheet_path.rsplit("/", 1)[-1]
    rels_path = f"xl/worksheets/_rels/{sheet_file}.rels"
    persons_part = "xl/persons/person.xml"
    tc_part = "xl/threadedComments/threadedComment_stress2.xml"

    persons_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<personList xmlns="{NS_THREAD}" xmlns:x="{NS_MAIN}">
  <person displayName="Alice" id="{PERSON_ALICE}" userId="alice@example.com" providerId="None"/>
  <person displayName="Bob" id="{PERSON_BOB}" userId="bob@example.com" providerId="None"/>
</personList>
"""
    tc_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<ThreadedComments xmlns="{NS_THREAD}" xmlns:x="{NS_MAIN}">
  <threadedComment ref="{ref}" personId="{PERSON_ALICE}" id="{TC_ID}" dT="2024-06-01T10:00:00Z">
    <text>{text}</text>
  </threadedComment>
</ThreadedComments>
"""
    parts[persons_part] = persons_xml.encode("utf-8")
    parts[tc_part] = tc_xml.encode("utf-8")

    # Content types
    ct = parts["[Content_Types].xml"].decode("utf-8")
    if "threadedComment" not in ct:
        ct = ct.replace(
            "</Types>",
            f'<Override PartName="/{tc_part}" ContentType="application/vnd.ms-excel.threadedcomments+xml"/>\n'
            f'<Override PartName="/{persons_part}" ContentType="application/vnd.ms-excel.person+xml"/>\n</Types>',
        )
        parts["[Content_Types].xml"] = ct.encode("utf-8")

    # Workbook rel for persons
    wb_rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    if "person" not in wb_rels:
        max_id = max(int(x) for x in re.findall(r'Id="rId(\d+)"', wb_rels))
        rid = f"rId{max_id + 1}"
        wb_rels = wb_rels.replace(
            "</Relationships>",
            f'<Relationship Id="{rid}" Type="{NS_ODR}/person" Target="persons/person.xml"/>\n</Relationships>',
        )
        parts["xl/_rels/workbook.xml.rels"] = wb_rels.encode("utf-8")

    # Sheet rel for threaded comments
    if rels_path in parts:
        srels = parts[rels_path].decode("utf-8")
    else:
        srels = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="{NS_REL}">\n</Relationships>'
    if "threadedComment" not in srels:
        max_id = 0
        ids = re.findall(r'Id="rId(\d+)"', srels)
        if ids:
            max_id = max(int(x) for x in ids)
        rid = f"rId{max_id + 1}"
        # relative target from worksheets/
        rel_target = "../threadedComments/threadedComment_stress2.xml"
        srels = srels.replace(
            "</Relationships>",
            f'<Relationship Id="{rid}" Type="{NS_ODR}/threadedComment" Target="{rel_target}"/>\n</Relationships>',
        )
        parts[rels_path] = srels.encode("utf-8")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in parts.items():
            zout.writestr(name, data)
    path.write_bytes(out.getvalue())


# ---------------------------------------------------------------------------
# 3. stress2_onecell_all.xlsx — cross-feature single cell
# ---------------------------------------------------------------------------


def gen_onecell_all(path: Path) -> None:
    """One cell: merge + named style + full border + validation + CF + hyperlink + formula + threaded."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OneCell"

    # Named style with full border
    thin = Side(style="thin", color="FF0000")
    thick = Side(style="medium", color="0000FF")
    ns = NamedStyle(
        name="StressAll",
        font=Font(bold=True, name="Arial", sz=12, color="0000FF"),
        fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        border=Border(left=thin, right=thin, top=thick, bottom=thick),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    if "StressAll" not in wb.named_styles:
        wb.add_named_style(ns)

    for i, h in enumerate(["A", "B", "C", "D"], 1):
        ws.cell(1, i, h)

    # Merge B2:C3 — top-left is the feature cell
    ws.merge_cells("B2:C3")
    cell = ws["B2"]
    cell.value = "=1+2"
    cell.style = "StressAll"
    cell.hyperlink = "https://example.com/onecell"
    cell.comment = Comment("legacy mirror for threaded", "Alice")

    # Neighbors so grid is non-empty
    ws["A2"] = 10
    ws["D2"] = 20
    ws["A4"] = 30

    # Validation covering B2
    dv = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
    dv.add("B2")
    ws.add_data_validation(dv)

    # CF covering B2
    ws.conditional_formatting.add(
        "B2:C3",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            font=Font(color="FF0000", bold=True),
        ),
    )

    wb.save(path)
    wb.close()

    # Inject cached formula value + threaded comment on B2
    _patch_onecell_cache_and_threaded(path)


def _patch_onecell_cache_and_threaded(path: Path) -> None:
    buf = path.read_bytes()
    with zipfile.ZipFile(io.BytesIO(buf), "r") as zin:
        parts = {n: zin.read(n) for n in zin.namelist()}

    # Inject <v>3</v> after formula on B2 if missing
    for name in list(parts):
        if not (name.startswith("xl/worksheets/sheet") and name.endswith(".xml")):
            continue
        text = parts[name].decode("utf-8")
        # Match cell B2 with formula, ensure cached value
        def add_v(m: re.Match) -> str:
            body = m.group(0)
            if "<v>" in body:
                return body
            # insert before </c>
            return body.replace("</c>", "<v>3</v></c>")

        text2 = re.sub(
            r'<c r="B2"[^>]*>.*?</c>',
            add_v,
            text,
            count=1,
            flags=re.DOTALL,
        )
        parts[name] = text2.encode("utf-8")

    # Threaded comment on B2
    tmp = Path(str(path) + ".tmp")
    with zipfile.ZipFile(io.BytesIO(), "w") as _:
        pass
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in parts.items():
            zout.writestr(name, data)
    path.write_bytes(out.getvalue())
    _inject_threaded_on_sheet(path, sheet_title="OneCell", ref="B2", text="threaded on feature cell")


# ---------------------------------------------------------------------------
# 4. stress2_tail_order.xlsx — permuted tail elements
# ---------------------------------------------------------------------------


def gen_tail_order(path: Path) -> None:
    """Hand-built sheet: conditionalFormatting + dataValidations BEFORE hyperlinks."""
    content_types = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="{NS_CT}">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""
    rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""
    wb_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <sheets>
    <sheet name="TailOrder" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>"""
    wb_rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="{NS_ODR}/styles" Target="styles.xml"/>
</Relationships>"""
    sheet_rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/hyperlink" Target="https://example.com/tail" TargetMode="External"/>
</Relationships>"""
    # Permuted order vs openpyxl: CF, then DV, then hyperlinks (openpyxl: hyperlinks often earlier)
    sheet = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <sheetData>
    <row r="1">
      <c r="A1" t="inlineStr"><is><t>h</t></is></c>
      <c r="B1" t="inlineStr"><is><t>v</t></is></c>
    </row>
    <row r="2">
      <c r="A2"><v>10</v></c>
      <c r="B2"><v>20</v></c>
    </row>
  </sheetData>
  <conditionalFormatting sqref="A2">
    <cfRule type="cellIs" dxfId="0" priority="1" operator="greaterThan">
      <formula>5</formula>
    </cfRule>
  </conditionalFormatting>
  <dataValidations count="1">
    <dataValidation type="whole" operator="greaterThan" allowBlank="1" sqref="A2">
      <formula1>0</formula1>
    </dataValidation>
  </dataValidations>
  <hyperlinks>
    <hyperlink ref="B2" r:id="rId1"/>
  </hyperlinks>
</worksheet>"""
    styles = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="{NS_MAIN}">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="1"><fill><patternFill patternType="none"/></fill></fills>
  <borders count="1"><border/></borders>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellXfs>
  <dxfs count="1">
    <dxf><font><b/><color rgb="FFFF0000"/></font></dxf>
  </dxfs>
</styleSheet>"""

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("xl/workbook.xml", wb_xml)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/worksheets/sheet1.xml", sheet)
        z.writestr("xl/worksheets/_rels/sheet1.xml.rels", sheet_rels)
        z.writestr("xl/styles.xml", styles)


# ---------------------------------------------------------------------------
# 5. stress2_malformed_*.xlsx — hand-built degradation cases (also covered in Rust)
# ---------------------------------------------------------------------------


def _write_store_zip(path: Path, entries: list[tuple[str, bytes]]) -> None:
    """Store-only zip (no deflate) for simple hand fixtures."""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_STORED) as z:
        for name, data in entries:
            z.writestr(name, data)


def _minimal_ct_wb_sheet(
    sheet_xml: str,
    extra_ct: str = "",
    extra_parts: list[tuple[str, bytes]] | None = None,
    wb_rels_extra: str = "",
    sheet_rels: str | None = None,
) -> list[tuple[str, bytes]]:
    ct = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="{NS_CT}">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  {extra_ct}
</Types>"""
    rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""
    wb = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <sheets><sheet name="S" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""
    wb_rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="{NS_ODR}/styles" Target="styles.xml"/>
  {wb_rels_extra}
</Relationships>"""
    styles = f"""<?xml version="1.0"?><styleSheet xmlns="{NS_MAIN}">
  <fonts count="1"><font/></fonts><fills count="1"><fill/></fills>
  <borders count="1"><border/></borders>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellXfs>
  <dxfs count="1"><dxf><font><b/></font></dxf></dxfs>
</styleSheet>"""
    out = [
        ("[Content_Types].xml", ct.encode()),
        ("_rels/.rels", rels.encode()),
        ("xl/workbook.xml", wb.encode()),
        ("xl/_rels/workbook.xml.rels", wb_rels.encode()),
        ("xl/worksheets/sheet1.xml", sheet_xml.encode()),
        ("xl/styles.xml", styles.encode()),
    ]
    if sheet_rels is not None:
        out.append(("xl/worksheets/_rels/sheet1.xml.rels", sheet_rels.encode()))
    if extra_parts:
        out.extend(extra_parts)
    return out


def gen_malformed(path_prefix: Path) -> None:
    base = path_prefix  # directory

    # empty sqref DV + min>max cols + oob dxfId
    sheet = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <cols>
    <col min="5" max="3" width="10"/>
    <col min="1" max="2" width="8"/>
  </cols>
  <sheetData>
    <row r="1"><c r="A1" t="inlineStr"><is><t>h</t></is></c></row>
    <row r="2"><c r="A2"><v>1</v></c></row>
  </sheetData>
  <conditionalFormatting sqref="A2">
    <cfRule type="cellIs" dxfId="99" priority="1" operator="equal">
      <formula>1</formula>
    </cfRule>
  </conditionalFormatting>
  <dataValidations count="2">
    <dataValidation type="whole" sqref="" allowBlank="1">
      <formula1>1</formula1>
    </dataValidation>
    <dataValidation type="whole" sqref="A2" allowBlank="1">
      <formula1>0</formula1>
    </dataValidation>
  </dataValidations>
</worksheet>"""
    _write_store_zip(
        base / "stress2_malformed_meta.xlsx",
        _minimal_ct_wb_sheet(sheet),
    )

    # truncated chart XML
    sheet_c = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <sheetData>
    <row r="1"><c r="A1" t="inlineStr"><is><t>h</t></is></c></row>
    <row r="2"><c r="A2"><v>1</v></c></row>
  </sheetData>
  <drawing r:id="rId1"/>
</worksheet>"""
    sheet_rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/drawing" Target="../drawings/drawing1.xml"/>
</Relationships>"""
    drawing = f"""<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
 xmlns:r="{NS_ODR}">
  <xdr:twoCellAnchor>
    <xdr:from><xdr:col>0</xdr:col><xdr:row>0</xdr:row></xdr:from>
    <xdr:to><xdr:col>2</xdr:col><xdr:row>2</xdr:row></xdr:to>
    <xdr:graphicFrame>
      <xdr:graphic>
        <a:graphicData xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
          <c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" r:id="rId1"/>
        </a:graphicData>
      </xdr:graphic>
    </xdr:graphicFrame>
  </xdr:twoCellAnchor>
</xdr:wsDr>"""
    drawing_rels = f"""<?xml version="1.0"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/chart" Target="../charts/chart1.xml"/>
</Relationships>"""
    # truncated mid-chart
    chart_trunc = b"""<?xml version="1.0"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">
  <c:chart><c:title><c:tx><c:rich><a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><a:r><a:t>Trunc"""
    ct_extra = """
  <Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>
  <Override PartName="/xl/charts/chart1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>
"""
    entries = _minimal_ct_wb_sheet(
        sheet_c,
        extra_ct=ct_extra,
        sheet_rels=sheet_rels,
        extra_parts=[
            ("xl/drawings/drawing1.xml", drawing.encode()),
            ("xl/drawings/_rels/drawing1.xml.rels", drawing_rels.encode()),
            ("xl/charts/chart1.xml", chart_trunc),
        ],
    )
    _write_store_zip(base / "stress2_malformed_chart.xlsx", entries)

    # pivot missing cacheDefinition rel
    sheet_p = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <sheetData>
    <row r="1"><c r="A1" t="inlineStr"><is><t>h</t></is></c></row>
    <row r="2"><c r="A2"><v>1</v></c></row>
  </sheetData>
</worksheet>"""
    # sheet rel to pivot table, table has NO cache rel
    srels_p = f"""<?xml version="1.0"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId9" Type="{NS_ODR}/pivotTable" Target="../pivotTables/pivotTable1.xml"/>
</Relationships>"""
    pivot_table = f"""<?xml version="1.0" encoding="UTF-8"?>
<pivotTableDefinition xmlns="{NS_MAIN}" name="P1" cacheId="0" dataCaption="Values">
  <location ref="E3:G8" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"/>
  <pivotFields count="1"><pivotField dataField="1" showAll="0"/></pivotFields>
</pivotTableDefinition>"""
    # empty pivot table rels (missing cacheDefinition)
    pivot_rels = f"""<?xml version="1.0"?>
<Relationships xmlns="{NS_REL}">
</Relationships>"""
    ct_p = """
  <Override PartName="/xl/pivotTables/pivotTable1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>
"""
    entries = _minimal_ct_wb_sheet(
        sheet_p,
        extra_ct=ct_p,
        sheet_rels=srels_p,
        extra_parts=[
            ("xl/pivotTables/pivotTable1.xml", pivot_table.encode()),
            ("xl/pivotTables/_rels/pivotTable1.xml.rels", pivot_rels.encode()),
        ],
    )
    _write_store_zip(base / "stress2_malformed_pivot.xlsx", entries)

    # threadedComments unknown personId
    sheet_t = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_ODR}">
  <sheetData>
    <row r="1"><c r="A1" t="inlineStr"><is><t>h</t></is></c></row>
    <row r="2"><c r="A2" t="inlineStr"><is><t>x</t></is></c></row>
  </sheetData>
</worksheet>"""
    srels_t = f"""<?xml version="1.0"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1" Type="{NS_ODR}/threadedComment" Target="../threadedComments/threadedComment1.xml"/>
</Relationships>"""
    persons = f"""<?xml version="1.0"?>
<personList xmlns="{NS_THREAD}">
  <person displayName="Alice" id="{PERSON_ALICE}"/>
</personList>"""
    tc = f"""<?xml version="1.0"?>
<ThreadedComments xmlns="{NS_THREAD}">
  <threadedComment ref="A2" personId="{TC_UNKNOWN_PERSON}" id="{TC_ID}" dT="2024-01-01T00:00:00Z">
    <text>orphan person</text>
  </threadedComment>
</ThreadedComments>"""
    ct_t = f"""
  <Override PartName="/xl/threadedComments/threadedComment1.xml" ContentType="application/vnd.ms-excel.threadedcomments+xml"/>
  <Override PartName="/xl/persons/person.xml" ContentType="application/vnd.ms-excel.person+xml"/>
"""
    wb_extra = f'<Relationship Id="rId3" Type="{NS_ODR}/person" Target="persons/person.xml"/>'
    entries = _minimal_ct_wb_sheet(
        sheet_t,
        extra_ct=ct_t,
        sheet_rels=srels_t,
        wb_rels_extra=wb_extra,
        extra_parts=[
            ("xl/threadedComments/threadedComment1.xml", tc.encode()),
            ("xl/persons/person.xml", persons.encode()),
        ],
    )
    _write_store_zip(base / "stress2_malformed_threaded.xlsx", entries)

    # vba content-type present but part missing
    ct_vba = """
  <Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/>
"""
    wb_vba = f'<Relationship Id="rId9" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/>'
    sheet_v = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}">
  <sheetData>
    <row r="1"><c r="A1" t="inlineStr"><is><t>h</t></is></c></row>
    <row r="2"><c r="A2"><v>1</v></c></row>
  </sheetData>
</worksheet>"""
    entries = _minimal_ct_wb_sheet(sheet_v, extra_ct=ct_vba, wb_rels_extra=wb_vba)
    # deliberately NO xl/vbaProject.bin
    _write_store_zip(base / "stress2_malformed_vba.xlsx", entries)


def main() -> None:
    _ensure_dir()
    targets = [
        ("stress2_sparse_meta.xlsx", gen_sparse_meta),
        ("stress2_multisheet_meta.xlsx", gen_multisheet_meta),
        ("stress2_onecell_all.xlsx", gen_onecell_all),
        ("stress2_tail_order.xlsx", gen_tail_order),
    ]
    for name, fn in targets:
        p = TESTDATA / name
        print(f"writing {p} ...")
        fn(p)
        print(f"  -> {p.stat().st_size} bytes")
    print("writing malformed variants ...")
    gen_malformed(TESTDATA)
    for name in (
        "stress2_malformed_meta.xlsx",
        "stress2_malformed_chart.xlsx",
        "stress2_malformed_pivot.xlsx",
        "stress2_malformed_threaded.xlsx",
        "stress2_malformed_vba.xlsx",
    ):
        p = TESTDATA / name
        print(f"  -> {p} ({p.stat().st_size} bytes)")
    print("done.")


if __name__ == "__main__":
    main()
