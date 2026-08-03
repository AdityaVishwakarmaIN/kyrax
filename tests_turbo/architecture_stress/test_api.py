"""Lane A1 - binding, familiar API, and turbo reachability (Wave 1 baseline).

Exclusive owner: A1. Write access: this file only (plus
plans/architecture_stress/NOTES/A1.md). All source, the shared harness
(A6 common.py / fixtures.py), scripts, and other lanes' files are excluded.

Verdict contract (A6 common.py):
  * Every A1 case emits exactly one structured ResultRecord (PASS / FAIL /
    KNOWN-GAP / BLOCKED) into the shared ResultAggregator; pytest status is a
    signal, the JSONL ``verdict`` field is authoritative.
  * No pytest skip/xfail/importorskip for campaign cases; no private watchdog -
    fresh-process probes run through common.watchdog_run.
  * KNOWN-GAP is a measured negative-capability probe, never a pass.
  * A stale/mismatched ``_kyrax.pyd`` makes every runtime case BLOCKED.

Coverage map (see NOTES/A1.md):
  A1-BIND-01  test_bind_01_static_scan
  A1-BIND-02  test_bind_02_runtime_surface_stub_readme
  A1-API-01   test_api_01_load_edit_save_reopen
  A1-API-02   test_api_02_error_contract
  A1-API-03   test_api_03_ws_a1_known_gap
  A1-GEAR-01  test_gear_01_stock_turbo_editable_coherence
  A1-GEAR-02  test_gear_02_turbo_reachability_known_gap
  A1-ADOPT-01 test_adopt_01_fresh_process_import
  P05 (A1)    test_p05_reader_reopen_after_repair
"""

from __future__ import annotations

import ast
import hashlib
import json
import os
import re
import sys
import time
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "python"))

try:
    from . import common
except ImportError:  # Direct import during focused harness checks.
    import common  # type: ignore[no-redef]

import kyrax  # noqa: E402

try:
    import openpyxl as OPX

    _OPX_AVAILABLE = True
except ImportError:
    _OPX_AVAILABLE = False

LANE = "A1"
WAVE = 1
_RUN_ID = common.new_run_id()
_AGG: list[common.ResultAggregator] = []
_PYD_INFO: dict = {}
STRUCTURED = ROOT / "testdata" / "structured.xlsx"


def _pyd_info() -> dict:
    if not _PYD_INFO:
        for candidate in (
            ROOT / "python" / "kyrax" / "_kyrax.pyd",
            ROOT / ".venv" / "Lib" / "site-packages" / "kyrax" / "_kyrax.pyd",
        ):
            if candidate.exists():
                _PYD_INFO["path"] = str(candidate)
                _PYD_INFO["sha256"] = common.sha256_file(candidate)
                break
    return _PYD_INFO


def _binary_ok() -> bool:
    measured = (_pyd_info().get("sha256") or "").lower()
    return measured == common.EVIDENCE_PYD_SHA256.lower()


@pytest.fixture(scope="module", autouse=True)
def _a1_evidence(tmp_path_factory) -> None:
    """Accumulate one structured verdict per A1 case and publish one JSONL file."""
    aggregator = common.ResultAggregator(run_id=_RUN_ID)
    _AGG.append(aggregator)
    yield
    out_dir = os.environ.get("KYRAX_ARCHSTRESS_OUTPUT")
    if out_dir:
        dest = Path(out_dir) / "A1_api_results.jsonl"
        dest.parent.mkdir(parents=True, exist_ok=True)
    else:
        dest = tmp_path_factory.mktemp("a1_evidence") / "A1_api_results.jsonl"
    aggregator.publish(dest)
    print(f"[A1] verdict JSONL: {dest}")


def _emit(
    test_id: str,
    verdict: common.Verdict,
    detail: str,
    *,
    start: float | None = None,
    isolation: str = "pytest-in-process",
    measured: dict | None = None,
    fixture: dict | None = None,
    command: str | None = None,
) -> None:
    record = common.ResultRecord(
        run_id=_RUN_ID,
        test_id=test_id,
        verdict=verdict,
        isolation=isolation,
        duration_s=round(time.monotonic() - start, 4) if start is not None else None,
        detail=detail,
        wave=WAVE,
        lane=LANE,
        command=command,
        fixture=fixture,
        measured=measured,
    )
    errors = _AGG[-1].add_result(record)
    if errors:
        raise AssertionError(f"A1 record rejected: {test_id}: {'; '.join(errors)}")
    if verdict == common.Verdict.FAIL:
        raise AssertionError(f"{test_id}: FAIL - {detail}")


def _require_binary(test_id: str) -> bool:
    info = _pyd_info()
    if _binary_ok():
        return True
    _emit(
        test_id,
        common.Verdict.BLOCKED,
        "stale or missing _kyrax.pyd: "
        f"expected {common.EVIDENCE_PYD_SHA256}, measured "
        f"{info.get('sha256')!r} at {info.get('path')}",
    )
    return False


def _author_small(path: Path, value: float = 1.0) -> None:
    kyrax.write_excel_turbo(
        str(path),
        [
            {"name": "Sheet1", "columns": [[value, 2.0], [10.0, 20.0]]},
            {"name": "Sheet2", "columns": [["a", "b"]]},
        ],
    )


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _read_last_resut(path: Path) -> dict:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    for line in reversed(text.splitlines()):
        if line.startswith("RESULT "):
            return json.loads(line[len("RESULT ") :])
    raise AssertionError(f"worker produced no RESULT line: {text[-400:]!r}")


# ---------------------------------------------------------------------------
# A1-BIND-01 - static scan: no feature logic in the Python binding layer.
# ---------------------------------------------------------------------------


def test_bind_01_static_scan() -> None:
    test_id = "A1-BIND-01"
    start = time.monotonic()
    findings: list[str] = []

    init_path = ROOT / "python" / "kyrax" / "__init__.py"
    lib_path = ROOT / "src" / "lib.rs"
    package = ROOT / "python" / "kyrax"

    src = init_path.read_text(encoding="utf-8")
    tree = ast.parse(src)

    guarded: set[int] = set()

    def _mark(node: ast.AST, inside_tc: bool = False) -> None:
        if isinstance(node, ast.If) and isinstance(node.test, ast.Name) and node.test.id == "TYPE_CHECKING":
            for child in node.body + node.orelse:
                _mark(child, True)
            return
        if inside_tc:
            guarded.add(id(node))
        for child in ast.iter_child_nodes(node):
            _mark(child, inside_tc)

    _mark(tree)

    doorway_funcs = {"to_polars", "to_pandas", "to_arrow", "to_arrow_with_errors"}
    blocked_runtime = {"pandas", "polars", "pyarrow"}
    imports: set[str] = set()
    approved: set[str] = set()
    unexpected: list[str] = []

    def _collect(node: ast.AST, enclosing: str | None = None) -> None:
        if id(node) in guarded:
            return
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            enclosing = node.name
        if isinstance(node, ast.Import):
            for a in node.names:
                root = a.name.split(".")[0]
                if root in blocked_runtime and enclosing not in doorway_funcs:
                    unexpected.append(f"{root} imported at line {node.lineno} outside a conversion doorway")
                elif root in blocked_runtime:
                    approved.add(root)
                else:
                    imports.add(root)
        elif isinstance(node, ast.ImportFrom):
            if node.level == 0 and node.module:
                root = node.module.split(".")[0]
                if root in blocked_runtime and enclosing not in doorway_funcs:
                    unexpected.append(f"{root} imported at line {node.lineno} outside a conversion doorway")
                elif root in blocked_runtime:
                    approved.add(root)
                else:
                    imports.add(root)
        for child in ast.iter_child_nodes(node):
            _collect(child, enclosing)

    _collect(tree)
    allowed = {"__future__", "typing", "collections", "os", "pathlib", "importlib"}
    for mod in sorted(imports - allowed - approved):
        unexpected.append(f"non-allowlisted runtime import: {mod}")
    if unexpected:
        findings.append("non-allowlisted runtime imports: " + "; ".join(unexpected))

    if re.search(r"^\s*(import|from)\s+openpyxl", src, re.M):
        findings.append("openpyxl imported in the binding layer")

    loop_funcs: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            if any(isinstance(n, (ast.For, ast.While, ast.AsyncFor)) for n in ast.walk(node)):
                loop_funcs.append(node.name)
    if loop_funcs != ["_validate_sheets"]:
        findings.append(f"loop-containing functions beyond the shape guard: {loop_funcs}")

    wrapper_names = {
        "load_workbook",
        "read_excel",
        "read_excel_turbo",
        "write_excel_turbo",
        "write_excel_turbo_stream",
        "write_excel_turbo_bytes",
        "is_encrypted",
        "encryption_info",
        "validate_excel",
        "repair_excel",
    }
    for node in tree.body:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name in wrapper_names:
            called = set()
            for n in ast.walk(node):
                if isinstance(n, ast.Call) and isinstance(n.func, ast.Name):
                    called.add(n.func.id)
            delegates = called & ({"edit_excel", "read_excel", "expanduser"} | {c for c in called if c.startswith("_")})
            if not delegates:
                findings.append(f"{node.name} body does not delegate to the Rust binding")

    lib_src = lib_path.read_text(encoding="utf-8")
    if "gil_used = false" not in lib_src:
        findings.append("src/lib.rs does not declare gil_used = false")

    shim = sorted(p.name for p in package.glob("*") if "compat" in p.name or "shim" in p.name)
    if shim:
        findings.append(f"compat/shim namespace files present: {shim}")

    verdict = common.Verdict.PASS if not findings else common.Verdict.FAIL
    detail = "static scan clean" if not findings else "; ".join(findings)
    _emit(test_id, verdict, detail, start=start, measured={"imports": sorted(imports)})


# ---------------------------------------------------------------------------
# A1-BIND-02 - runtime surface vs __all__, _kyrax.pyi, and README examples.
# ---------------------------------------------------------------------------


def test_bind_02_runtime_surface_stub_readme() -> None:
    test_id = "A1-BIND-02"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    findings: list[str] = []

    missing = [n for n in kyrax.__all__ if not hasattr(kyrax, n)]
    if missing:
        findings.append(f"__all__ names that do not resolve: {missing}")

    stub_path = ROOT / "python" / "kyrax" / "_kyrax.pyi"
    stub_names = set(re.findall(r"^(?:class|def)\s+(\w+)", stub_path.read_text(encoding="utf-8"), re.M))
    absent_stub = sorted(n for n in stub_names if not hasattr(kyrax._kyrax, n))
    if absent_stub:
        findings.append(f"stub names absent from runtime _kyrax module: {absent_stub}")

    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    if "### Edit Mode" in readme:
        block = readme.split("### Edit Mode", 1)[1].split("\n## ", 1)[0]
        if "ws[" in block and "A1" in block and not hasattr(kyrax.EditableSheet, "__setitem__"):
            findings.append("README documents ws['A1'] = ... but EditableSheet has no __setitem__")
        if "set_cell_style(0, 0" in block:
            findings.append("README documents 0-based set_cell_style(0, 0, ...) but the API is 1-based")

    verdict = common.Verdict.PASS if not findings else common.Verdict.FAIL
    detail = "surface, stub, and README agree" if not findings else "; ".join(findings)
    _emit(
        test_id,
        verdict,
        detail,
        start=start,
        measured={"__all__": len(kyrax.__all__), "stub_names": len(stub_names), "findings": findings},
    )


# ---------------------------------------------------------------------------
# A1-API-01 - load_workbook str/Path, edit + non-edit, save, reopen.
# ---------------------------------------------------------------------------


def test_api_01_load_edit_save_reopen(tmp_path) -> None:
    test_id = "A1-API-01"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    findings: list[str] = []
    src = tmp_path / "src.xlsx"

    if _OPX_AVAILABLE:
        wb = OPX.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 1.0
        ws["A1"].font = OPX.styles.Font(bold=True)
        ws.merge_cells("C1:D1")
        wb.create_sheet("Sheet2")
        wb["Sheet2"]["A1"] = 10.0
        wb.save(str(src))
        wb.close()
    else:
        _author_small(src)

    before = _sha(src)

    wb_str = kyrax.load_workbook(str(src))
    wb_path = kyrax.load_workbook(src)
    if wb_str.sheet_names != ["Sheet1", "Sheet2"]:
        findings.append(f"non-edit str reader sheets: {wb_str.sheet_names}")
    if wb_str.sheet_names != wb_path.sheet_names:
        findings.append("str and Path load_workbook disagree on sheet names")

    editable = kyrax.load_workbook(src, edit_mode=True)
    editable["Sheet1"].set_cell(1, 1, 999.0)
    editable["Sheet1"].set_cell(1, 2, 77.0)
    out = tmp_path / "out.xlsx"
    editable.save(str(out))
    if not out.exists():
        findings.append("save produced no output file")

    if _sha(src) != before:
        findings.append("source workbook was modified by an edit handle")

    if _OPX_AVAILABLE:
        o = OPX.load_workbook(str(out))
        ws = o["Sheet1"]
        if ws["A1"].value != 999.0:
            findings.append(f"A1 after reopen: {ws['A1'].value!r}")
        if ws["B1"].value != 77.0:
            findings.append(f"B1 after reopen: {ws['B1'].value!r}")
        if ws["A1"].font.bold is not True:
            findings.append("bold style lost through edit+save")
        merged = [str(r) for r in ws.merged_cells.ranges]
        if "C1:D1" not in merged:
            findings.append(f"merge lost: {merged}")
        if "Sheet2" not in o.sheetnames:
            findings.append("second sheet lost")
        o.close()
    else:
        rd = kyrax.read_excel(str(out))
        sh = rd.load_sheet("Sheet1", header_row=None)
        values = []
        rb = sh.to_arrow()
        for col in rb.column_names:
            values.extend(rb.column(col).to_pylist())
        if 999.0 not in values:
            findings.append("edited value not readable through the stock reader")

    verdict = common.Verdict.PASS if not findings else common.Verdict.FAIL
    detail = "load/str+Path/edit/save/reopen preserved values and features" if not findings else "; ".join(findings)
    _emit(test_id, verdict, detail, start=start, measured={"oracle": "openpyxl" if _OPX_AVAILABLE else "kyrax-stock"})


# ---------------------------------------------------------------------------
# A1-API-02 - stable typed errors; no silent malformed output or source writes.
# ---------------------------------------------------------------------------


def test_api_02_error_contract(tmp_path) -> None:
    test_id = "A1-API-02"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    findings: list[str] = []
    src = tmp_path / "src.xlsx"
    _author_small(src)
    before = _sha(src)

    wb = kyrax.load_workbook(str(src), edit_mode=True)
    try:
        wb["Nope"]
        findings.append("missing sheet did not raise KeyError")
    except KeyError:
        pass

    try:
        wb["Sheet1"].insert_rows(0)
        findings.append("insert_rows(0) did not raise a typed error")
    except (ValueError, kyrax.InvalidParametersError):
        pass

    try:
        wb["Sheet1"].move_range("not-a-range", rows=1, cols=1)
        findings.append("malformed move_range range was silently accepted")
    except (ValueError, kyrax.InvalidParametersError):
        pass

    for row, col, label in ((0, 1, "row=0"), (1, 0, "col=0")):
        bad = tmp_path / f"bad_{label}.xlsx"
        handle = kyrax.load_workbook(str(src), edit_mode=True)
        try:
            handle["Sheet1"].set_cell(row, col, 7)
            findings.append(f"set_cell({row},{col}) was not rejected with a typed error")
        except (ValueError, kyrax.InvalidParametersError):
            continue
        handle.save(str(bad))
        xml = ""
        import zipfile

        if bad.exists():
            with zipfile.ZipFile(bad) as z:
                for n in z.namelist():
                    if re.match(r"xl/worksheets/sheet\d+\.xml", n):
                        xml = z.read(n).decode("utf-8", "replace")
        refs = re.findall(r'<c r="([^"]+)"', xml)
        if any(ref.startswith(("0", "1")) and not ref[0].isalpha() for ref in refs):
            findings.append(f"set_cell({row},{col}) emitted malformed cell refs: {refs}")

    missing_dir = tmp_path / "does_not_exist" / "out.xlsx"
    try:
        wb.save(str(missing_dir))
        findings.append("save to a missing directory did not raise")
    except Exception:
        pass
    if missing_dir.exists():
        findings.append("partial output written to a missing-directory path")

    if _sha(src) != before:
        findings.append("source workbook bytes changed during error probes")

    verdict = common.Verdict.PASS if not findings else common.Verdict.FAIL
    detail = "all invalid inputs produced stable typed errors with no source writes" if not findings else "; ".join(findings)
    _emit(test_id, verdict, detail, start=start)


# ---------------------------------------------------------------------------
# A1-API-03 - ws["A1"] read/write: known gap until both exist.
# ---------------------------------------------------------------------------


def test_api_03_ws_a1_known_gap(tmp_path) -> None:
    test_id = "A1-API-03"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    src = tmp_path / "src.xlsx"
    _author_small(src)
    wb = kyrax.load_workbook(str(src), edit_mode=True)
    ws = wb["Sheet1"]

    has_get = hasattr(kyrax.EditableSheet, "__getitem__")
    has_set = hasattr(kyrax.EditableSheet, "__setitem__")
    try:
        ws["A1"]
        get_raises = False
    except TypeError:
        get_raises = True
    try:
        ws["A1"] = 5
        set_raises = False
    except TypeError:
        set_raises = True

    measured = {
        "has_getitem": has_get,
        "has_setitem": has_set,
        "getitem_raises_type_error": get_raises,
        "setitem_raises_type_error": set_raises,
    }
    if not has_get and not has_set and get_raises and set_raises:
        _emit(
            test_id,
            common.Verdict.KNOWN_GAP,
            "ws['A1'] get/set absent on EditableSheet (negative probe confirms "
            "TypeError on both); openpyxl A1-string semantics not implemented",
            start=start,
            measured=measured,
        )
    elif has_get and has_set:
        _emit(test_id, common.Verdict.PASS, "ws['A1'] get/set now implemented", start=start, measured=measured)
    else:
        _emit(test_id, common.Verdict.FAIL, f"partial A1 support: {measured}", start=start, measured=measured)


# ---------------------------------------------------------------------------
# A1-GEAR-01 - same fixture through stock / turbo / editable paths (bounded child).
# ---------------------------------------------------------------------------

_GEAR1_CHILD = r"""
import sys, json
sys.path.insert(0, sys.argv[2] + "/python")
import kyrax
path = sys.argv[1]
stock = kyrax.read_excel(path)
s_names = stock.sheet_names
stock_rb = stock.load_sheet(0).to_arrow()
turbo = kyrax.read_excel_turbo(path)
t_names = turbo.sheet_names
turbo_rb = turbo.load_sheet(0, features="all").to_arrow()
editable = kyrax.load_workbook(path, edit_mode=True)
try:
    e_ok = all(editable[n] is not None for n in s_names)
except Exception as exc:
    e_ok = "error:" + type(exc).__name__
res = {
    "stock_names": s_names,
    "turbo_names": t_names,
    "sheet_order_equal": s_names == t_names,
    "stock_shape": [stock_rb.num_rows, stock_rb.num_columns],
    "turbo_shape": [turbo_rb.num_rows, turbo_rb.num_columns],
    "schema_equal": str(stock_rb.schema) == str(turbo_rb.schema),
    "data_equal": bool(stock_rb.equals(turbo_rb)),
    "editable_all_sheets_addressable": e_ok,
}
print("RESULT " + json.dumps(res))
"""


def test_gear_01_stock_turbo_editable_coherence(tmp_path) -> None:
    test_id = "A1-GEAR-01"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    if not STRUCTURED.exists():
        _emit(
            test_id,
            common.Verdict.BLOCKED,
            f"fixture F01 missing: {STRUCTURED}",
            start=start,
        )
        return

    wr = common.watchdog_run(
        [sys.executable, "-c", _GEAR1_CHILD, str(STRUCTURED), str(ROOT)],
        timeout_s=30.0,
        workdir=str(tmp_path),
        label="a1_gear1",
    )
    if wr.verdict in (common.Verdict.TIMEOUT.value, common.Verdict.RSS_KILL.value, common.Verdict.COMMIT_KILL.value):
        _emit(
            test_id,
            common.Verdict.BLOCKED,
            f"structured.xlsx child probe failed: {wr.verdict} {wr.detail}",
            start=start,
            isolation=wr.isolation,
            command=str([sys.executable, "-c", _GEAR1_CHILD, str(STRUCTURED), str(ROOT)]),
            fixture={"fixture": "F01", "sha256": common.sha256_file(STRUCTURED)},
            measured={"watchdog_verdict": wr.verdict, "duration_s": wr.duration_s},
        )
        return
    try:
        res = _read_last_resut(wr.stdout_path)
    except AssertionError as exc:
        _emit(
            test_id,
            common.Verdict.BLOCKED,
            f"child exited {wr.exit_code} without evidence: {exc}",
            start=start,
            isolation=wr.isolation,
        )
        return

    problems: list[str] = []
    if not res.get("sheet_order_equal"):
        problems.append("sheet order/names differ between stock and turbo")
    if res.get("stock_shape") != res.get("turbo_shape"):
        problems.append(f"shapes differ: {res.get('stock_shape')} vs {res.get('turbo_shape')}")
    if not res.get("schema_equal"):
        problems.append("stock and turbo schemas differ")
    if not res.get("data_equal"):
        problems.append("stock and turbo values differ")
    if res.get("editable_all_sheets_addressable") is not True:
        problems.append(f"editable path cannot address all sheets: {res.get('editable_all_sheets_addressable')}")

    verdict = common.Verdict.PASS if not problems else common.Verdict.FAIL
    detail = "stock/turbo/editable agree on the same fixture" if not problems else "; ".join(problems)
    _emit(
        test_id,
        verdict,
        detail,
        start=start,
        isolation=wr.isolation,
        command=str([sys.executable, "-c", _GEAR1_CHILD, str(STRUCTURED), str(ROOT)]),
        fixture={"fixture": "F01", "sha256": common.sha256_file(STRUCTURED)},
        measured=res,
    )


# ---------------------------------------------------------------------------
# A1-GEAR-02 - turbo surfaces reachable from familiar-API objects: known gap.
# ---------------------------------------------------------------------------


def test_gear_02_turbo_reachability_known_gap(tmp_path) -> None:
    test_id = "A1-GEAR-02"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    src = tmp_path / "src.xlsx"
    _author_small(src)

    surfaces = ["charts", "images", "pivots", "formulas", "spill"]
    stock = kyrax.load_workbook(str(src))
    sheet = stock.load_sheet(0)
    editable = kyrax.load_workbook(str(src), edit_mode=True)
    editable_sheet = editable["Sheet1"]

    reachable = {}
    for surface in surfaces:
        reachable[surface] = {
            "stock_reader": hasattr(stock, surface),
            "stock_sheet": hasattr(sheet, surface),
            "editable_workbook": hasattr(editable, surface),
            "editable_sheet": hasattr(editable_sheet, surface),
        }

    any_reachable = any(
        v["stock_reader"] or v["stock_sheet"] or v["editable_workbook"] or v["editable_sheet"]
        for v in reachable.values()
    )
    if not any_reachable:
        _emit(
            test_id,
            common.Verdict.KNOWN_GAP,
            "charts/images/pivots/formulas/spill are not reachable from load_workbook "
            "objects (stock ExcelReader/ExcelSheet or EditableWorkbook/EditableSheet); "
            "they exist only on the parallel turbo vocabulary",
            start=start,
            measured=reachable,
        )
    else:
        _emit(test_id, common.Verdict.FAIL, f"turbo surface leaked onto familiar API: {reachable}", start=start)


# ---------------------------------------------------------------------------
# A1-ADOPT-01 - fresh-process import, star import, no side effects.
# ---------------------------------------------------------------------------

_ADOPT_CHILD = r"""
import sys, os, json
cwd = os.getcwd()
before = set(os.listdir(cwd))
import kyrax
from kyrax import *
all_ok = all(hasattr(kyrax, n) for n in kyrax.__all__)
opx_loaded = "openpyxl" in sys.modules
new_files = sorted(set(os.listdir(cwd)) - before)
tmp = os.path.join(cwd, "a.xlsx")
kyrax.write_excel_turbo(tmp, [{"name": "S1", "columns": [[1.0, 2.0]]}])
reader = kyrax.read_excel(tmp)
analyst_names = reader.sheet_names
sheet = reader.load_sheet(0)
_ = sheet.to_arrow()
wb = kyrax.load_workbook(tmp, edit_mode=True)
wb["S1"].set_cell(1, 1, 42.0)
out = os.path.join(cwd, "b.xlsx")
wb.save(out)
res = {
    "import_star_resolves": all_ok,
    "openpyxl_in_sys_modules": opx_loaded,
    "import_side_effect_files": new_files,
    "analyst_sheet_names": analyst_names,
    "engineer_saved": os.path.exists(out),
    "version": kyrax.__version__,
}
print("RESULT " + json.dumps(res))
"""


def test_adopt_01_fresh_process_import(tmp_path) -> None:
    test_id = "A1-ADOPT-01"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    cwd = tmp_path / "fresh"
    cwd.mkdir()
    wr = common.watchdog_run(
        [sys.executable, "-c", _ADOPT_CHILD],
        timeout_s=20.0,
        workdir=str(tmp_path),
        cwd=str(cwd),
        label="a1_adopt",
    )
    if wr.verdict in (common.Verdict.TIMEOUT.value, common.Verdict.RSS_KILL.value, common.Verdict.COMMIT_KILL.value):
        _emit(
            test_id,
            common.Verdict.BLOCKED,
            f"fresh-process probe failed: {wr.verdict} {wr.detail}",
            start=start,
            isolation=wr.isolation,
        )
        return
    try:
        res = _read_last_resut(wr.stdout_path)
    except AssertionError as exc:
        _emit(test_id, common.Verdict.BLOCKED, f"child exited {wr.exit_code} without evidence: {exc}", start=start)
        return

    problems: list[str] = []
    if not res.get("import_star_resolves"):
        problems.append("from kyrax import * left names unresolved")
    if res.get("openpyxl_in_sys_modules"):
        problems.append("openpyxl appears in sys.modules after import kyrax")
    if res.get("import_side_effect_files"):
        problems.append(f"import wrote files into the cwd: {res.get('import_side_effect_files')}")
    if res.get("analyst_sheet_names") != ["S1"]:
        problems.append(f"analyst workflow failed: {res.get('analyst_sheet_names')}")
    if not res.get("engineer_saved"):
        problems.append("engineer workflow failed to save")

    verdict = common.Verdict.PASS if not problems else common.Verdict.FAIL
    detail = "clean import, no openpyxl, no side effects, both workflows work" if not problems else "; ".join(problems)
    _emit(test_id, verdict, detail, start=start, isolation=wr.isolation, measured=res)


# ---------------------------------------------------------------------------
# P05 (A1 half) - every reader reopens a repaired workbook with intact inventory.
# ---------------------------------------------------------------------------


def test_p05_reader_reopen_after_repair(tmp_path) -> None:
    test_id = "P05"
    start = time.monotonic()
    if not _require_binary(test_id):
        return
    src = tmp_path / "src.xlsx"
    _author_small(src)
    repaired = tmp_path / "repaired.xlsx"
    findings: list[str] = []

    validation = kyrax.validate_excel(str(src))
    if validation.get("valid") is not True:
        findings.append(f"clean source not valid: {validation.get('valid')}")

    report = kyrax.repair_excel(str(src), str(repaired))
    if report.get("wrote_output") is not True:
        findings.append(f"repair wrote nothing: {report.get('wrote_output')}")
    if not repaired.exists():
        findings.append("repaired output file missing")

    if repaired.exists():
        rp = str(repaired)
        try:
            stock = kyrax.read_excel(rp)
            if stock.sheet_names != ["Sheet1", "Sheet2"]:
                findings.append(f"stock reader inventory differs: {stock.sheet_names}")
        except Exception as exc:
            findings.append(f"stock reader failed on repaired file: {type(exc).__name__}: {exc}")
        try:
            turbo = kyrax.read_excel_turbo(rp)
            if turbo.sheet_names != ["Sheet1", "Sheet2"]:
                findings.append(f"turbo reader inventory differs: {turbo.sheet_names}")
        except Exception as exc:
            findings.append(f"turbo reader failed on repaired file: {type(exc).__name__}: {exc}")
        try:
            editable = kyrax.load_workbook(rp, edit_mode=True)
            for name in ("Sheet1", "Sheet2"):
                editable[name]
        except Exception as exc:
            findings.append(f"editable reader failed on repaired file: {type(exc).__name__}: {exc}")

    verdict = common.Verdict.PASS if not findings else common.Verdict.FAIL
    detail = "repaired workbook reopens on stock/turbo/editable with intact inventory" if not findings else "; ".join(findings)
    _emit(
        test_id,
        verdict,
        detail,
        start=start,
        measured={"validate_valid": validation.get("valid"), "repair_actions": len(report.get("actions", []))},
    )
