"""A3 Wave-1 — overlay, mutation, refshift, preservation (architecture stress).

Owned by lane A3. Wave-1 is a small, self-contained baseline that uses only
verified public kyrax APIs (``kyrax.load_workbook(..., edit_mode=True)`` /
``kyrax.edit_excel``) plus openpyxl for authoring sources and reading back
structural results (openpyxl is a structural/range oracle here, never a formula
calculation oracle).

Verdict discipline (per corrected plan):
  * PASS  - the probe asserts an implemented, satisfied invariant.
  * xfail(strict=True) - a KNOWN current defect that is an agreed FAIL/S1 gate
    (stale formula caches never invalidated; retained invalid signatures).
    Strict xfail keeps the defect visible: an XPASS turns the test red.
  * skip(reason=...) - a named dependency that does not exist in this wave
    (A6 harness / COM / F05-F07 / cross-lane P02/P06 fixtures).

No A6 ``common.py``/``fixtures.py`` exists yet, so every ZIP/XML helper is
self-contained here. Nothing writes to shared fixtures; all sources are built
in ``tmp_path``.
"""

import hashlib
import zipfile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

import kyrax


# ---------------------------------------------------------------------------
# Self-contained ZIP/XML helpers (A6 common.py absent in Wave 1)
# ---------------------------------------------------------------------------


def _member_hashes(path):
    """{entry name: sha256 of inflated content} for every ZIP member."""
    with zipfile.ZipFile(path) as z:
        return {
            i.filename: hashlib.sha256(z.read(i.filename)).hexdigest()
            for i in z.infolist()
        }


def _changed_members(src, out):
    """Names of members whose inflated content differs between two files."""
    a, b = _member_hashes(src), _member_hashes(out)
    names = sorted(set(a) | set(b))
    return [n for n in names if a.get(n) != b.get(n)]


def _sheet_xml(path, sheet_no=1):
    """Raw bytes of ``xl/worksheets/sheet{sheet_no}.xml``."""
    name = f"xl/worksheets/sheet{sheet_no}.xml"
    with zipfile.ZipFile(path) as z:
        return z.read(name)


def _build_formula_sheet(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=A1+A2"
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# A3-OVR-01  no-op edit: member preservation
# ---------------------------------------------------------------------------


def test_ovr01_noop_edit_preserves_every_member(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _build_formula_sheet(str(src))

    ed = kyrax.edit_excel(str(src))
    ed.save(str(out))

    with zipfile.ZipFile(str(out)) as z:
        assert z.testzip() is None
    # Untouched members must be byte-identical (inflated). A whole-file byte
    # comparison is not required: the container may carry a documented minimal
    # ZIP-directory delta, so member-content hashes are the invariant.
    assert _changed_members(src, out) == [], _changed_members(src, out)


# ---------------------------------------------------------------------------
# A3-OVR-02  single numeric-cell edit: only the target sheet may change
# ---------------------------------------------------------------------------


def test_ovr02_single_numeric_cell_changes_only_target_sheet(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["B2"] = 2
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].set_cell(1, 1, 1000.0)
    ed.save(str(out))

    changed = _changed_members(src, out)
    assert changed == ["xl/worksheets/sheet1.xml"], changed


# ---------------------------------------------------------------------------
# A3-OVR-04  document properties survive no-op and edit
# ---------------------------------------------------------------------------


def test_ovr04_docprops_stable_through_noop_and_edit(tmp_path):
    src = tmp_path / "src.xlsx"
    noop = tmp_path / "noop.xlsx"
    edited = tmp_path / "edited.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    wb.properties.creator = "A3-test"
    wb.properties.title = "docprops probe"
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed.save(str(noop))

    ed = kyrax.edit_excel(str(src))
    ed["Data"].set_cell(1, 1, 5.0)
    ed.save(str(edited))

    for target in (noop, edited):
        changed = _changed_members(src, target)
        assert "docProps/core.xml" not in changed
        assert "docProps/app.xml" not in changed
        with zipfile.ZipFile(str(target)) as z:
            assert "docProps/core.xml" in z.namelist()
            assert "docProps/app.xml" in z.namelist()


# ---------------------------------------------------------------------------
# A3-MUT-01  formula text translation + stale-cache gate probe
# ---------------------------------------------------------------------------


def test_mut01_formula_body_and_cell_move_on_insert(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _build_formula_sheet(str(src))

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_rows(1, 1)
    ed.save(str(out))

    xml = _sheet_xml(out).decode("utf-8")
    assert '<c r="A4"' in xml, xml
    assert "<f>A2+A3</f>" in xml, xml
    assert "<f>A1+A2</f>" not in xml, xml


def _seed_stale_cache(path, out):
    """Copy ``path`` to ``out``, inject a nonempty cached `<v>` into the A3
    formula cell, and strip ``fullCalcOnLoad="1"`` from ``xl/workbook.xml``.
    Forces the engine to both clear the stale cached result and guarantee
    recalc itself. The injection accepts both the self-closing `<v />` form
    openpyxl emits and the paired `<v></v>` form."""
    with zipfile.ZipFile(str(path)) as zin, zipfile.ZipFile(str(out), "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                for old in (b"<f>A1+A2</f><v />", b"<f>A1+A2</f><v></v>"):
                    if old in data:
                        data = data.replace(old, b"<f>A1+A2</f><v>3</v>")
                        break
                assert b"<f>A1+A2</f><v>3</v>" in data, (
                    "failed to seed stale cache: formula cell not found in source"
                )
            elif item.filename == "xl/workbook.xml":
                data = data.replace(b' fullCalcOnLoad="1"', b"")
            zout.writestr(item, data)


def test_mut01_fail_probe_stale_cache_requires_full_calc_on_load(tmp_path):
    src = tmp_path / "src.xlsx"
    seeded = tmp_path / "seeded.xlsx"
    out = tmp_path / "out.xlsx"
    _build_formula_sheet(str(src))
    _seed_stale_cache(str(src), str(seeded))

    ed = kyrax.edit_excel(str(seeded))
    ed["Data"].insert_rows(1, 1)
    ed.save(str(out))

    with zipfile.ZipFile(str(out)) as z:
        wb_xml = z.read("xl/workbook.xml")
        sheet_xml = z.read("xl/worksheets/sheet1.xml")
    assert b'fullCalcOnLoad="1"' in wb_xml, wb_xml
    # The moved formula (A1+A2 -> A2+A3) must carry NO cached result: the seeded
    # <v>3</v> is gone and no <v> follows the formula.
    assert b"<f>A2+A3</f>" in sheet_xml, sheet_xml
    assert b"<f>A2+A3</f><v" not in sheet_xml, sheet_xml
    assert b"<f>A2+A3</f></c>" in sheet_xml, sheet_xml


# ---------------------------------------------------------------------------
# A3-MUT-02  merges: band policy
# ---------------------------------------------------------------------------


def test_mut02_merge_follows_insert_above(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "h"
    ws["A2"] = "m"
    ws.merge_cells("A2:A3")
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_rows(1, 1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    assert [str(m) for m in ld["Data"].merged_cells.ranges] == ["A3:A4"]
    ld.close()


def test_mut02_merge_trimmed_by_delete_inside(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "m"
    ws.merge_cells("A1:A5")
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].delete_rows(3, 1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    assert [str(m) for m in ld["Data"].merged_cells.ranges] == ["A1:A4"]
    ld.close()


# ---------------------------------------------------------------------------
# A3-MUT-06  boundary moves and typed refusal with source preservation
# ---------------------------------------------------------------------------


def _refusal_asserts(src, src_before, out, exc, ed):
    with pytest.raises(exc):
        ed.save(str(out))
    assert src.read_bytes() == src_before, "source file must be untouched"
    assert not out.exists(), "no output may be written on a refusal"


def test_mut06_row_overflow_refuses_and_preserves_source(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1048576, column=1, value=1)
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_rows(1048576, 1)
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


def test_mut06_col_overflow_at_xfd_refuses_and_preserves_source(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1, column=16384, value=1)
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].insert_cols(16384, 1)
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


def test_mut06_move_range_out_of_grid_refuses_and_preserves_source(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1048576, column=1, value=1)
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].move_range("A1048576:A1048576", rows=1, cols=0)
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


def test_mut06_insert_at_row1_and_colA_succeed(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "a"
    ws["B1"] = 10
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    sh = ed["Data"]
    sh.insert_rows(1, 1)
    sh.insert_cols(1, 1)
    ed.save(str(out))

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["B2"].value == "a"
    assert ws["C2"].value == 10
    ld.close()


def test_mut06_table_header_delete_refuses_and_preserves_source(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "H1"
    ws["B1"] = "H2"
    ws["A2"] = 1
    ws["B2"] = 2
    tab = Table(displayName="Tab1", ref="A1:B2")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    wb.save(str(src))
    wb.close()
    src_before = src.read_bytes()

    ed = kyrax.edit_excel(str(src))
    ed["Data"].delete_rows(1, 1)
    _refusal_asserts(src, src_before, out, kyrax.InvalidParametersError, ed)


# ---------------------------------------------------------------------------
# P01 (small evidence)  mutate one cell-everything sheet; member allowlist
# ---------------------------------------------------------------------------


def test_p01_mutation_keeps_coupled_features_and_member_allowlist(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    # Table headings live at A1/B1 and must be strings to avoid openpyxl's
    # header-coercion warning; numeric inputs and the formula move off the
    # header row (G1/G2 inputs, H1 formula referencing them).
    ws["G1"] = 1
    ws["G2"] = 2
    ws["H1"] = "=G1+G2"
    ws.merge_cells("E2:E3")
    ws["F1"].hyperlink = "https://example.com/x"
    ws.auto_filter.ref = "A1:B10"
    tab = Table(displayName="Tab1", ref="A1:B3")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws["A1"] = "H1"
    ws["B1"] = "H2"
    ws["A2"] = 10
    ws["B2"] = 20
    ws["A3"] = 30
    ws["B3"] = 40
    ws.add_table(tab)
    wb.save(str(src))
    wb.close()

    ed = kyrax.edit_excel(str(src))
    sh = ed["Data"]
    sh.insert_rows(2, 1)
    sh.set_cell(1, 7, 1000.0)
    ed.save(str(out))

    # Member allowlist: target sheet, its table part, and the workbook (the
    # mutation shifts the table's workbook-level defined-name metadata in
    # xl/workbook.xml). Every other member must be byte-identical.
    changed = _changed_members(src, out)
    assert set(changed) == {
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
        "xl/tables/table1.xml",
    }, changed

    ld = load_workbook(str(out))
    ws = ld["Data"]
    assert ws["G1"].value == 1000.0
    assert ws.tables["Tab1"].ref == "A1:B4"
    assert [str(m) for m in ws.merged_cells.ranges] == ["E3:E4"]
    assert ws.auto_filter.ref == "A1:B11"
    assert ws["F1"].hyperlink.target == "https://example.com/x"
    assert ws["H1"].value == "=G1+G3"
    ld.close()


# ---------------------------------------------------------------------------
# A3-OVR-03  hostile parse guard (named dependency skip for the synthetic
# fallback probe; small truncated-zip load refusal kept as evidence)
# ---------------------------------------------------------------------------

try:
    from kyrax import _kyrax as _kx

    HAS_KYRAX_BINDINGS = hasattr(_kx, "is_signed_workbook") and hasattr(_kx, "signature_info")
except Exception:  # pragma: no cover - build-dependent
    HAS_KYRAX_BINDINGS = False


def test_ovr03_truncated_zip_load_refuses(tmp_path):
    bad = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    wb.save(str(bad))
    wb.close()
    data = bad.read_bytes()[: len(bad.read_bytes()) // 2]
    bad.write_bytes(data)

    with pytest.raises(Exception):
        kyrax.edit_excel(str(bad))


@pytest.mark.skip(
    reason=(
        "A3-OVR-03 parse-fallback probe: needs the overlay parse-fallback gate "
        "(typed refusal vs silent synthetic sheet on defective <sheetData>) - "
        "source dependency; A6 common.py absent in Wave 1."
    )
)
def test_ovr03_parse_fallback_probe():
    pass


# ---------------------------------------------------------------------------
# A3-SIG-01  retained invalid signature = agreed FAIL/S1 gate (strict xfail)
# ---------------------------------------------------------------------------


def _add_synthetic_signature(path, out):
    """Copy ``path`` to ``out`` and inject signature parts + origin rel."""
    sig = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b"<Signature xmlns=\"http://www.w3.org/2000/09/xmldsig#\">"
        b"<SignatureValue>fake</SignatureValue></Signature>"
    )
    origin_rel = (
        b'<Relationship Id="rIdSIG" Type='
        b'"http://schemas.openxmlformats.org/package/2006/relationships/'
        b'digital-signature/origin" Target="_xmlsignatures/origin.sigs"/>'
    )
    with zipfile.ZipFile(str(path)) as zin, zipfile.ZipFile(str(out), "w") as zout:
        for item in zin.infolist():
            if item.filename == "_rels/.rels":
                root_rels = zin.read(item.filename).replace(
                    b"</Relationships>", origin_rel + b"</Relationships>"
                )
                zout.writestr(item, root_rels)
            else:
                zout.writestr(item, zin.read(item.filename))
        zout.writestr("_xmlsignatures/sig1.xml", sig)
        zout.writestr(
            "_xmlsignatures/origin.sigs",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b"<SignatureOriginList "
            b'xmlns="http://schemas.openxmlformats.org/package/2006/digital-signature" '
            b'><SignatureV1 Id="sig1" '
            b'Uri="/_xmlsignatures/sig1.xml" /></SignatureOriginList>',
        )


@pytest.mark.skipif(
    not HAS_KYRAX_BINDINGS,
    reason="kyrax._kyrax bindings unavailable in this build",
)
def test_sig01_fail_probe_edit_removes_signature_parts(tmp_path):
    """A3-SIG-01 (hard gate): a real content edit atomically removes the
    invalid digital signature.

    After an unrelated-value edit, every ``_xmlsignatures/*`` member is gone,
    the signature-origin/signature relationships are gone from ``_rels/.rels``,
    signature content-type declarations are gone from ``[Content_Types].xml``,
    and the output member set is the input set minus exactly the two signature
    parts -- no unrelated member is added or dropped.
    """
    src = tmp_path / "src.xlsx"
    signed = tmp_path / "signed.xlsx"
    out = tmp_path / "out.xlsx"
    _build_formula_sheet(str(src))
    _add_synthetic_signature(src, signed)

    assert _kx.is_signed_workbook(str(signed)) is True

    ed = kyrax.edit_excel(str(signed))
    ed["Data"].set_cell(1, 1, 999.0)
    ed.save(str(out))

    with zipfile.ZipFile(str(signed)) as z:
        input_members = set(z.namelist())
    with zipfile.ZipFile(str(out)) as z:
        members = set(z.namelist())
        assert not any(n.startswith("_xmlsignatures/") for n in members)
        rels = z.read("_rels/.rels")
        ct = z.read("[Content_Types].xml")

    assert _kx.is_signed_workbook(str(out)) is False
    assert b"digital-signature/origin" not in rels
    assert b'Target="_xmlsignatures/' not in rels
    assert b"/_xmlsignatures/" not in ct
    assert b'Extension="sigs"' not in ct
    assert members == input_members - {
        "_xmlsignatures/sig1.xml",
        "_xmlsignatures/origin.sigs",
    }


# ---------------------------------------------------------------------------
# Named dependency skips for cross-lane / later-wave work
# ---------------------------------------------------------------------------


@pytest.mark.skip(
    reason="P02 hydration half is lane A4 and needs A4 HYD baselines + A6 excel_com runner."
)
def test_p02_a3_half_edit_inputs_recalc_file():
    pass


@pytest.mark.skip(
    reason="P06 authored-image baseline is lane A2; image anchor-shift probe needs F04 corpus (A6)."
)
def test_p06_a3_half_image_anchor_shift_probe():
    pass


@pytest.mark.skip(
    reason="A3-PRV-01/02 and A3-AUTH-01 need the F05 preserve corpus generated by A6 fixtures.py."
)
def test_prv_auth_corpus_probes():
    pass
