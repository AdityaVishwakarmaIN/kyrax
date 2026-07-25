import pytest
import tempfile
from pathlib import Path
import kyrax

def test_load_modify_save():
    with tempfile.TemporaryDirectory() as tmpdir:
        # Step 1: Create initial excel file using write_excel_turbo
        init_path = Path(tmpdir) / "original.xlsx"
        modified_path = Path(tmpdir) / "modified.xlsx"

        kyrax.write_excel_turbo(
            str(init_path),
            [{"name": "Sheet1", "columns": [[1.0, 2.0], [10.0, 20.0]]}],
        )
        assert init_path.exists()

        # Step 2: Open with load_workbook in edit_mode
        wb = kyrax.load_workbook(init_path, edit_mode=True)
        ws = wb["Sheet1"]
        ws.set_cell(row=1, col=1, value=999.0)
        wb.save(str(modified_path))

        assert modified_path.exists()

        # Step 3: Verify with openpyxl if installed
        try:
            import openpyxl
            wb_read = openpyxl.load_workbook(modified_path, data_only=True)
            val = wb_read["Sheet1"].cell(row=1, column=1).value
            assert val == 999.0, f"Expected 999.0, got {val}"
        except ImportError:
            pass
