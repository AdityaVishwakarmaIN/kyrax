import tempfile
from pathlib import Path
import kyrax

def test_set_cell_style_persistence():
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / "original.xlsx"
        output_path = Path(tmpdir) / "styled_output.xlsx"

        # 1. Create a base workbook
        sheets_data = [{
            "name": "Sheet1",
            "rows": [
                ["Header1", "Header2"],
                [100.0, 200.0],
            ]
        }]
        kyrax.write_excel_turbo(str(input_path), sheets_data)

        # 2. Open via edit_excel / load_workbook, set style on A1 and B2
        wb = kyrax.load_workbook(str(input_path), edit_mode=True)
        ws = wb["Sheet1"]
        ws.set_cell(1, 1, "Modified Header")
        ws.set_cell_style(1, 1, font={"bold": True, "name": "Arial", "sz": 14})
        ws.set_cell_style(2, 2, fill={"patternType": "solid", "fg": "FFFF0000"})

        wb.save(str(output_path))
        assert output_path.exists()

        # 3. Verify via openpyxl
        try:
            import openpyxl
            wb_check = openpyxl.load_workbook(output_path)
            ws_check = wb_check["Sheet1"]
            assert ws_check.cell(row=1, column=1).value == "Modified Header"
            assert ws_check.cell(row=1, column=1).font.bold is True
            assert ws_check.cell(row=1, column=1).font.name == "Arial"
            assert ws_check.cell(row=2, column=2).value == 200.0
            assert ws_check.cell(row=2, column=2).fill.fill_type == "solid"
        except ImportError:
            pass
