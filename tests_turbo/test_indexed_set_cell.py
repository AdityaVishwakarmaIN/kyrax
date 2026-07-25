import tempfile
from pathlib import Path
import kyrax

def test_indexed_set_cell_and_style():
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "binary_search_set_cell.xlsx"

        # Create workbook and populate out-of-order cells via set_cell logic
        sheets_data = [{
            "name": "Sheet1",
            "cell_styles": [
                {"row": 5, "col": 5, "style": {"font": {"bold": True}}},
                {"row": 1, "col": 1, "style": {"font": {"italic": True}}},
            ],
            "rows": [
                [1.0, 2.0],
                [3.0, 4.0],
            ]
        }]

        kyrax.write_excel_turbo(str(output_path), sheets_data)
        assert output_path.exists()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Sheet1"]
            assert ws.cell(row=1, column=1).value == 1.0
            assert ws.cell(row=5, column=5).font.bold == True
        except ImportError:
            pass
