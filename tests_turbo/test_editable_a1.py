from pathlib import Path

import openpyxl
import pytest

import kyrax


def _author_book(path: Path) -> None:
    kyrax.write_excel_turbo(
        str(path),
        [
            {
                "name": "Sheet1",
                "columns": [
                    [1.0, 2.0, 3.0],
                    [True, False, None],
                    ["x", "y", "z"],
                ],
            }
        ],
    )


def _editable(path: Path):
    return kyrax.load_workbook(path, edit_mode=True)["Sheet1"]


def _range_values(ws, key):
    """Value matrix of a range key: tuple-of-tuples of Cells -> nested lists of values."""
    return [[cell.value for cell in row] for row in ws[key]]


def test_editable_a1_scalar_and_range_reads(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    _author_book(path)
    ws = _editable(path)

    # Scalar __getitem__ returns a Cell; the value lives at .value.
    assert ws["A1"].value == 1.0
    assert ws["a2"].value == 2.0
    assert ws["B1"].value is True
    assert ws["C2"].value == "y"
    assert ws["D4"].value is None
    # Range __getitem__ returns tuple-of-tuples of Cells, normalized to top-left.
    assert _range_values(ws, "B2:C3") == [[False, "y"], [None, "z"]]
    assert _range_values(ws, "$C$3:$B$2") == [[False, "y"], [None, "z"]]


def test_editable_a1_set_is_immediately_visible_and_round_trips(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _author_book(source)
    workbook = kyrax.load_workbook(source, edit_mode=True)
    ws = workbook["Sheet1"]

    ws["A1"] = 11.5
    ws["B2:C3"] = [[10, "alpha"], [20, "beta"]]

    assert ws["A1"].value == 11.5
    assert _range_values(ws, "B2:C3") == [[10.0, "alpha"], [20.0, "beta"]]
    assert _range_values(ws, "$C$3:$B$2") == [[10.0, "alpha"], [20.0, "beta"]]

    workbook.save(str(output))
    reopened = openpyxl.load_workbook(output, data_only=False)
    sheet = reopened["Sheet1"]
    assert sheet["A1"].value == 11.5
    assert sheet["B2"].value == 10.0
    assert sheet["C2"].value == "alpha"
    assert sheet["B3"].value == 20.0
    assert sheet["C3"].value == "beta"
    reopened.close()


class _BadValue:
    def __str__(self) -> str:
        raise RuntimeError("cannot stringify")


def test_editable_a1_range_assignment_is_atomic(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    _author_book(path)
    ws = _editable(path)

    # Range assignment validates the full matrix before any mutation, so compare
    # value snapshots, not freshly allocated Cell object identity.
    before = _range_values(ws, "A1:B1")
    with pytest.raises(TypeError):
        ws["A1:B1"] = [[99.0]]
    assert _range_values(ws, "A1:B1") == before

    with pytest.raises(TypeError):
        ws["A1:B1"] = [[99.0, _BadValue()]]
    assert _range_values(ws, "A1:B1") == before


@pytest.mark.parametrize(
    "key",
    ["", "A0", "XFE1", "A1048577", "0", "$A", "A1:B2:C3", " A1", "A-1", "A"],
)
def test_editable_a1_rejects_malformed_or_out_of_grid_keys(
    tmp_path: Path, key: str
) -> None:
    path = tmp_path / "source.xlsx"
    _author_book(path)
    ws = _editable(path)

    with pytest.raises(ValueError):
        _ = ws[key]
    with pytest.raises(ValueError):
        ws[key] = 1


@pytest.mark.parametrize(
    "row,col",
    [(0, 1), (1, 0), (1_048_577, 1), (1, 16_385)],
)
def test_editable_coordinate_methods_reject_out_of_grid(
    tmp_path: Path, row: int, col: int
) -> None:
    path = tmp_path / "source.xlsx"
    _author_book(path)
    ws = _editable(path)

    with pytest.raises(ValueError):
        ws.set_cell(row, col, 1)
    with pytest.raises(ValueError):
        ws.set_cell_style(row, col, font={"bold": True})


def test_editable_move_range_rejects_malformed_range(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    _author_book(path)
    ws = _editable(path)

    with pytest.raises(ValueError):
        ws.move_range("not-a-range", rows=1, cols=1)
