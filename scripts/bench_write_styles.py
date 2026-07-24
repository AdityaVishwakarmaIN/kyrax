"""W2 write bench: styles-dense 100k vs openpyxl + unstyled 1M numeric sanity.

Median of 3. Run from repo root:
  PYTHONUTF8=1 ./.venv/Scripts/python scripts/bench_write_styles.py
"""

from __future__ import annotations

import statistics
import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from kyrax import write_excel_turbo

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, numbers

N_RUNS = 3
N_DENSE = 100_000
N_STYLES = 200
N_NUMERIC = 1_000_000


def median(xs: list[float]) -> float:
    return statistics.median(xs)


def time_runs(fn, runs: int = N_RUNS) -> float:
    times = []
    for _ in range(runs):
        t0 = time.perf_counter()
        fn()
        times.append(time.perf_counter() - t0)
    return median(times)


def make_palette(n: int) -> list[dict]:
    palette = []
    for i in range(n):
        fg = f"{(i * 37) % 256:02X}{(i * 59) % 256:02X}{(i * 97) % 256:02X}"
        palette.append(
            {
                "font": {
                    "name": ["Calibri", "Arial", "Consolas", "Georgia", "Tahoma"][i % 5],
                    "sz": [9, 10, 11, 12, 14, 16][i % 6],
                    "bold": i % 3 == 0,
                    "color": fg,
                },
                "fill": {"patternType": "solid", "fg": fg},
                "num_fmt": (
                    f'0.000"u{i % 17}"'
                    if i % 7 == 0
                    else ["General", "0.00", "0%", "#,##0", "0.00E+00", "mm-dd-yy"][i % 6]
                ),
            }
        )
    return palette


def bench_styles_dense():
    print(f"=== styles_dense {N_DENSE} cells / ~{N_STYLES} styles ===")
    values = list(range(N_DENSE))
    palette = make_palette(N_STYLES)

    def turbo():
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_excel_turbo(
                path,
                [
                    {
                        "name": "Dense",
                        "columns": [values],
                        "style_palette": palette,
                    }
                ],
            )
        finally:
            Path(path).unlink(missing_ok=True)

    def opx_normal():
        wb = Workbook()
        ws = wb.active
        ws.title = "Dense"
        fonts = []
        fills = []
        for i in range(N_STYLES):
            fg = f"{(i * 37) % 256:02X}{(i * 59) % 256:02X}{(i * 97) % 256:02X}"
            fonts.append(
                Font(
                    name=["Calibri", "Arial", "Consolas", "Georgia", "Tahoma"][i % 5],
                    size=[9, 10, 11, 12, 14, 16][i % 6],
                    bold=i % 3 == 0,
                    color=fg,
                )
            )
            fills.append(PatternFill("solid", fgColor=fg))
        for i, v in enumerate(values, start=1):
            cell = ws.cell(i, 1, v)
            si = (i - 1) % N_STYLES
            cell.font = fonts[si]
            cell.fill = fills[si]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
        finally:
            Path(path).unlink(missing_ok=True)

    t_turbo = time_runs(turbo)
    print(f"  turbo write:     {t_turbo:.3f} s (median of {N_RUNS})")
    t_norm = time_runs(opx_normal)
    print(f"  openpyxl normal: {t_norm:.3f} s  ({t_norm / t_turbo:.1f}x turbo)")
    return {"turbo": t_turbo, "opx_normal": t_norm}


def bench_numeric_unstyled():
    print(f"=== unstyled numeric {N_NUMERIC} (W1 non-regression) ===")
    values = list(range(N_NUMERIC))

    def turbo():
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_excel_turbo(path, [{"name": "N", "columns": [values]}])
        finally:
            Path(path).unlink(missing_ok=True)

    t_turbo = time_runs(turbo)
    print(f"  turbo write:     {t_turbo:.3f} s (median of {N_RUNS})")
    return {"turbo": t_turbo}


def main():
    print("bench_write_styles.py — median of", N_RUNS)
    r1 = bench_styles_dense()
    r2 = bench_numeric_unstyled()
    print()
    print("SUMMARY")
    print(
        f"styles_dense 100k: turbo {r1['turbo']:.3f}s | "
        f"opx {r1['opx_normal']:.3f}s ({r1['opx_normal']/r1['turbo']:.1f}x)"
    )
    print(f"numeric 1M unstyled: turbo {r2['turbo']:.3f}s")


if __name__ == "__main__":
    main()
