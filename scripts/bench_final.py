"""Final nextexcel benchmarks: stock kyrax vs turbo vs openpyxl-naive.

Run from repo root:
  .venv/Scripts/python scripts/bench_final.py

Stock kyrax 0.20.2 is timed via system `python` (subprocess; wall time
measured inside the child). Turbo and openpyxl use the venv interpreter.
"""

from __future__ import annotations

import json
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TESTDATA = ROOT / "testdata"
sys.path.insert(0, str(ROOT / "python"))

import kyrax
from kyrax import read_excel_turbo

# ---------------------------------------------------------------------------
# Fixture metadata
# ---------------------------------------------------------------------------

FIXTURES = [
    "mixed",
    "strings_shared",
    "styled",
    "formulas",
    "structured",
    "comments",
]

# Selective feature set for A.4 (mixed / strings_shared skip)
SELECTIVE: dict[str, list[str]] = {
    "styled": ["styles"],
    "formulas": ["formulas"],
    "structured": ["merges", "hyperlinks", "tables", "defined_names"],
    "comments": ["comments"],
}

# openpyxl naive only for feature-heavy fixtures (B)
OPENPYXL_FIXTURES = ["styled", "formulas", "structured", "comments"]

N_TIMED = 3  # best-of-N after one warm-up


def path_for(name: str) -> Path:
    return TESTDATA / f"{name}.xlsx"


def cells_of(name: str) -> int:
    """Cell count from turbo values-only (header excluded, matches arrow)."""
    s = read_excel_turbo(str(path_for(name))).load_sheet(0, features="values")
    return int(s.nrows) * int(s.ncols)


def cells_per_sec(seconds: float, n_cells: int) -> float:
    if seconds <= 0:
        return float("inf")
    return n_cells / seconds


def fmt_s(s: float | None) -> str:
    if s is None:
        return "—"
    if s < 0.01:
        return f"{s:.4f}"
    return f"{s:.3f}"


def fmt_cps(cps: float | None) -> str:
    if cps is None:
        return "—"
    if cps >= 1e6:
        return f"{cps / 1e6:.2f}M"
    if cps >= 1e3:
        return f"{cps / 1e3:.1f}k"
    return f"{cps:.0f}"


def fmt_x(ratio: float | None) -> str:
    if ratio is None:
        return "—"
    return f"{ratio:.1f}x"


# ---------------------------------------------------------------------------
# A.1 stock kyrax via system python (time INSIDE subprocess)
# ---------------------------------------------------------------------------

STOCK_CHILD = r"""
import json, sys, time
import kyrax

path = sys.argv[1]
n = int(sys.argv[2])

# Verify we are stock (no turbo)
assert not hasattr(kyrax, "read_excel_turbo"), "system python has turbo — wrong env"
ver = getattr(kyrax, "__version__", "?")

# warm-up
kyrax.read_excel(path).load_sheet(0).to_arrow()

times = []
for _ in range(n):
    t0 = time.perf_counter()
    kyrax.read_excel(path).load_sheet(0).to_arrow()
    times.append(time.perf_counter() - t0)

print(json.dumps({"version": ver, "times": times, "best": min(times)}))
"""


def verify_stock_python() -> str:
    """Confirm system `python` has kyrax 0.20.2 without turbo."""
    code = (
        "import kyrax, json; "
        "print(json.dumps({"
        "'version': kyrax.__version__, "
        "'has_turbo': hasattr(kyrax, 'read_excel_turbo')"
        "}))"
    )
    r = subprocess.run(
        ["python", "-c", code],
        capture_output=True,
        text=True,
        check=False,
    )
    if r.returncode != 0:
        raise RuntimeError(f"system python kyrax check failed:\n{r.stderr}")
    info = json.loads(r.stdout.strip())
    if info["version"] != "0.20.2":
        raise RuntimeError(f"expected stock kyrax 0.20.2, got {info['version']}")
    if info["has_turbo"]:
        raise RuntimeError("system python kyrax has read_excel_turbo — not stock")
    return info["version"]


def bench_stock(path: Path) -> float:
    r = subprocess.run(
        ["python", "-c", STOCK_CHILD, str(path), str(N_TIMED)],
        capture_output=True,
        text=True,
        check=False,
    )
    if r.returncode != 0:
        raise RuntimeError(
            f"stock bench failed for {path.name}:\nstdout={r.stdout}\nstderr={r.stderr}"
        )
    # last line is JSON
    line = [ln for ln in r.stdout.strip().splitlines() if ln.strip()][-1]
    data = json.loads(line)
    return float(data["best"])


# ---------------------------------------------------------------------------
# Turbo helpers
# ---------------------------------------------------------------------------

def materialize_all(path: Path) -> None:
    """features='all' + force every accessor so nothing lazy is skipped."""
    reader = read_excel_turbo(str(path))
    sheet = reader.load_sheet(0, features="all")
    _ = sheet.to_arrow()
    _ = sheet.cell_errors()
    # styles
    st = sheet.style_table()
    si = sheet.style_indices()
    if st is not None:
        _ = len(st)
    if si is not None:
        _ = len(si)
        if si:
            _ = len(si[0])
    ns = sheet.named_styles()
    if ns is not None:
        _ = len(ns)
    # formulas (includes shared translations) — RecordBatch
    formulas = sheet.formulas()
    if formulas is not None:
        _ = formulas.num_rows
        if formulas.num_rows > 0:
            _ = formulas.column("text")[0].as_py()
    # merges
    merges = sheet.merges()
    if merges is not None:
        _ = len(merges)
    # workbook-level defined names + tables
    names = reader.defined_names()
    if names is not None:
        _ = len(names)
    tables = sheet.tables()
    if tables is None:
        tables = reader.tables()
    if tables is not None:
        _ = len(tables)
    # hyperlinks
    hlinks = sheet.hyperlinks()
    if hlinks is not None:
        _ = len(hlinks)
    # comments — RecordBatch + threaded
    comments = sheet.comments()
    if comments is not None:
        _ = comments.num_rows
        if comments.num_rows > 0:
            _ = comments.column("text")[0].as_py()
    authors = sheet.comment_authors()
    if authors is not None:
        _ = len(authors)
    thr = sheet.threaded_comments()
    if thr is not None:
        _ = len(thr)
    persons = reader.persons()
    if persons is not None:
        _ = len(persons)
    # Stream A/B/C gap surfaces
    for fn in (
        sheet.row_dimensions,
        sheet.column_dimensions,
        sheet.sheet_format,
        sheet.auto_filter,
        sheet.sheet_view,
        sheet.protection,
        sheet.page_setup,
        sheet.page_margins,
        sheet.print_options,
        sheet.header_footer,
        sheet.data_validations,
        sheet.conditional_formatting,
        sheet.charts,
        sheet.pivots,
    ):
        v = fn()
        if v is not None:
            _ = len(v) if hasattr(v, "__len__") else v
    _ = sheet.freeze_panes()
    _ = sheet.code_name
    _ = sheet.tab_color
    _ = sheet.sheet_state
    _ = sheet.sheet_kind
    wp = reader.workbook_props()
    if wp is not None:
        _ = len(wp)
    _ = reader.date1904
    _ = reader.has_vba
    vba = reader.vba_project()
    if vba is not None:
        _ = len(vba)


def turbo_values(path: Path) -> None:
    read_excel_turbo(str(path)).load_sheet(0, features="values").to_arrow()


def turbo_selective(path: Path, features: list[str]) -> None:
    reader = read_excel_turbo(str(path))
    sheet = reader.load_sheet(0, features=features)
    _ = sheet.to_arrow()
    # touch the requested feature accessors so cost is real
    if "styles" in features:
        _ = sheet.style_table()
        _ = sheet.style_indices()
    if "formulas" in features:
        f = sheet.formulas()
        if f is not None:
            _ = f.num_rows
            if f.num_rows > 0:
                _ = f.column("text")[0].as_py()
    if "merges" in features:
        m = sheet.merges()
        if m is not None:
            _ = len(m)
    if "hyperlinks" in features:
        h = sheet.hyperlinks()
        if h is not None:
            _ = len(h)
    if "tables" in features:
        t = sheet.tables()
        if t is None:
            t = reader.tables()
        if t is not None:
            _ = len(t)
    if "defined_names" in features:
        n = reader.defined_names()
        if n is not None:
            _ = len(n)
    if "comments" in features:
        c = sheet.comments()
        if c is not None:
            _ = c.num_rows
            if c.num_rows > 0:
                _ = c.column("text")[0].as_py()
        _ = sheet.comment_authors()


def best_of(fn, n: int = N_TIMED) -> float:
    """One warm-up, then best-of-n wall times."""
    fn()  # warm-up
    samples: list[float] = []
    for _ in range(n):
        t0 = time.perf_counter()
        fn()
        samples.append(time.perf_counter() - t0)
    return min(samples)


# ---------------------------------------------------------------------------
# B. openpyxl naive routes (1 run each; faithful to lab measurements)
# ---------------------------------------------------------------------------

def openpyxl_styled(path: Path) -> float:
    """Full load + iterate values and style attributes (number_format/font/fill)."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            _ = cell.number_format
            _ = cell.font
            _ = cell.fill
    wb.close()
    return time.perf_counter() - t0


def openpyxl_formulas(path: Path) -> float:
    """Double load (data_only False then True), iterating all cell values."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
    wb.close()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
    wb.close()
    return time.perf_counter() - t0


def openpyxl_structured(path: Path) -> float:
    """Full load + merges/hyperlinks/names/tables extraction + value iteration."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    # merges
    _ = [str(r) for r in ws.merged_cells.ranges]
    # defined names (global + sheet)
    _ = dict(wb.defined_names)
    for sn in wb.sheetnames:
        _ = dict(wb[sn].defined_names)
    # tables (TableList.items() yields name→ref str; index for Table objects)
    tables = getattr(ws, "tables", None) or {}
    for tname in list(tables):
        tbl = tables[tname]
        _ = (tname, tbl.ref, list(tbl.tableColumns))
    # value + hyperlink iteration (openpyxl needs per-cell scan for hyperlinks)
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            _ = cell.hyperlink
    wb.close()
    return time.perf_counter() - t0


def openpyxl_comments(path: Path) -> float:
    """Full load + read all comments."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            c = cell.comment
            if c is not None:
                _ = (c.text, c.author)
                n += 1
    wb.close()
    return time.perf_counter() - t0


OPENPYXL_FNS = {
    "styled": openpyxl_styled,
    "formulas": openpyxl_formulas,
    "structured": openpyxl_structured,
    "comments": openpyxl_comments,
}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    print("=" * 72)
    print("nextexcel final benchmarks")
    print(f"repo        {ROOT}")
    print(f"venv turbo  kyrax {kyrax.__version__}  turbo={hasattr(kyrax, 'read_excel_turbo')}")
    stock_ver = verify_stock_python()
    print(f"stock py    kyrax {stock_ver}  (system python, no turbo)")
    import openpyxl

    print(f"openpyxl    {openpyxl.__version__}")
    print(f"method      best-of-{N_TIMED} after 1 warm-up (openpyxl: 1 run)")
    print("=" * 72)

    # cell counts
    cell_counts: dict[str, int] = {}
    for name in FIXTURES:
        cell_counts[name] = cells_of(name)
        p = path_for(name)
        print(f"  {name:16s}  {cell_counts[name]:>10,} cells  ({p.stat().st_size / 1e6:.1f} MB)")
    print()

    # results collectors
    stock: dict[str, float] = {}
    turbo_vals: dict[str, float] = {}
    turbo_all: dict[str, float] = {}
    turbo_sel: dict[str, float] = {}
    oxl: dict[str, float] = {}

    # ---- A: stock + turbo ----
    print("--- A. stock kyrax / turbo values / turbo all / turbo selective ---")
    for name in FIXTURES:
        p = path_for(name)
        n_cells = cell_counts[name]
        print(f"\n[{name}] ({n_cells:,} cells)")

        # A.1 stock
        print("  stock kyrax value-only ...", flush=True)
        t = bench_stock(p)
        stock[name] = t
        print(f"    stock values:     {fmt_s(t)} s   {fmt_cps(cells_per_sec(t, n_cells))} cells/s")

        # A.2 turbo values
        print("  turbo values-only ...", flush=True)
        t = best_of(lambda: turbo_values(p))
        turbo_vals[name] = t
        print(f"    turbo values:     {fmt_s(t)} s   {fmt_cps(cells_per_sec(t, n_cells))} cells/s")

        # A.3 turbo all
        print("  turbo ALL features (materialized) ...", flush=True)
        t = best_of(lambda: materialize_all(p))
        turbo_all[name] = t
        print(f"    turbo all:        {fmt_s(t)} s   {fmt_cps(cells_per_sec(t, n_cells))} cells/s")

        # A.4 selective
        if name in SELECTIVE:
            feats = SELECTIVE[name]
            print(f"  turbo selective {feats} ...", flush=True)
            t = best_of(lambda f=feats: turbo_selective(p, f))
            turbo_sel[name] = t
            print(f"    turbo selective:  {fmt_s(t)} s   {fmt_cps(cells_per_sec(t, n_cells))} cells/s")
        else:
            print("  turbo selective: skip (mixed/strings_shared)")

    # ---- B: openpyxl naive ----
    print("\n--- B. openpyxl naive (1 run each) ---")
    for name in OPENPYXL_FIXTURES:
        p = path_for(name)
        n_cells = cell_counts[name]
        print(f"  openpyxl {name} ...", flush=True)
        t = OPENPYXL_FNS[name](p)
        oxl[name] = t
        print(f"    openpyxl naive:   {fmt_s(t)} s   {fmt_cps(cells_per_sec(t, n_cells))} cells/s")

    # ---- C: feature cost ----
    print("\n--- C. Feature-cost: turbo-all vs turbo-values ---")
    print(f"{'fixture':16s}  {'values_s':>10s}  {'all_s':>10s}  {'delta_s':>10s}  {'overhead_%':>10s}")
    for name in FIXTURES:
        v = turbo_vals[name]
        a = turbo_all[name]
        d = a - v
        pct = (d / v * 100.0) if v > 0 else 0.0
        print(f"{name:16s}  {v:10.3f}  {a:10.3f}  {d:10.3f}  {pct:9.1f}%")

    # ---- Summary tables ----
    print("\n" + "=" * 72)
    print("SUMMARY TABLE A — stock / turbo-values / turbo-all / selective")
    print(
        f"{'fixture':16s}  {'stock_s':>8s}  {'t_vals_s':>8s}  {'t_all_s':>8s}  "
        f"{'t_sel_s':>8s}  {'vals/stock':>10s}  {'all/stock':>10s}"
    )
    for name in FIXTURES:
        s, v, a = stock[name], turbo_vals[name], turbo_all[name]
        sel = turbo_sel.get(name)
        print(
            f"{name:16s}  {fmt_s(s):>8s}  {fmt_s(v):>8s}  {fmt_s(a):>8s}  "
            f"{fmt_s(sel):>8s}  {fmt_x(s / v if v else None):>10s}  "
            f"{fmt_x(s / a if a else None):>10s}"
        )

    print("\nSUMMARY TABLE A (cells/s)")
    print(
        f"{'fixture':16s}  {'stock':>10s}  {'t_vals':>10s}  {'t_all':>10s}  {'t_sel':>10s}"
    )
    for name in FIXTURES:
        n = cell_counts[name]
        sel = turbo_sel.get(name)
        print(
            f"{name:16s}  {fmt_cps(cells_per_sec(stock[name], n)):>10s}  "
            f"{fmt_cps(cells_per_sec(turbo_vals[name], n)):>10s}  "
            f"{fmt_cps(cells_per_sec(turbo_all[name], n)):>10s}  "
            f"{fmt_cps(cells_per_sec(sel, n) if sel else None):>10s}"
        )

    print("\nSUMMARY TABLE B — openpyxl naive vs turbo-all")
    print(
        f"{'fixture':16s}  {'openpyxl_s':>10s}  {'turbo_all_s':>11s}  {'speedup':>8s}"
    )
    for name in OPENPYXL_FIXTURES:
        o, a = oxl[name], turbo_all[name]
        print(f"{name:16s}  {fmt_s(o):>10s}  {fmt_s(a):>11s}  {fmt_x(o / a if a else None):>8s}")

    # machine-readable dump for RESULTS doc generation
    payload = {
        "machine": "12 cores, Windows",
        "stock_version": stock_ver,
        "turbo_version": kyrax.__version__,
        "openpyxl_version": openpyxl.__version__,
        "cell_counts": cell_counts,
        "stock_s": stock,
        "turbo_values_s": turbo_vals,
        "turbo_all_s": turbo_all,
        "turbo_selective_s": turbo_sel,
        "openpyxl_s": oxl,
        "selective_features": SELECTIVE,
    }
    out = ROOT / "scripts" / "_bench_final_raw.json"
    out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"\nraw JSON → {out}")
    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
