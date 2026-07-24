"""Gap / coherence benchmark refresh (MERGE M3).

Re-runs the six original fixture benches (stock / turbo-values / turbo-all),
compares against RESULTS_NEXTEXCEL.md baselines, and measures openpyxl-naive
vs turbo-all on gap fixtures (richmeta, sheetmeta, threaded, charts).

Run from repo root:
  .venv/Scripts/python scripts/bench_gap.py
"""

from __future__ import annotations

import json
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TESTDATA = ROOT / "testdata"
REPORTS = ROOT.parent / "nextexcel-reports"
sys.path.insert(0, str(ROOT / "python"))

import kyrax
from kyrax import read_excel_turbo

# Import shared helpers from bench_final
sys.path.insert(0, str(ROOT / "scripts"))
from bench_final import (  # noqa: E402
    FIXTURES,
    N_TIMED,
    SELECTIVE,
    bench_stock,
    best_of,
    cells_of,
    cells_per_sec,
    fmt_cps,
    fmt_s,
    fmt_x,
    materialize_all,
    path_for,
    turbo_selective,
    turbo_values,
    verify_stock_python,
)

# Prior RESULTS_NEXTEXCEL.md wall seconds (best-of-3)
PRIOR = {
    "mixed": {"stock": 2.876, "turbo_values": 1.246, "turbo_all": 1.126},
    "strings_shared": {"stock": 4.254, "turbo_values": 0.881, "turbo_all": 1.031},
    "styled": {"stock": 5.372, "turbo_values": 0.879, "turbo_all": 1.068},
    "formulas": {"stock": 2.089, "turbo_values": 1.006, "turbo_all": 0.745},
    "structured": {"stock": 4.659, "turbo_values": 1.213, "turbo_all": 1.636},
    "comments": {"stock": 0.260, "turbo_values": 0.123, "turbo_all": 0.268},
}

REGRESSION_PCT = 10.0  # flag if new > prior * (1 + REGRESSION_PCT/100)

GAP_FIXTURES = {
    "richmeta": TESTDATA / "gap_richmeta.xlsx",
    "sheetmeta": TESTDATA / "gap_sheetmeta.xlsx",
    "threaded": TESTDATA / "threaded.xlsx",
    "charts": TESTDATA / "charts.xlsx",
}


def materialize_all_sheets(path: Path) -> None:
    """features=all on every sheet (charts workbook has chartsheets)."""
    reader = read_excel_turbo(str(path))
    for name in reader.sheet_names:
        sheet = reader.load_sheet(name, features="all")
        _ = sheet.to_arrow()
        _ = sheet.cell_errors()
        for fn in (
            sheet.style_table,
            sheet.style_indices,
            sheet.named_styles,
            sheet.formulas,
            sheet.merges,
            sheet.hyperlinks,
            sheet.comments,
            sheet.comment_authors,
            sheet.threaded_comments,
            sheet.charts,
            sheet.pivots,
            sheet.tables,
            sheet.row_dimensions,
            sheet.column_dimensions,
            sheet.sheet_format,
            sheet.auto_filter,
            sheet.sheet_view,
            sheet.protection,
            sheet.page_setup,
            sheet.page_margins,
            sheet.print_options,
            sheet.header_footer,
            sheet.data_validations,
            sheet.conditional_formatting,
        ):
            v = fn()
            if v is not None and hasattr(v, "num_rows"):
                _ = v.num_rows
            elif v is not None and hasattr(v, "__len__"):
                _ = len(v)
        _ = sheet.freeze_panes()
        _ = sheet.code_name
        _ = sheet.tab_color
    _ = reader.defined_names()
    _ = reader.tables()
    _ = reader.workbook_props()
    _ = reader.persons()
    _ = reader.has_vba
    _ = reader.vba_project()
    _ = reader.date1904


# ---------------------------------------------------------------------------
# openpyxl naive routes for gap fixtures
# ---------------------------------------------------------------------------


def openpyxl_richmeta(path: Path) -> float:
    """Full load + iterate value+border+alignment+named style + validations + CF."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            _ = cell.border
            _ = cell.alignment
            _ = cell.style  # named style
            _ = cell.number_format
            _ = cell.font
            _ = cell.fill
            _ = cell.protection
    # data validations
    dvs = getattr(ws, "data_validations", None)
    if dvs is not None:
        data = getattr(dvs, "dataValidation", None) or getattr(dvs, "data_validations", None) or []
        try:
            for dv in data:
                _ = (dv.sqref, dv.type, dv.formula1, dv.formula2)
        except TypeError:
            _ = list(dvs)
    # conditional formatting
    cf = getattr(ws, "conditional_formatting", None)
    if cf is not None:
        try:
            for sqref, rules in cf._cf_rules.items():  # noqa: SLF001
                for rule in rules:
                    _ = (sqref, getattr(rule, "type", None), getattr(rule, "dxf", None))
        except Exception:
            _ = list(cf)
    wb.close()
    return time.perf_counter() - t0


def openpyxl_sheetmeta(path: Path) -> float:
    """Full load + extract dims/views/protection/pagesetup on all sheets."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    _ = wb.properties
    _ = wb.epoch
    for name in wb.sheetnames:
        ws = wb[name]
        if getattr(ws, "sheet_state", None) is not None:
            _ = ws.sheet_state
        # chartsheet has no row dims
        if not hasattr(ws, "row_dimensions"):
            continue
        for r, dim in ws.row_dimensions.items():
            _ = (r, dim.height, dim.hidden)
        for letter, dim in ws.column_dimensions.items():
            _ = (letter, dim.width, dim.hidden)
        _ = ws.auto_filter.ref if ws.auto_filter else None
        _ = ws.freeze_panes
        _ = ws.sheet_view
        _ = ws.protection.sheet
        _ = ws.page_setup.orientation
        _ = ws.page_margins.left
        _ = ws.print_options.horizontalCentered
        _ = ws.oddHeader.center if ws.oddHeader else None
        _ = ws.sheet_format.defaultRowHeight
        # values pass so route is honest full-load+work
        for row in ws.iter_rows():
            for cell in row:
                _ = cell.value
    wb.close()
    return time.perf_counter() - t0


def openpyxl_threaded(path: Path) -> float:
    """Full load + all comments (openpyxl sees only legacy mirrors; no threaded)."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            c = cell.comment
            if c is not None:
                _ = (c.text, c.author)
                n += 1
    # openpyxl has no persons / threaded API
    _ = hasattr(wb, "persons")
    wb.close()
    return time.perf_counter() - t0


def openpyxl_charts(path: Path) -> float:
    """Full load including charts on all sheets (read_only=False required)."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    total_charts = 0
    for name in wb.sheetnames:
        ws = wb[name]
        charts = getattr(ws, "_charts", None) or []
        total_charts += len(charts)
        for ch in charts:
            _ = type(ch).__name__
            _ = getattr(ch, "title", None)
        if hasattr(ws, "iter_rows"):
            try:
                for row in ws.iter_rows():
                    for cell in row:
                        _ = cell.value
            except Exception:
                pass
    wb.close()
    return time.perf_counter() - t0


GAP_OXL = {
    "richmeta": openpyxl_richmeta,
    "sheetmeta": openpyxl_sheetmeta,
    "threaded": openpyxl_threaded,
    "charts": openpyxl_charts,
}


def main() -> int:
    print("=" * 72)
    print("nextexcel gap / coherence benchmarks (M3)")
    print(f"repo        {ROOT}")
    print(f"venv turbo  kyrax {kyrax.__version__}  turbo={hasattr(kyrax, 'read_excel_turbo')}")
    stock_ver = verify_stock_python()
    print(f"stock py    kyrax {stock_ver}")
    import openpyxl

    print(f"openpyxl    {openpyxl.__version__}")
    print(f"method      best-of-{N_TIMED} after 1 warm-up (openpyxl gap: 1 run)")
    print("=" * 72)

    cell_counts: dict[str, int] = {}
    for name in FIXTURES:
        cell_counts[name] = cells_of(name)
        p = path_for(name)
        print(f"  {name:16s}  {cell_counts[name]:>10,} cells  ({p.stat().st_size / 1e6:.1f} MB)")
    print()

    stock: dict[str, float] = {}
    turbo_vals: dict[str, float] = {}
    turbo_all: dict[str, float] = {}
    turbo_sel: dict[str, float] = {}

    print("--- A. stock / turbo values / turbo all (six original fixtures) ---")
    for name in FIXTURES:
        p = path_for(name)
        n_cells = cell_counts[name]
        print(f"\n[{name}] ({n_cells:,} cells)")

        print("  stock ...", flush=True)
        t = bench_stock(p)
        stock[name] = t
        print(f"    stock values:  {fmt_s(t)} s")

        print("  turbo values ...", flush=True)
        t = best_of(lambda: turbo_values(p))
        turbo_vals[name] = t
        print(f"    turbo values:  {fmt_s(t)} s")

        print("  turbo all ...", flush=True)
        t = best_of(lambda: materialize_all(p))
        turbo_all[name] = t
        print(f"    turbo all:     {fmt_s(t)} s")

        if name in SELECTIVE:
            feats = SELECTIVE[name]
            t = best_of(lambda f=feats: turbo_selective(p, f))
            turbo_sel[name] = t
            print(f"    turbo sel:     {fmt_s(t)} s  {feats}")

    # Regression check
    print("\n--- A.reg regression vs RESULTS_NEXTEXCEL.md (>10% slower flags) ---")
    regressions: list[dict] = []
    print(
        f"{'fixture':16s}  {'metric':14s}  {'prior_s':>8s}  {'now_s':>8s}  {'delta_%':>8s}  {'verdict':8s}"
    )
    for name in FIXTURES:
        for metric, now_map in (
            ("turbo_values", turbo_vals),
            ("turbo_all", turbo_all),
        ):
            prior = PRIOR[name][metric]
            now = now_map[name]
            pct = ((now - prior) / prior) * 100.0 if prior > 0 else 0.0
            bad = pct > REGRESSION_PCT
            verdict = "REGRESS" if bad else "ok"
            if bad:
                regressions.append(
                    {
                        "fixture": name,
                        "metric": metric,
                        "prior": prior,
                        "now": now,
                        "pct": pct,
                    }
                )
            print(
                f"{name:16s}  {metric:14s}  {prior:8.3f}  {now:8.3f}  {pct:7.1f}%  {verdict}"
            )

    # Gap fixtures
    print("\n--- B. openpyxl-naive vs turbo-all on gap fixtures ---")
    gap_oxl: dict[str, float] = {}
    gap_turbo: dict[str, float] = {}
    gap_notes: dict[str, str] = {}

    for name, path in GAP_FIXTURES.items():
        print(f"\n[{name}] {path.name}")
        print("  openpyxl naive (1 run) ...", flush=True)
        t_oxl = GAP_OXL[name](path)
        gap_oxl[name] = t_oxl
        print(f"    openpyxl:  {fmt_s(t_oxl)} s")

        print("  turbo all best-of-3 ...", flush=True)
        t_t = best_of(lambda p=path: materialize_all_sheets(p))
        gap_turbo[name] = t_t
        print(f"    turbo-all: {fmt_s(t_t)} s   speedup {fmt_x(t_oxl / t_t if t_t else None)}")

        if name == "threaded":
            gap_notes[name] = (
                "openpyxl sees only legacy comment mirrors; no persons/threaded API. "
                "turbo surfaces both legacy + threaded + persons under features=all."
            )
        elif name == "charts":
            gap_notes[name] = (
                "openpyxl full mode (not read_only) required for worksheet charts; "
                "turbo single path surfaces charts on worksheets + chartsheets."
            )
        else:
            gap_notes[name] = ""

    # Summary tables
    print("\n" + "=" * 72)
    print("SUMMARY A — six fixtures (stock / turbo-values / turbo-all)")
    print(
        f"{'fixture':16s}  {'stock_s':>8s}  {'t_vals_s':>8s}  {'t_all_s':>8s}  "
        f"{'vals/stock':>10s}  {'all/stock':>10s}"
    )
    for name in FIXTURES:
        s, v, a = stock[name], turbo_vals[name], turbo_all[name]
        print(
            f"{name:16s}  {fmt_s(s):>8s}  {fmt_s(v):>8s}  {fmt_s(a):>8s}  "
            f"{fmt_x(s / v if v else None):>10s}  {fmt_x(s / a if a else None):>10s}"
        )

    print("\nSUMMARY B — gap openpyxl vs turbo-all")
    print(f"{'fixture':12s}  {'openpyxl_s':>10s}  {'turbo_all_s':>11s}  {'speedup':>8s}")
    for name in GAP_FIXTURES:
        o, a = gap_oxl[name], gap_turbo[name]
        print(f"{name:12s}  {fmt_s(o):>10s}  {fmt_s(a):>11s}  {fmt_x(o / a if a else None):>8s}")

    payload = {
        "machine": "12 cores, Windows",
        "stock_version": stock_ver,
        "turbo_version": kyrax.__version__,
        "openpyxl_version": openpyxl.__version__,
        "cell_counts": cell_counts,
        "stock_s": stock,
        "turbo_values_s": turbo_vals,
        "turbo_all_s": turbo_all,
        "turbo_selective_s": turbo_sel,
        "prior": PRIOR,
        "regressions": regressions,
        "gap_openpyxl_s": gap_oxl,
        "gap_turbo_all_s": gap_turbo,
        "gap_notes": gap_notes,
    }
    out = ROOT / "scripts" / "_bench_gap_raw.json"
    out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"\nraw JSON → {out}")
    print(f"regressions: {len(regressions)}")
    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
