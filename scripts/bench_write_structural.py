"""W3 write benches: plain, feature-dense, 100 charts vs openpyxl."""
from __future__ import annotations

import os
import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from kyrax import write_excel_turbo

from tests_turbo.test_write_structural import _base_dense_sheet


def turbo_plain(path: str) -> None:
    write_excel_turbo(
        path,
        [
            {
                "name": "S",
                "columns": [
                    ["Category", "A", "B", "C", "D"],
                    ["Sales", 10, 20, 30, 40],
                    ["Cost", 5, 10, 15, 20],
                ],
            }
        ],
    )


def turbo_dense(path: str) -> None:
    write_excel_turbo(
        path,
        [_base_dense_sheet()],
        props={"title": "T", "creator": "C", "custom": {"Dept": "E"}},
        defined_names=[{"name": "MyRange", "value": "'Data'!$A$1:$B$2"}],
        lock_structure=True,
        features="all",
    )


def turbo_charts100(path: str) -> None:
    cat = "'Data'!$A$2:$A$5"
    val = "'Data'!$B$2:$B$5"
    charts = [
        {
            "type": "col",
            "title": f"C{i}",
            "series": [{"cat_ref": cat, "val_ref": val}],
            "anchor": {"cell": "E3", "width_cm": 10, "height_cm": 6},
        }
        for i in range(100)
    ]
    write_excel_turbo(
        path,
        [
            {
                "name": "Data",
                "columns": [
                    ["Category", "A", "B", "C", "D"],
                    ["Sales", 10, 20, 30, 40],
                    ["Cost", 5, 10, 15, 20],
                ],
                "charts": charts,
            }
        ],
        features="all",
    )


def o_charts100(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["B1"] = "Sales"
    for i, c in enumerate(["A", "B", "C", "D"], 1):
        ws.cell(i + 1, 1, c)
        ws.cell(i + 1, 2, i * 10)
    for i in range(100):
        ch = BarChart()
        ch.title = f"C{i}"
        ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=5), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=5))
        ch.anchor = "E3"
        ws.add_chart(ch)
    wb.save(path)


def med(fn, n=5) -> float:
    times = []
    for _ in range(n):
        fd, p = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        t0 = time.perf_counter()
        fn(p)
        t1 = time.perf_counter()
        times.append((t1 - t0) * 1000)
        try:
            os.remove(p)
        except OSError:
            pass
    times.sort()
    return times[len(times) // 2]


def main() -> None:
    plain = med(turbo_plain)
    dense = med(turbo_dense)
    tp = med(turbo_charts100)
    op = med(o_charts100)
    print(f"plain_ms {plain:.2f}")
    print(f"dense_ms {dense:.2f}")
    print(f"dense_overhead_ms {dense - plain:.2f}")
    print(f"charts100_turbo_ms {tp:.2f}")
    print(f"charts100_openpyxl_ms {op:.2f}")
    print(f"charts_per_s_turbo {100 / (tp / 1000):.1f}")
    print(f"charts_per_s_openpyxl {100 / (op / 1000):.1f}")
    print(f"speedup {op / tp:.2f}x")


if __name__ == "__main__":
    main()
