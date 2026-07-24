"""S2 production verdict benches: openpyxl naive vs turbo-all vs stock kyrax.

Run from repo root:
  .venv/Scripts/python scripts/bench_production.py

- openpyxl naive: 1 run each (slow)
- turbo features="all" full materialize: best-of-3 after 1 warm-up
- stock kyrax value-only: system python subprocess, best-of-3 after warm-up
"""

from __future__ import annotations

import json
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TESTDATA = ROOT / "testdata"
REPORTS = ROOT.parent / "nextexcel-reports"
sys.path.insert(0, str(ROOT / "python"))

import kyrax
from kyrax import read_excel_turbo

# Four big fidelity fixtures + richmeta
BIG = ["styled", "formulas", "structured", "comments"]
N_TIMED = 3


def path_for(name: str) -> Path:
    if name == "richmeta":
        return TESTDATA / "gap_richmeta.xlsx"
    return TESTDATA / f"{name}.xlsx"


def cells_of(name: str) -> int:
    s = read_excel_turbo(str(path_for(name))).load_sheet(0, features="values")
    return int(s.nrows) * int(s.ncols)


STOCK_CHILD = r"""
import json, sys, time
import kyrax
path = sys.argv[1]
n = int(sys.argv[2])
assert not hasattr(kyrax, "read_excel_turbo"), "system python has turbo"
kyrax.read_excel(path).load_sheet(0).to_arrow()  # warm
times = []
for _ in range(n):
    t0 = time.perf_counter()
    kyrax.read_excel(path).load_sheet(0).to_arrow()
    times.append(time.perf_counter() - t0)
print(json.dumps({"version": getattr(kyrax, "__version__", "?"), "times": times, "best": min(times)}))
"""


def bench_stock(path: Path) -> float:
    r = subprocess.run(
        ["python", "-c", STOCK_CHILD, str(path), str(N_TIMED)],
        capture_output=True,
        text=True,
        check=False,
    )
    if r.returncode != 0:
        raise RuntimeError(f"stock bench failed: {r.stderr}\n{r.stdout}")
    line = [ln for ln in r.stdout.strip().splitlines() if ln.strip()][-1]
    return float(json.loads(line)["best"])


def materialize_all(path: Path) -> None:
    """features='all' + force every accessor; formulas() called once (cached)."""
    reader = read_excel_turbo(str(path))
    sheet = reader.load_sheet(0, features="all")
    _ = sheet.to_arrow()
    _ = sheet.cell_errors()
    st = sheet.style_table()
    si = sheet.style_indices()
    if st is not None:
        _ = len(st)
    if si is not None:
        _ = len(si)
        if si:
            _ = len(si[0])
    ns = sheet.named_styles()
    if ns is not None:
        _ = len(ns)
    # formulas once — OnceLock cache in rust
    formulas = sheet.formulas()
    if formulas is not None:
        _ = formulas.num_rows
        if formulas.num_rows > 0:
            _ = formulas.column("text")[0].as_py()
    merges = sheet.merges()
    if merges is not None:
        _ = len(merges)
    names = reader.defined_names()
    if names is not None:
        _ = len(names)
    tables = sheet.tables()
    if tables is None:
        tables = reader.tables()
    if tables is not None:
        _ = len(tables)
    hlinks = sheet.hyperlinks()
    if hlinks is not None:
        _ = len(hlinks)
    comments = sheet.comments()
    if comments is not None:
        _ = comments.num_rows
        if comments.num_rows > 0:
            _ = comments.column("text")[0].as_py()
    authors = sheet.comment_authors()
    if authors is not None:
        _ = len(authors)
    thr = sheet.threaded_comments()
    if thr is not None:
        _ = len(thr)
    persons = reader.persons()
    if persons is not None:
        _ = len(persons)
    for fn in (
        sheet.row_dimensions,
        sheet.column_dimensions,
        sheet.sheet_format,
        sheet.auto_filter,
        sheet.sheet_view,
        sheet.protection,
        sheet.page_setup,
        sheet.page_margins,
        sheet.print_options,
        sheet.header_footer,
        sheet.data_validations,
        sheet.conditional_formatting,
        sheet.charts,
        sheet.pivots,
    ):
        v = fn()
        if v is not None:
            _ = len(v) if hasattr(v, "__len__") else v
    _ = sheet.freeze_panes()
    _ = sheet.code_name
    _ = sheet.tab_color
    _ = sheet.sheet_state
    _ = sheet.sheet_kind
    wp = reader.workbook_props()
    if wp is not None:
        _ = len(wp)
    _ = reader.date1904
    _ = reader.has_vba
    vba = reader.vba_project()
    if vba is not None:
        _ = len(vba)


def best_of(fn, n: int = N_TIMED) -> float:
    fn()
    samples = []
    for _ in range(n):
        t0 = time.perf_counter()
        fn()
        samples.append(time.perf_counter() - t0)
    return min(samples)


def openpyxl_styled(path: Path) -> float:
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            _ = cell.number_format
            _ = cell.font
            _ = cell.fill
    wb.close()
    return time.perf_counter() - t0


def openpyxl_formulas(path: Path) -> float:
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
    wb.close()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
    wb.close()
    return time.perf_counter() - t0


def openpyxl_structured(path: Path) -> float:
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    _ = [str(r) for r in ws.merged_cells.ranges]
    _ = dict(wb.defined_names)
    for sn in wb.sheetnames:
        _ = dict(wb[sn].defined_names)
    tables = getattr(ws, "tables", None) or {}
    for tname in list(tables):
        tbl = tables[tname]
        _ = (tname, tbl.ref, list(tbl.tableColumns))
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            _ = cell.hyperlink
    wb.close()
    return time.perf_counter() - t0


def openpyxl_comments(path: Path) -> float:
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            c = cell.comment
            if c is not None:
                _ = (c.text, c.author)
    wb.close()
    return time.perf_counter() - t0


def openpyxl_richmeta(path: Path) -> float:
    """Full load + value/border/alignment/named style + validations + CF."""
    import openpyxl

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            _ = cell.value
            _ = cell.number_format
            _ = cell.font
            _ = cell.fill
            _ = cell.border
            _ = cell.alignment
            _ = cell.style
    _ = list(ws.data_validations.dataValidation) if ws.data_validations else []
    _ = list(ws.conditional_formatting)
    wb.close()
    return time.perf_counter() - t0


OXL_FNS = {
    "styled": openpyxl_styled,
    "formulas": openpyxl_formulas,
    "structured": openpyxl_structured,
    "comments": openpyxl_comments,
    "richmeta": openpyxl_richmeta,
}


def main() -> int:
    import openpyxl

    print("=" * 72)
    print("nextexcel S2 production verdict benches")
    print(f"repo        {ROOT}")
    print(f"venv turbo  kyrax {kyrax.__version__} turbo={hasattr(kyrax, 'read_excel_turbo')}")
    print(f"openpyxl    {openpyxl.__version__}")
    r = subprocess.run(
        [
            "python",
            "-c",
            "import kyrax,json; print(json.dumps({'v':kyrax.__version__,'t':hasattr(kyrax,'read_excel_turbo')}))",
        ],
        capture_output=True,
        text=True,
        check=True,
    )
    stock_info = json.loads(r.stdout.strip())
    print(f"stock py    kyrax {stock_info['v']} has_turbo={stock_info['t']}")
    print(f"method      turbo best-of-{N_TIMED} after warm-up; openpyxl 1 run; stock best-of-{N_TIMED}")
    print("=" * 72)

    results: dict = {"meta": {}, "fixtures": {}}
    results["meta"] = {
        "turbo_version": kyrax.__version__,
        "openpyxl": openpyxl.__version__,
        "stock_version": stock_info["v"],
        "method": f"turbo best-of-{N_TIMED}; openpyxl 1-run; stock best-of-{N_TIMED}",
    }

    all_names = BIG + ["richmeta"]
    for name in all_names:
        p = path_for(name)
        n_cells = cells_of(name)
        size_mb = p.stat().st_size / 1e6
        print(f"\n[{name}] {n_cells:,} cells  {size_mb:.2f} MB  {p.name}", flush=True)

        row: dict = {"cells": n_cells, "size_mb": size_mb, "file": p.name}

        print("  stock kyrax value-only ...", flush=True)
        t_stock = bench_stock(p)
        row["stock_s"] = t_stock
        print(f"    stock:     {t_stock:.4f} s", flush=True)

        print("  turbo features=all (materialize, formulas once) ...", flush=True)
        t_turbo = best_of(lambda: materialize_all(p))
        row["turbo_all_s"] = t_turbo
        print(f"    turbo-all: {t_turbo:.4f} s", flush=True)

        print("  openpyxl naive (1 run) ...", flush=True)
        t_oxl = OXL_FNS[name](p)
        row["openpyxl_s"] = t_oxl
        print(f"    openpyxl:  {t_oxl:.4f} s", flush=True)

        row["oxl_over_turbo"] = t_oxl / t_turbo if t_turbo > 0 else None
        row["stock_over_turbo"] = t_stock / t_turbo if t_turbo > 0 else None
        row["turbo_over_stock"] = t_turbo / t_stock if t_stock > 0 else None
        print(
            f"    multipliers: oxl/turbo={row['oxl_over_turbo']:.1f}x  "
            f"stock/turbo={row['stock_over_turbo']:.2f}x  "
            f"(turbo carries full surface; stock is values-only)",
            flush=True,
        )
        results["fixtures"][name] = row

    out = REPORTS / "bench_production_raw.json"
    REPORTS.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(results, indent=2), encoding="utf-8")
    print(f"\nWrote {out}")

    print("\n" + "=" * 72)
    print("SUMMARY")
    print(
        f"{'fixture':12s}  {'cells':>10s}  {'stock_s':>9s}  {'turbo_s':>9s}  "
        f"{'oxl_s':>9s}  {'oxl/turbo':>10s}  {'stock/turbo':>11s}"
    )
    for name in all_names:
        r = results["fixtures"][name]
        print(
            f"{name:12s}  {r['cells']:10,}  {r['stock_s']:9.3f}  {r['turbo_all_s']:9.3f}  "
            f"{r['openpyxl_s']:9.3f}  {r['oxl_over_turbo']:9.1f}x  {r['stock_over_turbo']:10.2f}x"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
