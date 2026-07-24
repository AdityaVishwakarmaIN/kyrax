"""Quick write bench: turbo write vs openpyxl normal / write_only save.

Median of 3 runs. Fixtures: 1M numeric, 1M repetitive strings.
Run from repo root:
  PYTHONUTF8=1 ./.venv/Scripts/python scripts/bench_write.py
"""

from __future__ import annotations

import statistics
import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from kyrax import write_excel_turbo

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell

N_RUNS = 3
N = 1_000_000


def median(xs: list[float]) -> float:
    return statistics.median(xs)


def time_runs(fn, runs: int = N_RUNS) -> float:
    times = []
    for _ in range(runs):
        t0 = time.perf_counter()
        fn()
        times.append(time.perf_counter() - t0)
    return median(times)


def bench_numeric():
    print(f"=== numeric {N} cells ===")
    values = list(range(N))

    def turbo():
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_excel_turbo(path, [{"name": "N", "columns": [values]}])
        finally:
            Path(path).unlink(missing_ok=True)

    def opx_normal():
        wb = Workbook()
        ws = wb.active
        ws.title = "N"
        for i, v in enumerate(values, start=1):
            ws.cell(i, 1, v)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
        finally:
            Path(path).unlink(missing_ok=True)

    def opx_wo():
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("N")
        for v in values:
            ws.append([v])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
        finally:
            Path(path).unlink(missing_ok=True)

    t_turbo = time_runs(turbo)
    print(f"  turbo write:          {t_turbo:.3f} s (median of {N_RUNS})")
    t_norm = time_runs(opx_normal)
    print(f"  openpyxl normal:      {t_norm:.3f} s  ({t_norm / t_turbo:.1f}x turbo)")
    t_wo = time_runs(opx_wo)
    print(f"  openpyxl write_only:  {t_wo:.3f} s  ({t_wo / t_turbo:.1f}x turbo)")
    return {
        "fixture": "numeric1m",
        "turbo": t_turbo,
        "opx_normal": t_norm,
        "opx_wo": t_wo,
    }


def bench_strings_rep():
    print(f"=== strings rep {N} cells (100 unique) ===")
    values = [f"s{i % 100}" for i in range(N)]

    def turbo_inline():
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_excel_turbo(
                path, [{"name": "S", "columns": [values]}], string_mode="inline"
            )
        finally:
            Path(path).unlink(missing_ok=True)

    def turbo_sst():
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_excel_turbo(
                path, [{"name": "S", "columns": [values]}], string_mode="sst"
            )
        finally:
            Path(path).unlink(missing_ok=True)

    def opx_normal():
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for i, v in enumerate(values, start=1):
            ws.cell(i, 1, v)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
        finally:
            Path(path).unlink(missing_ok=True)

    def opx_wo():
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("S")
        for v in values:
            ws.append([v])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
        finally:
            Path(path).unlink(missing_ok=True)

    t_inline = time_runs(turbo_inline)
    print(f"  turbo inline:         {t_inline:.3f} s")
    t_sst = time_runs(turbo_sst)
    print(f"  turbo sst:            {t_sst:.3f} s")
    t_norm = time_runs(opx_normal)
    print(f"  openpyxl normal:      {t_norm:.3f} s  (vs inline {t_norm / t_inline:.1f}x)")
    t_wo = time_runs(opx_wo)
    print(f"  openpyxl write_only:  {t_wo:.3f} s  (vs inline {t_wo / t_inline:.1f}x)")
    return {
        "fixture": "strings1m_rep",
        "turbo_inline": t_inline,
        "turbo_sst": t_sst,
        "opx_normal": t_norm,
        "opx_wo": t_wo,
    }


def main():
    print("bench_write.py — median of", N_RUNS, "runs")
    print("N =", N)
    r1 = bench_numeric()
    r2 = bench_strings_rep()
    print()
    print("SUMMARY")
    print(
        f"numeric1m: turbo {r1['turbo']:.3f}s | "
        f"opx normal {r1['opx_normal']:.3f}s ({r1['opx_normal']/r1['turbo']:.0f}x) | "
        f"opx wo {r1['opx_wo']:.3f}s ({r1['opx_wo']/r1['turbo']:.1f}x)"
    )
    print(
        f"strings_rep inline: turbo {r2['turbo_inline']:.3f}s | "
        f"opx normal {r2['opx_normal']:.3f}s ({r2['opx_normal']/r2['turbo_inline']:.0f}x) | "
        f"opx wo {r2['opx_wo']:.3f}s ({r2['opx_wo']/r2['turbo_inline']:.1f}x)"
    )
    print(
        f"strings_rep sst:    turbo {r2['turbo_sst']:.3f}s | "
        f"vs normal {r2['opx_normal']/r2['turbo_sst']:.0f}x"
    )


if __name__ == "__main__":
    main()
